#!/usr/bin/env python3
"""
Умный помощник / Smart Assistant — Excel/CSV + Claude.
Запуск: python3 app.py
"""

from __future__ import annotations

import csv
import os
import platform
import tempfile
import re
import sys
import threading
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from io import StringIO

import pandas as pd
from anthropic import Anthropic

from complaints_registry import enrich_complaints_draft, mark_complaints_sent
from history_store import (
    HistoryRecord,
    delete_history_entry,
    list_history_entries,
    load_history_dataframes,
    load_history_entry,
    save_history_entry,
)
import user_data
from user_data import APP_DIR, ensure_user_data_ready

ensure_user_data_ready()
from PyQt6.QtCore import (
    QEasingCurve,
    QEvent,
    QObject,
    QPoint,
    QPointF,
    QPropertyAnimation,
    QRectF,
    QSize,
    Qt,
    QTimer,
    QUrl,
    pyqtSignal,
)
from PyQt6.QtGui import (
    QBrush,
    QColor,
    QDesktopServices,
    QFont,
    QFontMetrics,
    QIcon,
    QKeyEvent,
    QMouseEvent,
    QPainter,
    QPalette,
    QPen,
    QPixmap,
    QKeySequence,
    QResizeEvent,
    QShortcut,
)
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QMenu,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

# ── Темы ─────────────────────────────────────
THEME_PALETTES: dict[str, dict[str, str]] = {
    "light": {
        "BG_APP": "#f3f2f1",
        "BG_CARD": "#ffffff",
        "SIDEBAR_BG": "#faf9f8",
        "SIDEBAR_TEXT": "#323130",
        "SIDEBAR_MUTED": "#605e5c",
        "SIDEBAR_ACCENT": "#0078d4",
        "SIDEBAR_BORDER": "#edebe9",
        "BG_HOVER": "#edebe9",
        "BG_ACCENT": "#eff6fc",
        "BORDER": "#edebe9",
        "BORDER_FOCUS": "#0078d4",
        "BLUE": "#0078d4",
        "BLUE_HOVER": "#106ebe",
        "GREEN": "#107c10",
        "GREEN_HOVER": "#0b5a0b",
        "WARN_BG": "#fff4ce",
        "WARN_BORDER": "#fde047",
        "WARN_TEXT": "#7a5c00",
        "INFO_BG": "#faf9f8",
        "INFO_BORDER": "#edebe9",
        "INFO_TEXT": "#605e5c",
        "TEXT": "#323130",
        "TEXT_SEC": "#605e5c",
        "TEXT_MUTED": "#8a8886",
        "ERROR": "#a4262c",
        "STATS_BG": "#faf9f8",
        "RESULT_CANVAS": "#f3f2f1",
        "RESULT_PANEL": "#ffffff",
        "RESULT_READY_BG": "#eff6fc",
        "RESULT_READY_BORDER": "#0078d4",
        "FILTER_BAR_BG": "#faf9f8",
        "TPL_BADGE_BG": "#eff6fc",
        "TPL_BADGE_BORDER": "#c7e0f4",
        "TPL_BADGE_TEXT": "#0078d4",
        "CHIP_BG": "#faf9f8",
        "CHIP_BORDER": "#c7e0f4",
        "DETACH_BG": "#ffffff",
        "DETACH_HOVER": "#fde7e9",
        "DISABLED_BTN": "#c8c6c4",
        "DISABLED_TEXT": "#ffffff",
    },
    "dark": {
        "BG_APP": "#1b1a19",
        "BG_CARD": "#292827",
        "SIDEBAR_BG": "#201f1e",
        "SIDEBAR_TEXT": "#f3f2f1",
        "SIDEBAR_MUTED": "#a19f9d",
        "SIDEBAR_ACCENT": "#479ef5",
        "SIDEBAR_BORDER": "#3b3a39",
        "BG_HOVER": "#323130",
        "BG_ACCENT": "#243a5e",
        "BORDER": "#3b3a39",
        "BORDER_FOCUS": "#479ef5",
        "BLUE": "#479ef5",
        "BLUE_HOVER": "#62abf5",
        "GREEN": "#6ccb5f",
        "GREEN_HOVER": "#92d873",
        "WARN_BG": "#433519",
        "WARN_BORDER": "#8a7119",
        "WARN_TEXT": "#fde047",
        "INFO_BG": "#292827",
        "INFO_BORDER": "#3b3a39",
        "INFO_TEXT": "#a19f9d",
        "TEXT": "#f3f2f1",
        "TEXT_SEC": "#d2d0ce",
        "TEXT_MUTED": "#a19f9d",
        "ERROR": "#f1707b",
        "STATS_BG": "#252423",
        "RESULT_CANVAS": "#1b1a19",
        "RESULT_PANEL": "#292827",
        "RESULT_READY_BG": "#1a2f4a",
        "RESULT_READY_BORDER": "#479ef5",
        "FILTER_BAR_BG": "#252423",
        "TPL_BADGE_BG": "#243a5e",
        "TPL_BADGE_BORDER": "#479ef5",
        "TPL_BADGE_TEXT": "#62abf5",
        "CHIP_BG": "#252423",
        "CHIP_BORDER": "#479ef5",
        "DETACH_BG": "#292827",
        "DETACH_HOVER": "#4a2028",
        "DISABLED_BTN": "#484644",
        "DISABLED_TEXT": "#a19f9d",
    },
}

APP_VERSION = "1.1.0"
APP_ICON_PATH = os.path.join(APP_DIR, "assets", "app_icon.png")
RULES_TASK_MARKER = "@@company_rules@@"
TABLE_EXTENSIONS = (".xlsx", ".csv", ".txt", ".tsv")
RULES_EXTENSIONS = (".txt", ".md", ".csv", ".xlsx", ".xlsm")
RULES_TABLE_FOOTER = (
    "\n\nВАЖНО: при классификации используй ТОЛЬКО колонки и логику из этих правил. "
    "Не добавляй колонки, которых нет в правилах компании."
)

SYSTEM_PROMPT = (
    "You are a deterministic CSV table processor. "
    "Apply the command from <user_task> to data from <source_csv>.\n\n"
    "STABILITY RULES (mandatory):\n"
    "1. First row is column headers. Never delete or rename unless explicitly asked.\n"
    "2. Do not add or remove columns unless specified.\n"
    "3. Never invent data. If unclear, keep the original value.\n"
    "4. Preserve column order unless the task requires otherwise.\n"
    "5. Keep numbers as numbers, dates as dates.\n"
    "6. Only do what the task asks. No extra transformations.\n"
    "7. When sorting or deduplicating, do not lose rows that should remain.\n"
    "8. UTF-8 encoding. Comma separator. Double-quote any field that contains "
    "commas, semicolons, quotes, or line breaks.\n"
    "9. Every row must have the same number of columns as the header.\n\n"
    "OUTPUT: Valid RFC-4180 CSV only. No markdown fences, no comments, "
    "no explanations before or after the table."
)

RULES_SYSTEM_APPEND = (
    "\n\nWHEN <company_rules> IS PRESENT (complaint routing):\n"
    "1. <company_rules> is the ONLY source for output column names, allowed departments, "
    "categories, priorities, composite-complaint logic, and fallback behavior.\n"
    "2. Do NOT add columns that are not listed in <company_rules>.\n"
    "3. Do NOT use column names from generic examples — only names from <company_rules>.\n"
    "4. For EVERY row: read the FULL complaint text before deciding.\n"
    "5. Never route by a single keyword; use surrounding context.\n"
    "6. Composite complaints: follow composite rules in <company_rules>.\n"
    "7. If ambiguous — use the fallback defined in <company_rules>.\n"
    "8. Leave empty any columns that <company_rules> says the application will fill.\n"
    "9. <context> describes the input table — use it together with <company_rules>."
)

_COMPLAINTS_TASK_BODIES: dict[str, str] = {
    "ru": (
        "ЗАДАЧА: классифицировать каждую жалобу строго по <company_rules>.\n\n"
        "<company_rules> — единственный источник: какие колонки добавить, как их назвать, "
        "разрешённые отделы, категории, приоритеты, составные жалобы и неясные случаи.\n"
        "НЕ добавляй колонки, которых нет в <company_rules>. "
        "НЕ подставляй названия из шаблонов или примеров.\n\n"
        "АЛГОРИТМ (для КАЖДОЙ строки, без пропусков):\n"
        "1. Найди колонку с текстом жалобы — как указано в <company_rules> и <context>.\n"
        "2. Прочитай текст жалобы ЦЕЛИКОМ — не по одному слову.\n"
        "3. Сопоставь смысл с категориями и отделами из <company_rules>.\n"
        "4. Назначь главную категорию и отдел — только из разрешённого списка в <company_rules>.\n"
        "5. Несколько проблем в одной жалобе — следуй разделу о составных жалобах в <company_rules>.\n"
        "6. Неясный случай — fallback из <company_rules>.\n"
        "7. Заполни колонки результата из списка в <company_rules>; "
        "колонки, которые правила просят оставить пустыми, — не заполняй.\n\n"
        "ЗАПРЕЩЕНО: колонки вне списка из <company_rules>; отделы вне разрешённого списка; "
        "удалять или переименовывать исходные колонки; менять исходные ячейки.\n"
        "Верни только валидный CSV; поля с запятыми — в двойных кавычках."
    ),
    "en": (
        "TASK: classify each complaint strictly per <company_rules>.\n\n"
        "<company_rules> is the ONLY source for output column names, allowed departments, "
        "categories, priorities, composite complaints, and fallback behavior.\n"
        "Do NOT add columns not listed in <company_rules>. "
        "Do NOT use column names from generic templates or examples.\n\n"
        "ALGORITHM (for EVERY row, no skips):\n"
        "1. Find the complaint text column per <company_rules> and <context>.\n"
        "2. Read the FULL complaint text — not a single keyword.\n"
        "3. Match meaning to categories and departments in <company_rules>.\n"
        "4. Assign primary category and department — only from the allowed list in <company_rules>.\n"
        "5. Multiple issues — follow the composite-complaint section in <company_rules>.\n"
        "6. Ambiguous case — use fallback from <company_rules>.\n"
        "7. Fill result columns listed in <company_rules>; leave empty any columns "
        "the rules say the application will fill.\n\n"
        "FORBIDDEN: columns outside <company_rules>; departments outside the allowed list; "
        "deleting or renaming source columns; changing source cell values.\n"
        "Output valid CSV only; quote fields that contain commas."
    ),
    "de": (
        "AUFGABE: Jede Beschwerde strikt nach <company_rules> einordnen.\n\n"
        "<company_rules> ist die EINZIGE Quelle für Ausgabespalten, erlaubte Abteilungen, "
        "Kategorien, Prioritäten, Mehrfachbeschwerden und Fallback-Verhalten.\n"
        "KEINE Spalten hinzufügen, die nicht in <company_rules> stehen.\n\n"
        "ALGORITHMUS (für JEDE Zeile):\n"
        "1. Beschwerdetext-Spalte laut <company_rules> und <context> finden.\n"
        "2. Den GESAMTEN Beschwerdetext lesen.\n"
        "3. Bedeutung mit Kategorien und Abteilungen in <company_rules> abgleichen.\n"
        "4. Hauptkategorie und Abteilung — nur aus der erlaubten Liste in <company_rules>.\n"
        "5. Mehrere Probleme — Abschnitt zu Mehrfachbeschwerden in <company_rules> befolgen.\n"
        "6. Unklar — Fallback aus <company_rules>.\n"
        "7. Ergebnisspalten aus <company_rules> füllen; vom Programm vorgesehene Felder leer lassen.\n\n"
        "VERBOTEN: Spalten außerhalb <company_rules>; Abteilungen außerhalb der Liste; "
        "Quellspalten löschen oder umbenennen.\n"
        "Nur gültiges CSV ausgeben."
    ),
}


def complaints_task_text(lang: str) -> str:
    body = _COMPLAINTS_TASK_BODIES.get(lang, _COMPLAINTS_TASK_BODIES["ru"])
    return f"{RULES_TASK_MARKER}\n{body}"

COMPLAINT_TEXT_NAMES = (
    "Текст_жалобы", "Текст", "Жалоба", "Обращение", "Описание",
    "Text", "Complaint", "Beschreibung", "Beschwerde",
)

def apply_platform_theme(app: QApplication, theme: str) -> None:
    """На macOS нативная тема ломает фон у текста — задаём Fusion + палитру явно."""
    if platform.system() == "Darwin":
        app.setStyle("Fusion")

    p = THEME_PALETTES.get(theme, THEME_PALETTES["light"])
    pal = QPalette()
    pal.setColor(QPalette.ColorRole.Window, QColor(p["BG_APP"]))
    pal.setColor(QPalette.ColorRole.WindowText, QColor(p["TEXT"]))
    pal.setColor(QPalette.ColorRole.Base, QColor(p["BG_CARD"]))
    pal.setColor(QPalette.ColorRole.AlternateBase, QColor(p["BG_HOVER"]))
    pal.setColor(QPalette.ColorRole.Text, QColor(p["TEXT"]))
    pal.setColor(QPalette.ColorRole.Button, QColor(p["BG_CARD"]))
    pal.setColor(QPalette.ColorRole.ButtonText, QColor(p["TEXT"]))
    pal.setColor(QPalette.ColorRole.ToolTipBase, QColor(p["BG_CARD"]))
    pal.setColor(QPalette.ColorRole.ToolTipText, QColor(p["TEXT"]))
    pal.setColor(QPalette.ColorRole.Highlight, QColor(p["BLUE"]))
    pal.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
    pal.setColor(QPalette.ColorRole.PlaceholderText, QColor(p["TEXT_MUTED"]))
    app.setPalette(pal)


def build_stylesheet(palette: dict[str, str]) -> str:
    p = palette
    info_bg = p.get("INFO_BG", p["STATS_BG"])
    info_border = p.get("INFO_BORDER", p["BORDER"])
    info_text = p.get("INFO_TEXT", p["TEXT_MUTED"])
    return f"""
QMainWindow, QWidget {{
    background-color: {p["BG_APP"]};
    color: {p["TEXT"]};
    font-family: "Segoe UI", "Helvetica Neue", Arial, sans-serif;
    font-size: 13px;
}}
QFrame#card, QFrame#card_result {{
    background-color: {p["BG_CARD"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 4px;
}}
QLabel {{
    background-color: transparent;
    color: {p["TEXT"]};
}}
QLabel#app_title {{
    font-size: 22px; font-weight: 700; color: {p["TEXT"]}; background: transparent;
    letter-spacing: -0.3px;
}}
QLabel#app_sub {{ font-size: 13px; color: {p["TEXT_MUTED"]}; line-height: 1.4; }}
QLabel#app_icon {{
    background: transparent;
    border: none;
    padding: 0;
}}
QLabel#card_title {{
    font-size: 12px; font-weight: 600; color: {p["TEXT"]};
    letter-spacing: 0.4px; padding-bottom: 2px;
}}
QFrame#command_bar {{
    background: {p["BG_CARD"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 4px;
}}
QLabel#command_title {{
    font-size: 16px; font-weight: 600; color: {p["TEXT"]};
    background: transparent; letter-spacing: -0.1px;
}}
QLabel#command_sub {{
    font-size: 12px; color: {p["TEXT_MUTED"]}; background: transparent;
}}
QLabel#version_badge {{
    font-size: 11px; font-weight: 600; color: {p["TEXT_SEC"]};
    background: {p["STATS_BG"]}; border: 1px solid {p["BORDER"]};
    border-radius: 3px; padding: 4px 10px;
}}
QLabel#step_title {{
    font-size: 12px; font-weight: 600; color: {p["TEXT_MUTED"]};
    letter-spacing: 0.6px; padding: 0;
}}
QLabel#step_badge {{
    background: {p["BG_ACCENT"]}; color: {p["BLUE"]};
    border-radius: 10px;
    min-width: 20px; max-width: 20px;
    min-height: 20px; max-height: 20px;
    font-size: 11px; font-weight: 700;
    qproperty-alignment: AlignCenter;
}}
QFrame#result_empty_panel {{
    background: transparent; border: none;
}}
QLabel#result_empty_hint {{
    color: {p["TEXT_MUTED"]}; font-size: 15px; font-weight: 400;
    background: transparent; padding: 8px 24px; line-height: 1.55;
}}
QLabel#result_empty_title {{
    color: {p["TEXT"]}; font-size: 16px; font-weight: 600;
    background: transparent; padding: 0 24px 6px 24px;
}}
QFrame#app_header {{
    background: {p["BG_CARD"]};
    border: none;
    border-bottom: 1px solid {p["BORDER"]};
}}
QFrame#nav_sidebar {{
    background: {p.get("SIDEBAR_BG", p["BG_CARD"])};
    border: none;
    border-right: 1px solid {p.get("SIDEBAR_BORDER", p["BORDER"])};
}}
QLabel#sidebar_title {{
    font-size: 14px; font-weight: 600; color: {p.get("SIDEBAR_TEXT", p["TEXT"])};
    background: transparent; letter-spacing: -0.1px;
}}
QLabel#sidebar_sub {{
    font-size: 11px; color: {p.get("SIDEBAR_MUTED", p["TEXT_MUTED"])};
    background: transparent; line-height: 1.35;
}}
QLabel#sidebar_section {{
    font-size: 10px; font-weight: 600; color: {p.get("SIDEBAR_MUTED", p["TEXT_MUTED"])};
    background: transparent; letter-spacing: 0.8px;
}}
QFrame#wf_item {{
    background: transparent; border: none; border-radius: 2px;
    border-left: 3px solid transparent;
}}
QFrame#wf_item:hover {{
    background: {p.get("BG_HOVER", p["STATS_BG"])};
}}
QFrame#wf_item[state="active"] {{
    background: {p["BG_ACCENT"]};
    border-left: 3px solid {p["BLUE"]};
}}
QFrame#wf_item[state="done"] {{
    background: transparent;
    border-left: 3px solid {p["GREEN"]};
}}
QLabel#wf_dot {{
    background: {p["STATS_BG"]};
    color: {p.get("SIDEBAR_MUTED", p["TEXT_MUTED"])};
    border: 1px solid {p["BORDER"]};
    border-radius: 10px;
    min-width: 20px; max-width: 20px;
    min-height: 20px; max-height: 20px;
    font-size: 10px; font-weight: 600;
    qproperty-alignment: AlignCenter;
}}
QFrame#wf_item[state="active"] QLabel#wf_dot {{
    background: {p["BLUE"]}; color: white; border-color: {p["BLUE"]};
}}
QFrame#wf_item[state="done"] QLabel#wf_dot {{
    background: {p["GREEN"]}; color: white; border-color: {p["GREEN"]};
}}
QLabel#wf_label {{
    color: {p.get("SIDEBAR_MUTED", p["TEXT_MUTED"])};
    font-size: 12px; font-weight: 400; background: transparent;
}}
QFrame#wf_item[state="active"] QLabel#wf_label,
QFrame#wf_item[state="done"] QLabel#wf_label {{
    color: {p.get("SIDEBAR_TEXT", p["TEXT"])};
    font-weight: 600;
}}
QWidget#main_stage {{
    background: {p["BG_APP"]};
}}
QFrame#stage_panel {{
    background-color: {p["BG_CARD"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 4px;
}}
QFrame#panel_sep {{
    background: {p["BORDER"]}; border: none; max-height: 1px;
}}
QLabel#panel_label {{
    color: {p["TEXT_SEC"]}; font-size: 10px; font-weight: 600;
    letter-spacing: 0.6px; background: transparent;
}}
QWidget#input_column {{
    background: transparent;
}}
QFrame#card_result {{
    background-color: {p.get("RESULT_PANEL", p["BG_CARD"])};
    border: 1px solid {p["BORDER"]};
    border-radius: 4px;
}}
QFrame#card_result[ready="true"] {{
    border-color: {p["RESULT_READY_BORDER"]};
    background-color: {p.get("RESULT_READY_BG", p["BG_CARD"])};
}}
QFrame#nav_sidebar QLabel#lang_label {{
    color: {p.get("SIDEBAR_MUTED", p["TEXT_MUTED"])};
    font-size: 11px;
}}
QFrame#nav_sidebar QComboBox {{
    background: {p["BG_CARD"]};
    color: {p.get("SIDEBAR_TEXT", p["TEXT"])};
    border: 1px solid {p["BORDER"]};
    border-radius: 3px;
    font-size: 12px;
}}
QFrame#nav_sidebar QPushButton#info_icon {{
    background: {p["BG_CARD"]}; color: {p.get("SIDEBAR_MUTED", p["TEXT_SEC"])};
    border: 1px solid {p["BORDER"]};
}}
QFrame#nav_sidebar QPushButton#info_icon:hover {{
    background: {p["BG_ACCENT"]};
    border-color: {p["BLUE"]};
    color: {p["BLUE"]};
}}
QFrame#nav_sidebar QPushButton#info_icon[active="true"] {{
    background: {p["BG_ACCENT"]};
    border-color: {p["BLUE"]};
    color: {p["BLUE"]};
}}
QWidget#work_zone {{
    background: {p["BG_APP"]};
}}
QFrame#result_canvas {{
    background: {p.get("RESULT_CANVAS", p["STATS_BG"])};
    border: none;
    border-top: 1px solid {p["BORDER"]};
}}
QFrame#card_divider {{
    background: transparent; border: none;
    border-top: 1px solid {p["BORDER"]};
    min-height: 1px; max-height: 1px;
}}
QFrame#result_hdr {{
    background: {p["BG_CARD"]}; border: none;
    border-bottom: 1px solid {p["BORDER"]};
}}
QLabel#result_hdr_sub {{
    color: {p["TEXT_MUTED"]}; font-size: 12px; font-weight: 400;
    background: transparent;
}}
QFrame#result_body {{
    background: transparent; border: none;
}}
QFrame#result_ready_panel {{
    background: {p.get("STATS_BG", p["BG_ACCENT"])};
    border: 1px solid {p["BORDER"]};
    border-radius: 4px;
}}
QLabel#result_ready_title {{
    color: {p["TEXT"]}; font-size: 14px; font-weight: 600;
    background: transparent;
}}
QLabel#result_ready_hint {{
    color: {p["TEXT_SEC"]}; font-size: 12px; background: transparent;
}}
QLabel#result_alert {{
    color: {p["WARN_TEXT"]}; font-size: 12px; font-weight: 500;
    background: {p["WARN_BG"]}; border: 1px solid {p["WARN_BORDER"]};
    border-radius: 4px; padding: 10px 12px;
}}
QPushButton#info_nav_btn {{
    background: {p["BG_CARD"]};
    color: {p.get("SIDEBAR_TEXT", p["TEXT"])};
    border: 1px solid {p["BORDER"]}; border-radius: 3px;
    padding: 8px 10px; font-size: 12px; font-weight: 500;
    text-align: left; min-height: 32px;
}}
QPushButton#info_nav_btn:hover {{
    background: {p["BG_ACCENT"]}; border-color: {p["BLUE"]}; color: {p["BLUE"]};
}}
QPushButton#info_nav_btn[active="true"] {{
    background: {p["BG_ACCENT"]}; border-color: {p["BLUE"]};
    color: {p["BLUE"]}; font-weight: 600;
}}
QProgressBar#run_progress {{
    background: {p["STATS_BG"]}; border: none; border-radius: 2px;
    max-height: 4px; min-height: 4px;
}}
QProgressBar#run_progress::chunk {{
    background: {p["BLUE"]}; border-radius: 2px;
}}
QFrame#filter_bar {{
    background: {p.get("FILTER_BAR_BG", p["BG_CARD"])};
    border: 1px solid {p["BORDER"]};
    border-radius: 10px;
}}
QLabel#filter_bar_label {{
    color: {p["TEXT"]}; font-size: 12px; font-weight: 600;
    background: transparent;
}}
QLabel#filter_count {{
    color: {p["TEXT_SEC"]}; font-size: 12px; font-weight: 500;
    background: {p.get("BG_HOVER", p["STATS_BG"])};
    border-radius: 6px; padding: 4px 10px;
}}
QTableWidget#draft_table {{
    background: {p["BG_CARD"]}; color: {p["TEXT"]};
    border: 1px solid {p["BORDER"]}; border-radius: 8px;
    gridline-color: {p["BORDER"]}; font-size: 12px;
}}
QHeaderView::section {{
    background: {p.get("FILTER_BAR_BG", p["STATS_BG"])};
    color: {p["TEXT_SEC"]}; border: none;
    border-bottom: 1px solid {p["BORDER"]};
    padding: 6px 8px; font-size: 12px; font-weight: 600;
}}
QLabel#lang_label {{ color: {p["TEXT_SEC"]}; font-size: 12px; }}
QLabel#tpl_label {{
    background: {p["TPL_BADGE_BG"]};
    color: {p["TPL_BADGE_TEXT"]};
    border: 1px solid {p["TPL_BADGE_BORDER"]};
    border-radius: 8px;
    padding: 10px 14px;
    font-size: 13px;
    font-weight: 600;
    min-width: 100px;
}}
QLabel#inline_label {{
    background: transparent; border: none;
    color: {p["TEXT_SEC"]}; font-size: 13px; font-weight: 500;
    padding: 0;
}}
QLabel#rules_hint {{
    color: {p["WARN_TEXT"]}; font-size: 12px;
}}
QLabel#section_label {{
    color: {p["TEXT_SEC"]}; font-size: 13px; font-weight: 500;
    margin-top: 4px;
}}
QComboBox, QComboBox#lang_combo, QComboBox#theme_combo {{
    background: {p["BG_CARD"]}; color: {p["TEXT"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 3px; padding: 6px 10px; min-width: 110px;
    font-size: 12px;
}}
QComboBox:hover {{
    border-color: {p["BLUE"]};
}}
QComboBox::drop-down, QComboBox#lang_combo::drop-down, QComboBox#theme_combo::drop-down {{
    border: none; width: 22px;
}}
QComboBox QAbstractItemView {{
    background: {p["BG_CARD"]}; color: {p["TEXT"]};
    border: 1px solid {p["BORDER"]};
    selection-background-color: {p["BG_ACCENT"]};
    selection-color: {p["BLUE"]};
    padding: 4px; outline: none;
}}
QPushButton#tpl_pick {{
    background: {p["BG_CARD"]}; color: {p["TEXT"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 3px; padding: 7px 12px;
    min-height: 20px; max-height: 32px;
    font-size: 12px; text-align: left;
}}
QPushButton#tpl_pick:hover {{
    border-color: {p["BLUE"]}; color: {p["BLUE"]}; background: {p["BG_ACCENT"]};
}}
QPushButton#tpl_pick[selected="true"] {{
    background: {p["BG_ACCENT"]}; color: {p["BLUE"]};
    border-color: {p["BLUE"]}; font-weight: 600;
}}
QLineEdit {{
    background-color: {p["BG_CARD"]}; color: {p["TEXT"]};
    border: 1px solid {p["BORDER"]}; border-radius: 3px; padding: 8px 10px;
    font-size: 12px;
    selection-background-color: {p["BLUE"]}; selection-color: white;
}}
QLineEdit:focus {{ border: 1px solid {p["BORDER_FOCUS"]}; }}
QFrame#setup_strip {{
    background-color: {p["BG_CARD"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 12px;
}}
QFrame#setup_sep {{
    background-color: {p["BORDER"]};
    border: none; max-width: 1px;
}}
QFrame#task_dock {{
    background-color: {p["BG_CARD"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 12px;
}}
QFrame#task_composer {{
    background-color: {p["STATS_BG"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 3px;
}}
QFrame#task_composer[focused="true"] {{
    border: 1px solid {p["BORDER_FOCUS"]};
}}
QTextEdit#task_input {{
    background: transparent; color: {p["TEXT"]};
    border: none; border-radius: 0;
    padding: 8px 10px; font-size: 13px; line-height: 1.4;
    selection-background-color: {p["BLUE"]}; selection-color: white;
}}
QWidget#task_send_row {{
    background: transparent;
}}
QLabel#task_progress {{
    color: {p["BLUE"]}; font-size: 12px; font-weight: 500;
    background: transparent; padding: 4px 2px 0 4px;
}}
QPushButton#task_send {{
    background-color: {p["BLUE"]}; color: white; border: none;
    border-radius: 3px; padding: 0 16px;
    min-width: 148px; max-width: 220px;
    min-height: 32px; max-height: 32px;
    font-size: 12px; font-weight: 600;
}}
QPushButton#task_send:hover {{ background-color: {p["BLUE_HOVER"]}; }}
QPushButton#task_send:disabled {{
    background-color: {p["DISABLED_BTN"]}; color: {p["DISABLED_TEXT"]};
}}
QTextEdit#output {{
    background-color: {p["BG_CARD"]}; color: {p["TEXT"]};
    border: none; border-radius: 0;
    padding: 14px 16px;
    selection-background-color: {p["BLUE"]}; selection-color: white;
}}
QTextEdit#output:focus {{
    border: 1px solid {p["BORDER_FOCUS"]};
}}
QTextEdit#output {{
    font-family: "Menlo", "Consolas", monospace; font-size: 12px;
}}
QTextEdit#task_input QScrollBar:horizontal,
QTextEdit#output QScrollBar:horizontal {{
    height: 0px; max-height: 0px; background: transparent;
}}
QTextEdit#task_input QScrollBar::handle:horizontal,
QTextEdit#output QScrollBar::handle:horizontal {{
    background: transparent; min-width: 0px;
}}
QFrame#disclaimer {{
    background-color: {info_bg}; border: 1px solid {info_border};
    border-radius: 8px; min-height: 0px;
}}
QLabel#disclaimer_text {{
    color: {info_text}; font-size: 12px; font-weight: 400;
    background: transparent; padding: 2px 0; line-height: 1.4;
}}
QPushButton#primary {{
    background-color: {p["BLUE"]}; color: white; border: none;
    border-radius: 3px; padding: 8px 16px; font-weight: 600; font-size: 12px;
}}
QPushButton#primary:hover {{ background-color: {p["BLUE_HOVER"]}; }}
QPushButton#primary:disabled {{
    background-color: {p["DISABLED_BTN"]}; color: {p["DISABLED_TEXT"]};
}}
QPushButton#action {{
    background-color: {p["BLUE"]}; color: white; border: none;
    border-radius: 10px; padding: 0 24px; font-size: 14px; font-weight: 600;
    min-height: 42px; max-height: 42px;
}}
QPushButton#action:hover {{ background-color: {p["BLUE_HOVER"]}; }}
QPushButton#action:disabled {{
    background-color: {p["DISABLED_BTN"]}; color: {p["DISABLED_TEXT"]};
}}
QPushButton#secondary {{
    background-color: {p["BG_CARD"]}; color: {p["TEXT_SEC"]};
    border: 1px solid {p["BORDER"]}; border-radius: 3px;
    padding: 6px 12px; font-size: 12px; min-height: 30px; max-height: 30px;
}}
QPushButton#secondary:hover {{
    background-color: {p["BG_HOVER"]}; border-color: {p["BLUE"]}; color: {p["BLUE"]};
}}
QPushButton#save {{
    background-color: {p["BLUE"]}; color: white;
    border: none; border-radius: 8px; font-weight: 600;
    padding: 7px 12px; font-size: 13px; min-height: 34px; max-height: 34px;
}}
QPushButton#save:hover {{ background-color: {p["BLUE_HOVER"]}; }}
QPushButton#save:disabled {{
    background-color: {p["DISABLED_BTN"]}; color: {p["DISABLED_TEXT"]};
}}
QPushButton#save_outline {{
    background-color: {p["BG_CARD"]}; color: {p["BLUE"]};
    border: 1px solid {p["BLUE"]}; border-radius: 8px; font-weight: 600;
    padding: 7px 12px; font-size: 13px; min-height: 34px; max-height: 34px;
}}
QPushButton#save_outline:hover {{ background-color: {p["BG_ACCENT"]}; }}
QPushButton#save_outline:disabled {{
    color: {p["TEXT_MUTED"]}; border-color: {p["BORDER"]}; background: {p["BG_CARD"]};
}}
QPushButton#menu_btn {{
    background-color: {p["BG_CARD"]}; color: {p["TEXT_SEC"]};
    border: 1px solid {p["BORDER"]}; border-radius: 8px; padding: 7px 10px;
    font-size: 13px; font-weight: 500;
}}
QPushButton#menu_btn:hover {{
    border-color: {p["BLUE"]}; color: {p["BLUE"]}; background: {p["BG_ACCENT"]};
}}
QPushButton#menu_btn:disabled {{ color: {p["TEXT_MUTED"]}; border-color: {p["BORDER"]}; }}
QLabel#file_empty {{ color: {p["TEXT_MUTED"]}; font-size: 13px; }}
QFrame#file_chip QLabel#file_chip_name {{
    background: transparent; border: none;
    color: {p["BLUE"]}; font-size: 13px; font-weight: 600;
    padding: 0; margin: 0;
}}
QLabel#stats_title {{
    color: {p["TEXT"]}; font-size: 14px; font-weight: 600;
    background: transparent; padding-bottom: 2px;
}}
QFrame#file_chip {{
    background: {p["CHIP_BG"]}; border: 1px solid {p["CHIP_BORDER"]};
    border-radius: 8px; min-height: 40px; max-height: 40px;
}}
QLabel#detach {{
    background: {p["DETACH_BG"]}; color: {p["ERROR"]};
    border: 1px solid {p["ERROR"]}; border-radius: 3px;
    min-width: 24px; max-width: 24px;
    min-height: 24px; max-height: 24px;
    padding: 0; margin: 0;
}}
QLabel#detach:hover {{ background: {p["DETACH_HOVER"]}; }}
QLabel#status_ok {{
    color: {p["GREEN"]}; font-size: 11px; font-weight: 600;
}}
QLabel#status_pending {{
    color: {p["TEXT_MUTED"]}; font-size: 11px; font-weight: 500;
}}
QLabel#run_hint {{
    color: {p["TEXT_MUTED"]}; font-size: 11px; font-weight: 500;
    background: transparent;
}}
QLabel#file_meta {{
    color: {p["TEXT_SEC"]}; font-size: 11px; background: transparent;
}}
QFrame#task_composer[locked="true"] {{
    background-color: {p["STATS_BG"]};
    border-color: {p["BORDER"]};
}}
QLabel#progress {{ color: {p["BLUE"]}; font-size: 13px; font-weight: 500; }}
QPushButton#attach_rules {{
    background: {p["STATS_BG"]}; color: {p["TEXT_SEC"]};
    border: 1px solid {p["BORDER"]}; border-radius: 8px;
    padding: 6px 12px; font-size: 12px; font-weight: 500;
    min-height: 32px; max-height: 32px;
}}
QPushButton#toolbar_btn {{
    background: {p["BG_CARD"]}; color: {p["TEXT_SEC"]};
    border: 1px solid {p["BORDER"]}; border-radius: 3px;
    padding: 4px 10px; font-size: 11px; font-weight: 500;
    min-height: 28px; max-height: 28px;
}}
QLabel#result_warn {{
    color: {p["WARN_TEXT"]}; font-size: 12px; font-weight: 500;
    background: {p["WARN_BG"]}; border: 1px solid {p["WARN_BORDER"]};
    border-radius: 8px; padding: 10px 14px;
}}
QFrame#changes_box {{
    background: {p["STATS_BG"]}; border: 1px solid {p["BORDER"]};
    border-radius: 4px;
}}
QLabel#changes_summary {{
    color: {p["TEXT"]}; font-size: 12px; font-weight: 600; background: transparent;
}}
QLabel#save_hint {{
    color: {p["TEXT_MUTED"]}; font-size: 12px; background: transparent;
}}
QFrame#result_panel {{
    background-color: {p.get("RESULT_PANEL", p["BG_CARD"])};
    border: 1px solid {p["BORDER"]};
    border-radius: 14px;
}}
QFrame#result_panel[ready="true"] {{
    border: 2px solid {p.get("RESULT_READY_BORDER", p["GREEN"])};
    background-color: {p.get("RESULT_READY_BG", p["BG_ACCENT"])};
}}
QFrame#result_ready {{
    background: transparent; border: none;
}}
QLabel#result_empty {{
    color: {p["TEXT_MUTED"]}; font-size: 14px; background: transparent;
    padding: 36px 28px;
}}
QLabel#result_ready_title {{
    color: {p["TEXT"]}; font-size: 17px; font-weight: 700; background: transparent;
}}
QLabel#result_ready_hint {{
    color: {p["TEXT_SEC"]}; font-size: 13px; background: transparent;
    line-height: 1.45;
}}
QLabel#result_summary {{
    color: {p["TEXT"]}; font-size: 13px; font-weight: 600;
    background: {p.get("BG_HOVER", p["STATS_BG"])};
    border-radius: 8px; padding: 10px 12px;
}}
QFrame#info_rail {{
    background-color: {p["BG_CARD"]};
    border: none;
    border-right: 1px solid {p["BORDER"]};
}}
QPushButton#info_icon {{
    background: transparent; color: {p["TEXT_SEC"]};
    border: 1px solid transparent; border-radius: 11px;
    padding: 0; min-width: 44px; max-width: 44px;
    min-height: 44px; max-height: 44px;
}}
QPushButton#info_icon:hover {{
    background: {p["BG_HOVER"]}; border-color: {p["BORDER"]};
}}
QPushButton#info_icon:pressed {{
    background: {p["BG_ACCENT"]}; border-color: {p["BLUE"]};
}}
QPushButton#info_icon[active="true"] {{
    background: {p["BG_ACCENT"]}; border-color: {p["BLUE"]};
}}
QFrame#info_hover_card {{
    background: {p["BG_CARD"]};
    border: 1px solid {p["BORDER"]};
    border-radius: 10px;
}}
QLabel#info_hover_text {{
    color: {p["TEXT"]}; font-size: 12px; font-weight: 600;
    background: transparent; padding: 0;
}}
QFrame#info_drawer {{
    background: {p["BG_CARD"]};
    border: none;
    border-right: 1px solid {p["BORDER"]};
}}
QLabel#info_drawer_title {{
    color: {p["TEXT"]}; font-size: 15px; font-weight: 700;
    background: transparent;
}}
QPushButton#info_drawer_close {{
    background: {p["STATS_BG"]}; color: {p["TEXT_SEC"]};
    border: 1px solid {p["BORDER"]}; border-radius: 8px;
    min-width: 30px; max-width: 30px;
    min-height: 30px; max-height: 30px;
    padding: 0; font-size: 16px; font-weight: 500;
}}
QPushButton#info_drawer_close:hover {{
    background: {p["BG_HOVER"]}; color: {p["TEXT"]};
    border-color: {p["BLUE"]};
}}
QTextBrowser#info_drawer_body {{
    background: {p["STATS_BG"]}; color: {p["TEXT_SEC"]};
    border: 1px solid {p["BORDER"]}; border-radius: 10px;
    padding: 12px 14px; font-size: 13px;
}}
QTextBrowser#info_drawer_body a {{
    color: {p["BLUE"]}; text-decoration: none;
}}
QFrame#operation_log_frame {{
    background: {p["STATS_BG"]}; border: 1px solid {p["BORDER"]};
    border-radius: 8px;
}}
QLabel#operation_log_title {{
    color: {p["TEXT"]}; font-size: 12px; font-weight: 600; background: transparent;
}}
QTextEdit#operation_log {{
    background: transparent; color: {p["TEXT_SEC"]};
    border: none; padding: 0 2px;
    font-family: "Menlo", "Consolas", monospace; font-size: 11px;
}}
QTextEdit#output:!focus {{
    color: {p["TEXT_MUTED"]};
}}
"""

# ── Переводы UI ──────────────────────────────
TEXTS: dict[str, dict[str, str]] = {
    "de": {
        "window_title": "Smart Assistant  |  Intelligenter Assistent",
        "app_title": "Intelligenter Assistent",
        "app_sub": "KI-Helfer für Excel und CSV — spart Zeit, Sie prüfen das Ergebnis",
        "disclaimer": (
            "Der Assistent erstellt einen Entwurf — bitte vor der Verwendung prüfen."
        ),
        "ai_banner": (
            "Von KI erstellt — kann Fehler enthalten. "
            "Bitte vor der Nutzung manuell prüfen."
        ),
        "output_empty": (
            "Hier erscheint nach der Verarbeitung eine kurze Zusammenfassung. "
            "Die Tabelle öffnen Sie am besten in Excel — nicht in diesem Fenster."
        ),
        "result_ready_title": "Entwurf fertig",
        "result_ready_hint": (
            "Öffnen Sie die Tabelle in Excel oder Numbers, prüfen Sie sie "
            "und speichern Sie mit «Speichern» oben."
        ),
        "result_rows": "Zeilen im Entwurf: {n}",
        "result_no_file": "Keine Datei — oben eine Tabelle öffnen",
        "result_empty_hint": (
            "Schlüssel und Datei links verbinden, Aufgabe beschreiben — "
            "der Entwurf erscheint in diesem Bereich."
        ),
        "result_empty_title": "Hier erscheint das Ergebnis",
        "result_file_waiting": "In Bearbeitung: {name}",
        "result_file_ready": "Entwurf fertig: {name}",
        "result_actions_hint": "Speichern nicht nötig — Entwurf direkt ansehen",
        "result_ready_title": "Entwurf ist bereit",
        "btn_view_draft": "Ergebnis ansehen",
        "btn_view_draft_tip": "Bearbeitete Tabelle im Programm öffnen (ohne Speichern)",
        "dlg_draft_title": "Entwurf",
        "dlg_draft_rows": "Zeilen: {total} · angezeigt: {shown}",
        "btn_open_draft": "In Excel öffnen",
        "btn_open_draft_tip": "Entwurf in Excel oder Numbers öffnen",
        "progress_opened": "Geöffnet: {name}",
        "btn_more": "Mehr",
        "menu_tip": "Weitere Aktionen",
        "lang_label": "Sprache:",
        "theme_label": "Design:",
        "theme_light": "Hell",
        "theme_dark": "Dunkel",
        "card_api": "API-Schlüssel",
        "card_file": "Datei",
        "card_task": "Aufgabe",
        "wf_setup": "Verbinden",
        "wf_work": "Aufgabe",
        "wf_result": "Ergebnis",
        "sidebar_workflow": "ABLAUF",
        "sidebar_help": "HILFE",
        "workspace_title": "Arbeitsbereich",
        "workspace_sub": "Entwurf erstellen und vor der Nutzung prüfen",
        "workflow_need_key": "Schritt 1: API-Schlüssel eingeben",
        "workflow_need_file": "Schritt 2: Excel- oder CSV-Datei öffnen",
        "workflow_need_task": "Schritt 3: Vorlage wählen oder Aufgabe beschreiben",
        "workflow_need_rules": "Für Beschwerden: Firmenregeln anhängen",
        "workflow_ready": "Bereit — «Entwurf erstellen» klicken",
        "workflow_done": "Entwurf bereit — Ergebnis rechts prüfen",
        "workflow_busy": "Assistent arbeitet…",
        "setup_key_ok": "Schlüssel verbunden",
        "setup_key_pending": "Schlüssel fehlt",
        "setup_file_ok": "{rows} Zeilen · {cols} Spalten",
        "setup_file_pending": "Keine Datei geöffnet",
        "card_result": "Ergebnis",
        "info_panel_title": "Hilfe",
        "info_about_title": "Über die App",
        "info_about": (
            "Intelligenter Assistent ist eine Desktop-App für Excel (.xlsx) und CSV.\n\n"
            "Was die App macht:\n"
            "• Öffnet Ihre Datei lokal auf dem Computer\n"
            "• Sendet Aufgabe und Tabellendaten an Claude (Anthropic)\n"
            "• Erstellt einen bearbeiteten Entwurf und einen Änderungsbericht\n"
            "• Speichert den Entwurf — Sie prüfen und entscheiden über die Nutzung\n\n"
            "Typische Aufgaben: Duplikate, Datumsformat, Sortierung, Leerzeichen, "
            "Klassifizierung von Beschwerden nach Firmenregeln und mehr.\n\n"
            "Für wen: Buchhaltung, Analyse, Support, HR — überall, wo Tabellen "
            "viel manuelle Routine erfordern.\n\n"
            "Wichtig:\n"
            "• Das Ergebnis ist ein Entwurf — immer manuell prüfen\n"
            "• Daten bleiben auf Ihrem Gerät; kein Upload an den Entwickler\n"
            "• Eigener Anthropic-API-Schlüssel nötig (Pay-as-you-go)"
        ),
        "info_steps_title": "So geht's",
        "info_steps": (
            "1. API-Schlüssel eintragen (Anthropic)\n"
            "2. Datei öffnen (.xlsx / .csv)\n"
            "3. Vorlage wählen oder Aufgabe beschreiben\n"
            "4. «Entwurf erstellen» klicken\n"
            "5. Bericht prüfen und Ergebnis speichern\n"
            "6. Bei Beschwerden: «Mehr ▼» → Export in Ordner oder Datei speichern"
        ),
        "info_templates_title": "Vorlagen",
        "info_templates": (
            "Vorlagen sind fertige Aufgabenformulierungen für typische Tabellenarbeit.\n\n"
            "So funktionieren sie:\n"
            "1. Unten auf «— Vorlage wählen —» klicken\n"
            "2. Einen Eintrag wählen — der Text erscheint im Aufgabenfeld\n"
            "3. Die Schaltfläche bleibt blau markiert, solange der Text der Vorlage entspricht\n"
            "4. Text manuell ändern — die Markierung verschwindet\n"
            "5. Oder die Aufgabe ohne Vorlage frei beschreiben\n\n"
            "Enthalten u. a.: Duplikate, Datumsformat, Leerzeichen, Sortierung, "
            "leere Zeilen, Namen aufteilen, Tippfehler und «Beschwerden einordnen».\n\n"
            "Für «Beschwerden einordnen» zusätzlich eine Regeldatei verbinden. "
            "Nach dem Entwurf: «Mehr ▼» → «In Ordner» (Ordner selbst wählen), "
            "«E-Mail» oder «Speichern». «An Abt.» nur bei konfigurierten Ordnerpfaden."
        ),
        "info_data_title": "Daten & Schlüssel",
        "info_data": (
            "<p style='margin-top:0'><b>Wo Daten liegen</b></p>"
            "<p>API-Schlüssel, Einstellungen und der Entwurfsverlauf werden "
            "<b>nur auf Ihrem Gerät</b> gespeichert (im App-Ordner auf der Festplatte). "
            "Es gibt keinen eigenen Server — Tabellen werden nicht zu uns hochgeladen "
            "und nicht beim Entwickler gespeichert.</p>"
            "<p>An Anthropic gehen nur Dateiinhalt und Aufgabentext <b>beim Start</b> "
            "des Assistenten. Ohne Ihren Klick wird nichts gesendet.</p>"
            "<p><b>API-Schlüssel besorgen</b></p>"
            "<p>1. Bei Anthropic registrieren<br>"
            "2. Guthaben aufladen (Pay-as-you-go)<br>"
            "3. Schlüssel in der Konsole erstellen und oben bei «API-Schlüssel» einfügen</p>"
            "<p>Schlüssel erstellen:<br>"
            "<a href='https://console.anthropic.com/settings/keys'>"
            "console.anthropic.com/settings/keys</a></p>"
        ),
        "info_pick": "Wählen Sie ein Thema — der Text erscheint unten.",
        "btn_show": "Anzeigen",
        "btn_hide": "Verbergen",
        "key_loaded": "Schlüssel geladen",
        "key_saved": "Schlüssel gespeichert",
        "btn_open": "Datei wählen",
        "btn_change_file": "Datei wechseln",
        "btn_detach": "Datei entfernen",
        "no_file": "Unterstützt: .xlsx und .csv",
        "stat_rows": "Zeilen",
        "stat_cols": "Spalten",
        "stat_size": "Größe",
        "stat_format": "Format",
        "tpl_label": "Vorlagen:",
        "tpl_choose": "— Vorlage wählen —",
        "file_info": "Datei-Informationen",
        "rules_label": "Firmenregeln:",
        "btn_attach_rules": "Regeldatei verbinden",
        "rules_none": "Keine Regeldatei — für Beschwerden erforderlich",
        "dlg_open_rules": "Regeldatei wählen",
        "filter_rules": "Regeln (*.txt *.md *.csv *.xlsx)",
        "filter_rules_txt": "Text (*.txt *.md)",
        "filter_rules_table": "Tabellen (*.csv *.xlsx)",
        "filter_txt": "Text/TSV (*.txt *.tsv)",
        "filter_all": "Alle Dateien (*)",
        "err_rules_type": "Dieses Regelformat wird nicht unterstützt.",
        "err_rules_type_hint": "Erlaubt: .txt, .md, .csv oder .xlsx",
        "err_no_rules": "Für die Beschwerden-Vorlage fehlen Firmenregeln.",
        "err_no_rules_hint": (
            "Bei «Regeldatei verbinden» klicken "
            "(z. B. company_rules.example.txt)."
        ),
        "own_task": "Aufgabe",
        "task_ph": "Beschreiben Sie, was der Assistent tun soll…",
        "btn_run": "Entwurf erstellen",
        "progress_run": "Assistent arbeitet…",
        "progress_done": "Entwurf fertig — Aktionen oben rechts im Ergebnisbereich.",
        "progress_saved": "Gespeichert: {name}",
        "result_warn": (
            "Entwurf fertig — öffnen Sie den Bericht und prüfen Sie die Änderungen."
        ),
        "report_title": "Arbeitsbericht",
        "btn_view_report": "Bericht anzeigen",
        "btn_close": "Schließen",
        "changes_title": "Was hat sich geändert?",
        "changes_summary": (
            "Zeilen: {before} → {after}. Geänderte Zellen: {cells}. "
            "Entfernte Zeilen: {removed}. Neue Zeilen: {added}."
        ),
        "changes_structure": (
            "Die Tabellenstruktur wurde verändert (Spalten hinzugefügt, entfernt oder "
            "umbenannt). Details unten beziehen sich auf gemeinsame Spalten."
        ),
        "changes_rows_fewer": (
            "Insgesamt {n} Zeile(n) weniger. Einzelne gelöschte Zeilen siehe unten."
        ),
        "changes_none": "Keine erkennbaren Änderungen.",
        "changes_removed_header": "Entfernte Zeilen ({n}):",
        "changes_added_header": "Neue Zeilen ({n}):",
        "changes_row_removed": "  Zeile {row}: {preview}",
        "changes_row_added": "  Zeile {row}: {preview}",
        "changes_cells_header": "Geänderte Zellen ({n}):",
        "changes_cell": "  Zeile {row}, „{col}“: {old} → {new}",
        "changes_more": "  … und {n} weitere.",
        "draft_label": "Vollständiger Entwurf:",
        "save_hint": "Mit dem Speichern bestätigen Sie, dass Sie den Entwurf geprüft haben.",
        "err_draft_parse": "Die Antwort konnte nicht als Tabelle gelesen werden.",
        "err_draft_parse_hint": (
            "Der Entwurf steht unten im Text — oft hat die KI Text vor oder nach der Tabelle ergänzt. "
            "Erneut ausführen oder die Aufgabe kürzer formulieren."
        ),
        "err_no_draft": "Es gibt noch keinen Entwurf zum Speichern.",
        "err_no_draft_hint": "Zuerst «Entwurf erstellen» ausführen und das Ergebnis prüfen.",
        "btn_save": "Speichern",
        "btn_save_tip": "Entwurf als CSV oder Excel speichern (Menü ▼)",
        "btn_history": "Verlauf",
        "btn_history_tip": "Frühere Entwürfe ansehen und erneut speichern (ohne API)",
        "history_title": "Entwurfs-Verlauf",
        "history_empty": "Noch keine Entwürfe gespeichert.",
        "history_task_label": "Aufgabe: {task}",
        "history_open": "In Arbeit öffnen",
        "history_download_csv": "CSV speichern",
        "history_download_xlsx": "Excel speichern",
        "history_delete": "Löschen",
        "history_delete_confirm": "Diesen Eintrag aus dem Verlauf löschen?",
        "history_deleted": "Eintrag gelöscht.",
        "history_opened": "Entwurf aus dem Verlauf geladen.",
        "history_err_load": "Eintrag konnte nicht gelesen werden.",
        "history_preview_summary": "Zusammenfassung",
        "btn_export_routing": "Nach Abteilungen",
        "menu_xlsx": "Als Excel speichern (.xlsx)",
        "menu_csv": "Als CSV speichern (.csv)",
        "menu_routing": "Routing-Export (Abteilungen)",
        "dlg_save_routing": "Routing-Export speichern",
        "sheet_summary": "Zusammenfassung",
        "sheet_review": "Prüfung",
        "sum_dept": "Abteilung",
        "sum_total": "Gesamt",
        "sum_high": "Hohe Priorität",
        "sum_review": "Prüfung nötig",
        "sum_unknown": "Unbekannt",
        "err_no_dept": "Im Entwurf fehlt die Spalte «Abteilung».",
        "err_no_dept_hint": (
            "Zuerst Vorlage «Beschwerden einordnen» ausführen und Firmenregeln verbinden, "
            "dann erneut über «Mehr ▼» → «In Ordner» exportieren."
        ),
        "progress_routing_saved": "Routing exportiert: {name}",
        "btn_deliver_inboxes": "An Abt.",
        "btn_deliver_inboxes_tip": (
            "Nur bei konfigurierten Ordnern (department_folders.demo.txt). "
            "Normalerweise «In Ordner» nutzen."
        ),
        "menu_deliver_inboxes": "An Abt. (konfigurierte Ordner)",
        "dlg_routing_review_title": "Prüfung vor dem Senden",
        "dlg_routing_review_hint": (
            "Prüfen Sie die Aufteilung nach Abteilungen. "
            "Nach Bestätigung landen die Dateien in den Ordnern, "
            "der Status wird «Gesendet»."
        ),
        "dlg_routing_review_intro": "Routing-Übersicht:",
        "btn_confirm_deliver": "An Abteilungen senden",
        "progress_status_sent": "Status aktualisiert: Gesendet ({n} Zeilen)",
        "btn_export_folder": "In Ordner",
        "btn_copy_email": "E-Mail",
        "btn_copy_email_tip": "E-Mail-Text für eine Abteilung in die Zwischenablage kopieren",
        "dlg_pick_folder": "Zielordner wählen",
        "menu_folder": "Routing-Ordner exportieren",
        "progress_folder_saved": "Ordner erstellt: {name}",
        "progress_delivered": "An {n} Abteilung(en) gesendet",
        "deliver_done_title": "An Abteilungen gesendet",
        "deliver_done_body": (
            "Dateien liegen in den Abteilungsordnern:\n{root}\n\n{details}"
        ),
        "err_no_inboxes": "Abteilungsordner sind nicht konfiguriert.",
        "err_no_inboxes_hint": "Datei anlegen oder prüfen:\n{path}",
        "err_nothing_delivered": "Es wurden keine Dateien in Abteilungsordner gelegt.",
        "err_nothing_delivered_hint": (
            "Prüfen Sie Spalte «Abteilung» im Entwurf und Einträge in department_folders.demo.txt."
        ),
        "sheet_unknown_dept": "Abteilung_prüfen",
        "warn_unknown_dept": (
            "{n} Zeile(n) mit unbekannter Abteilung (nicht in den Firmenregeln). "
            "Siehe Filter «Abteilung prüfen» oder Export."
        ),
        "filter_label": "Filter:",
        "filter_all": "Alle",
        "filter_high": "Hohe Priorität",
        "filter_review": "Prüfung nötig",
        "filter_unknown_dept": "Abteilung prüfen",
        "filter_count": "Angezeigt: {n} von {total}",
        "filter_reset": "Alle zeigen",
        "email_subject": "Beschwerden — {department} — {date}",
        "email_body": (
            "Betreff: Beschwerden — {department} — {date}\n\n"
            "Guten Tag,\n\n"
            "im Anhang {count} Beschwerde(n) für die Abteilung «{department}».\n"
            "Hohe Priorität: {high}\n"
            "Prüfung nötig: {review}\n\n"
            "{attach_hint}\n\n"
            "Mit freundlichen Grüßen"
        ),
        "email_attach": "Bitte die passende Excel-Datei aus dem Routing-Ordner anhängen.",
        "email_copied": "E-Mail-Text in die Zwischenablage kopiert.",
        "dlg_email_title": "E-Mail-Text",
        "dlg_email_hint": (
            "Abteilung wählen — unten erscheint der fertige Text. "
            "«Kopieren» drücken und in der E-Mail einfügen (⌘V / Strg+V)."
        ),
        "btn_copy_clipboard": "Kopieren",
        "email_copied_detail": (
            "Text für «{department}» ist in der Zwischenablage.\n\n"
            "E-Mail öffnen und einfügen (⌘V auf Mac, Strg+V auf Windows)."
        ),
        "log_title": "Routing-Export — Protokoll",
        "log_date": "Datum",
        "log_total": "Gesamt",
        "log_review": "Prüfung nötig",
        "log_unknown": "Unbekannte Abteilung",
        "log_by_dept": "Nach Abteilung:",
        "log_unknown_list": "Unbekannte Abteilungsnamen:",
        "operation_log_title": "Protokoll der letzten Aktion",
        "operation_log_deliver": "An Abteilungen gesendet",
        "operation_log_folder": "Routing-Ordner erstellt",
        "operation_log_path": "Ordner: {path}",
        "operation_log_files": "Dateien:",
        "btn_open_folder": "Ordner öffnen",
        "dlg_open": "Datei öffnen",
        "dlg_save_csv": "CSV speichern",
        "dlg_save_xlsx": "Excel speichern",
        "filter_tables": "Tabellen (*.xlsx *.csv *.txt *.tsv)",
        "filter_xlsx": "Excel (*.xlsx)",
        "filter_csv": "CSV (*.csv)",
        "warn_title": "Hinweis",
        "err_title": "Fehler",
        "err_api_title": "API-Fehler",
        "err_api": "Es ist kein API-Schlüssel eingetragen.",
        "err_api_hint": (
            "Tragen Sie Ihren Claude-Schlüssel (sk-ant-…) oben links ein. "
            "Schlüssel erstellen: console.anthropic.com → API Keys."
        ),
        "err_file": "Es ist keine Datei geöffnet.",
        "err_file_hint": "Oben bei «Datei» auf «Datei öffnen» klicken und Excel oder CSV wählen.",
        "err_file_empty": "Die geöffnete Datei enthält keine Datenzeilen.",
        "err_file_empty_hint": "Prüfen Sie die Tabelle oder wählen Sie eine andere Datei.",
        "err_task": "Das Aufgabenfeld ist leer.",
        "err_task_hint": "Wählen Sie eine Vorlage oder beschreiben Sie kurz, was der Assistent tun soll.",
        "err_type": "Dieses Dateiformat wird nicht unterstützt.",
        "err_type_hint": "Erlaubt: .xlsx, .csv, .txt, .tsv",
        "err_read": "Die Datei konnte nicht gelesen werden.",
        "err_read_hint": "Ist die Datei beschädigt, passwortgeschützt oder kein gültiges Tabellenformat?",
        "err_excel": "Excel konnte nicht gespeichert werden.",
        "err_excel_hint": "Technische Details:\n{err}",
        "err_api_auth": "Der API-Schlüssel wurde abgelehnt (ungültig oder abgelaufen).",
        "err_api_auth_hint": "Prüfen Sie den Schlüssel oben links und erstellen Sie bei Bedarf einen neuen.",
        "err_api_rate": "Zu viele Anfragen an die API (Rate-Limit).",
        "err_api_rate_hint": "Warten Sie 1–2 Minuten und starten Sie erneut. Bei großen Dateien weniger Zeilen testen.",
        "err_api_timeout": "Die Anfrage hat zu lange gedauert (Zeitüberschreitung).",
        "err_api_timeout_hint": "Internet prüfen und erneut versuchen. Sehr große Tabellen ggf. aufteilen.",
        "err_api_network": "Keine Verbindung zur API.",
        "err_api_network_hint": "Internetverbindung und Firewall/VPN prüfen, dann erneut versuchen.",
        "err_api_billing": "Problem mit dem API-Guthaben oder Abrechnung.",
        "err_api_billing_hint": "Im Anthropic-Konto prüfen, ob Guthaben oder Zahlungsmethode aktiv ist.",
        "err_api_overload": "Der API-Dienst ist vorübergehend überlastet.",
        "err_api_overload_hint": "In ein paar Minuten erneut versuchen.",
        "err_api_generic": "Beim Aufruf der API ist ein Fehler aufgetreten.",
        "err_api_generic_hint": "Details:\n{detail}",
    },
    "en": {
        "window_title": "Smart Assistant",
        "app_title": "Smart Assistant",
        "app_sub": "Your AI helper for Excel and CSV — saves time, you verify the result",
        "disclaimer": (
            "The assistant creates a draft — please review before use."
        ),
        "ai_banner": (
            "Generated by AI — may contain errors. "
            "Review the result manually before use."
        ),
        "output_empty": (
            "A short summary will appear here after processing. "
            "Open the table in Excel — not in this window."
        ),
        "result_ready_title": "Draft ready",
        "result_ready_hint": (
            "Open the spreadsheet in Excel or Numbers, review it, "
            "then save with «Save» above."
        ),
        "result_rows": "Rows in draft: {n}",
        "result_no_file": "No file — open a spreadsheet above",
        "result_empty_hint": (
            "Connect your key and file on the left, describe the task — "
            "the draft will appear in this panel."
        ),
        "result_empty_title": "Your result will appear here",
        "result_file_waiting": "Working on: {name}",
        "result_file_ready": "Draft ready: {name}",
        "result_actions_hint": "No need to save — preview the draft right away",
        "result_ready_title": "Draft is ready",
        "btn_view_draft": "View result",
        "btn_view_draft_tip": "Open the processed table in the app (no save needed)",
        "dlg_draft_title": "Draft",
        "dlg_draft_rows": "Rows: {total} · shown: {shown}",
        "btn_open_draft": "Open in Excel",
        "btn_open_draft_tip": "Open the draft in Excel or Numbers",
        "progress_opened": "Opened: {name}",
        "btn_more": "More",
        "menu_tip": "More actions",
        "lang_label": "Language:",
        "theme_label": "Theme:",
        "theme_light": "Light",
        "theme_dark": "Dark",
        "card_api": "API key",
        "card_file": "File",
        "card_task": "Task",
        "wf_setup": "Connect",
        "wf_work": "Task",
        "wf_result": "Result",
        "sidebar_workflow": "WORKFLOW",
        "sidebar_help": "HELP",
        "workspace_title": "Workspace",
        "workspace_sub": "Create a draft and review before use",
        "workflow_need_key": "Step 1: enter your Anthropic API key",
        "workflow_need_file": "Step 2: open an Excel or CSV file",
        "workflow_need_task": "Step 3: pick a template or describe the task",
        "workflow_need_rules": "For complaints: attach company rules",
        "workflow_ready": "Ready — click «Create draft»",
        "workflow_done": "Draft ready — review the result on the right",
        "workflow_busy": "Assistant is working…",
        "setup_key_ok": "Key connected",
        "setup_key_pending": "No API key",
        "setup_file_ok": "{rows} rows · {cols} columns",
        "setup_file_pending": "No file open",
        "card_result": "Result",
        "info_panel_title": "Guide",
        "info_about_title": "About",
        "info_about": (
            "Smart Assistant is a desktop app for Excel (.xlsx) and CSV files.\n\n"
            "What it does:\n"
            "• Opens your file locally on your computer\n"
            "• Sends the task and table data to Claude (Anthropic)\n"
            "• Returns an edited draft and a change report\n"
            "• Saves the draft — you review and decide whether to use it\n\n"
            "Typical tasks: duplicates, date format, sorting, whitespace, "
            "complaint classification by company rules, and more.\n\n"
            "Who it's for: finance, analytics, support, HR — anywhere "
            "spreadsheets need repetitive manual work.\n\n"
            "Important:\n"
            "• The output is a draft — always verify manually\n"
            "• Data stays on your device; nothing is uploaded to the developer\n"
            "• You need your own Anthropic API key (pay-as-you-go)"
        ),
        "info_steps_title": "How to use",
        "info_steps": (
            "1. Enter your Anthropic API key\n"
            "2. Open a file (.xlsx / .csv)\n"
            "3. Pick a template or describe the task\n"
            "4. Click «Create draft»\n"
            "5. Review the report and save the result\n"
            "6. For complaints: «More ▼» → export to folder or save file"
        ),
        "info_templates_title": "Templates",
        "info_templates": (
            "Templates are ready-made task descriptions for common spreadsheet work.\n\n"
            "How they work:\n"
            "1. Click «— Select a template —» at the bottom\n"
            "2. Pick an item — its text fills the task field\n"
            "3. The button stays highlighted in blue while the text matches the template\n"
            "4. Edit the text manually — the highlight clears\n"
            "5. Or describe the task from scratch without a template\n\n"
            "Includes: duplicates, date format, whitespace, sorting, empty rows, "
            "name split, typos, and «Classify complaints».\n\n"
            "For «Classify complaints», also attach a company rules file. "
            "After the draft: «More ▼» → «To folder» (you pick the folder), "
            "«Email», or «Save». «To depts» only if department folders are configured."
        ),
        "info_data_title": "Data & key",
        "info_data": (
            "<p style='margin-top:0'><b>Where data is stored</b></p>"
            "<p>Your API key, settings, and draft history are kept "
            "<b>only on your device</b> (in the app folder on disk). "
            "There is no app server — tables are not uploaded to us "
            "or stored by the developer.</p>"
            "<p>Anthropic receives only the file content and task text "
            "<b>when you run</b> the assistant. Nothing is sent without your action.</p>"
            "<p><b>Get an API key</b></p>"
            "<p>1. Sign up at Anthropic<br>"
            "2. Add billing (pay-as-you-go)<br>"
            "3. Create a key in the console and paste it into «API key» at the top</p>"
            "<p>Create a key:<br>"
            "<a href='https://console.anthropic.com/settings/keys'>"
            "console.anthropic.com/settings/keys</a></p>"
        ),
        "info_pick": "Pick a topic — the text will appear below.",
        "btn_show": "Show",
        "btn_hide": "Hide",
        "key_loaded": "Key loaded",
        "key_saved": "Key saved",
        "btn_open": "Select file",
        "btn_change_file": "Change file",
        "btn_detach": "Remove file",
        "no_file": "Supports .xlsx and .csv",
        "stat_rows": "Rows",
        "stat_cols": "Columns",
        "stat_size": "Size",
        "stat_format": "Format",
        "tpl_label": "Templates:",
        "tpl_choose": "— Select a template —",
        "file_info": "File details",
        "rules_label": "Company rules:",
        "btn_attach_rules": "Attach rules file",
        "rules_none": "No rules file selected — required for complaints",
        "dlg_open_rules": "Select rules file",
        "filter_rules": "Rules (*.txt *.md *.csv *.xlsx)",
        "filter_rules_txt": "Text (*.txt *.md)",
        "filter_rules_table": "Tables (*.csv *.xlsx)",
        "filter_txt": "Text/TSV (*.txt *.tsv)",
        "filter_all": "All files (*)",
        "err_rules_type": "This rules file format is not supported.",
        "err_rules_type_hint": "Allowed: .txt, .md, .csv, or .xlsx",
        "err_no_rules": "Company rules are required for the complaints template.",
        "err_no_rules_hint": (
            "Click «Attach rules file» "
            "(e.g. company_rules.example.txt)."
        ),
        "own_task": "Task",
        "task_ph": "Describe what the assistant should do…",
        "btn_run": "Create draft",
        "progress_run": "Assistant working…",
        "progress_done": "Draft ready — use the actions in the result header.",
        "progress_saved": "Saved: {name}",
        "result_warn": (
            "Draft ready — open the report and review changes before saving."
        ),
        "report_title": "Work report",
        "btn_view_report": "View report",
        "btn_close": "Close",
        "changes_title": "What changed?",
        "changes_summary": (
            "Rows: {before} → {after}. Changed cells: {cells}. "
            "Rows removed: {removed}. Rows added: {added}."
        ),
        "changes_structure": (
            "Table structure changed (columns added, removed, or renamed). "
            "Details below use common columns only."
        ),
        "changes_rows_fewer": (
            "{n} fewer row(s) in total. See deleted rows below."
        ),
        "changes_none": "No detectable changes.",
        "changes_removed_header": "Removed rows ({n}):",
        "changes_added_header": "Added rows ({n}):",
        "changes_row_removed": "  Row {row}: {preview}",
        "changes_row_added": "  Row {row}: {preview}",
        "changes_cells_header": "Changed cells ({n}):",
        "changes_cell": "  Row {row}, \"{col}\": {old} → {new}",
        "changes_more": "  … and {n} more.",
        "draft_label": "Full draft:",
        "save_hint": "By saving, you confirm you have reviewed the draft.",
        "err_draft_parse": "The response could not be read as a table.",
        "err_draft_parse_hint": (
            "See the draft text below — the AI may have added text before or after the table. "
            "Run again or shorten the task."
        ),
        "err_no_draft": "There is no draft to save yet.",
        "err_no_draft_hint": "Run «Create draft» first and review the result.",
        "btn_save": "Save",
        "btn_save_tip": "Save draft as CSV or Excel (▼ menu)",
        "btn_history": "History",
        "btn_history_tip": "Browse past drafts and save again (no API call)",
        "history_title": "Draft history",
        "history_empty": "No saved drafts yet.",
        "history_task_label": "Task: {task}",
        "history_open": "Open in workspace",
        "history_download_csv": "Save CSV",
        "history_download_xlsx": "Save Excel",
        "history_delete": "Delete",
        "history_delete_confirm": "Delete this entry from history?",
        "history_deleted": "Entry deleted.",
        "history_opened": "Draft loaded from history.",
        "history_err_load": "Could not read this history entry.",
        "history_preview_summary": "Summary",
        "btn_export_routing": "By department",
        "menu_xlsx": "Save as Excel (.xlsx)",
        "menu_csv": "Save as CSV (.csv)",
        "menu_routing": "Routing export (departments)",
        "dlg_save_routing": "Save routing export",
        "sheet_summary": "Summary",
        "sheet_review": "Needs review",
        "sum_dept": "Department",
        "sum_total": "Total",
        "sum_high": "High priority",
        "sum_review": "Needs review",
        "sum_unknown": "Unknown",
        "err_no_dept": "The draft has no «Department» column.",
        "err_no_dept_hint": (
            "Run the «Classify complaints» template with company rules attached, "
            "then export again via «More ▼» → «To folder»."
        ),
        "progress_routing_saved": "Routing exported: {name}",
        "btn_deliver_inboxes": "To depts",
        "btn_deliver_inboxes_tip": (
            "Only for configured folders (department_folders.demo.txt). "
            "Usually use «To folder» instead."
        ),
        "menu_deliver_inboxes": "To depts (configured folders)",
        "dlg_routing_review_title": "Review before sending",
        "dlg_routing_review_hint": (
            "Check how many complaints go to each department. "
            "After you confirm, files are placed in department folders "
            "and row status becomes «Sent»."
        ),
        "dlg_routing_review_intro": "Routing summary:",
        "btn_confirm_deliver": "Send to departments",
        "progress_status_sent": "Status updated: Sent ({n} rows)",
        "btn_export_folder": "To folder",
        "btn_copy_email": "Email",
        "btn_copy_email_tip": "Copy ready email text for a department to the clipboard",
        "dlg_pick_folder": "Choose destination folder",
        "menu_folder": "Export routing folder",
        "progress_folder_saved": "Folder created: {name}",
        "progress_delivered": "Sent to {n} department(s)",
        "deliver_done_title": "Sent to departments",
        "deliver_done_body": (
            "Files are in department folders:\n{root}\n\n{details}"
        ),
        "err_no_inboxes": "Department folders are not configured.",
        "err_no_inboxes_hint": "Create or check this file:\n{path}",
        "err_nothing_delivered": "No files were placed in department folders.",
        "err_nothing_delivered_hint": (
            "Check the «Department» column in the draft and entries in department_folders.demo.txt."
        ),
        "sheet_unknown_dept": "Check_department",
        "warn_unknown_dept": (
            "{n} row(s) with unknown department (not in company rules). "
            "See filter «Check department» or export."
        ),
        "filter_label": "Filter:",
        "filter_all": "All",
        "filter_high": "High priority",
        "filter_review": "Needs review",
        "filter_unknown_dept": "Check department",
        "filter_count": "Showing: {n} of {total}",
        "filter_reset": "Show all",
        "email_subject": "Complaints — {department} — {date}",
        "email_body": (
            "Subject: Complaints — {department} — {date}\n\n"
            "Hello,\n\n"
            "Attached are {count} complaint(s) for «{department}».\n"
            "High priority: {high}\n"
            "Needs review: {review}\n\n"
            "{attach_hint}\n\n"
            "Best regards"
        ),
        "email_attach": "Please attach the matching Excel file from the routing folder.",
        "email_copied": "Email text copied to clipboard.",
        "dlg_email_title": "Email text",
        "dlg_email_hint": (
            "Choose a department — the ready text appears below. "
            "Click «Copy» and paste it into your email (⌘V / Ctrl+V)."
        ),
        "btn_copy_clipboard": "Copy",
        "email_copied_detail": (
            "Text for «{department}» is in the clipboard.\n\n"
            "Open your email app and paste (⌘V on Mac, Ctrl+V on Windows)."
        ),
        "log_title": "Routing export — log",
        "log_date": "Date",
        "log_total": "Total",
        "log_review": "Needs review",
        "log_unknown": "Unknown department",
        "log_by_dept": "By department:",
        "log_unknown_list": "Unknown department names:",
        "operation_log_title": "Last action log",
        "operation_log_deliver": "Sent to departments",
        "operation_log_folder": "Routing folder created",
        "operation_log_path": "Folder: {path}",
        "operation_log_files": "Files:",
        "btn_open_folder": "Open folder",
        "dlg_open": "Open file",
        "dlg_save_csv": "Save CSV",
        "dlg_save_xlsx": "Save Excel",
        "filter_tables": "Tables (*.xlsx *.csv *.txt *.tsv)",
        "filter_xlsx": "Excel (*.xlsx)",
        "filter_csv": "CSV (*.csv)",
        "warn_title": "Notice",
        "err_title": "Error",
        "err_api_title": "API error",
        "err_api": "No API key entered.",
        "err_api_hint": (
            "Enter your Claude key (sk-ant-…) at the top left. "
            "Create a key at console.anthropic.com → API Keys."
        ),
        "err_file": "No file is open.",
        "err_file_hint": "At the top, under «File», click «Select file» and choose Excel or CSV.",
        "err_file_empty": "The open file has no data rows.",
        "err_file_empty_hint": "Check the table or choose another file.",
        "err_task": "The task field is empty.",
        "err_task_hint": "Pick a template or briefly describe what the assistant should do.",
        "err_type": "This file format is not supported.",
        "err_type_hint": "Allowed: .xlsx, .csv, .txt, .tsv",
        "err_read": "The file could not be read.",
        "err_read_hint": "Is the file corrupted, password-protected, or not a valid table?",
        "err_excel": "Could not save Excel.",
        "err_excel_hint": "Technical details:\n{err}",
        "err_api_auth": "The API key was rejected (invalid or expired).",
        "err_api_auth_hint": "Check the key at the top left and create a new one if needed.",
        "err_api_rate": "Too many API requests (rate limit).",
        "err_api_rate_hint": "Wait 1–2 minutes and try again. For large files, test with fewer rows.",
        "err_api_timeout": "The request took too long (timeout).",
        "err_api_timeout_hint": "Check your internet and try again. Split very large tables if needed.",
        "err_api_network": "Could not connect to the API.",
        "err_api_network_hint": "Check internet, firewall, or VPN, then try again.",
        "err_api_billing": "API credit or billing issue.",
        "err_api_billing_hint": "In your Anthropic account, check balance or payment method.",
        "err_api_overload": "The API service is temporarily overloaded.",
        "err_api_overload_hint": "Try again in a few minutes.",
        "err_api_generic": "An error occurred while calling the API.",
        "err_api_generic_hint": "Details:\n{detail}",
    },
    "ru": {
        "window_title": "Умный помощник  |  Smart Assistant",
        "app_title": "Умный помощник",
        "app_sub": "ИИ-помощник для Excel и CSV — ускоряет рутину, финальное решение за вами",
        "disclaimer": (
            "Помощник создаёт черновик — перед использованием проверьте результат."
        ),
        "ai_banner": (
            "Черновик создан ИИ и может содержать ошибки. "
            "Перед использованием обязательно проверьте результат вручную."
        ),
        "output_empty": (
            "Здесь появится краткий итог после обработки. "
            "Таблицу удобнее смотреть в Excel — не в этом окне."
        ),
        "result_ready_title": "Черновик готов",
        "result_ready_hint": (
            "Откройте таблицу в Excel или Numbers, проверьте и сохраните "
            "кнопкой «Сохранить» вверху."
        ),
        "result_rows": "Строк в черновике: {n}",
        "result_no_file": "Файл не выбран — откройте таблицу вверху",
        "result_empty_hint": (
            "Подключите ключ и файл слева, опишите задачу — "
            "готовый черновик появится в этой панели."
        ),
        "result_empty_title": "Здесь будет результат",
        "result_file_waiting": "Сейчас в работе: {name}",
        "result_file_ready": "Черновик готов: {name}",
        "result_actions_hint": "Сохранять не нужно — смотрите результат сразу",
        "result_ready_title": "Черновик готов",
        "btn_view_draft": "Посмотреть результат",
        "btn_view_draft_tip": "Открыть таблицу в приложении без сохранения",
        "dlg_draft_title": "Черновик",
        "dlg_draft_rows": "Строк: {total} · показано: {shown}",
        "btn_open_draft": "Открыть в Excel",
        "btn_open_draft_tip": "Открыть черновик в Excel или Numbers",
        "progress_opened": "Открыто: {name}",
        "btn_more": "Ещё",
        "menu_tip": "Дополнительные действия",
        "lang_label": "Язык:",
        "theme_label": "Тема:",
        "theme_light": "Светлая",
        "theme_dark": "Тёмная",
        "card_api": "API-ключ",
        "card_file": "Файл",
        "card_task": "Задача",
        "wf_setup": "Подключение",
        "wf_work": "Задача",
        "wf_result": "Результат",
        "sidebar_workflow": "ШАГИ",
        "sidebar_help": "СПРАВКА",
        "workspace_title": "Рабочая область",
        "workspace_sub": "Черновик для проверки перед использованием",
        "workflow_need_key": "Шаг 1: введите API-ключ Anthropic",
        "workflow_need_file": "Шаг 2: откройте таблицу Excel или CSV",
        "workflow_need_task": "Шаг 3: выберите шаблон или опишите задачу",
        "workflow_need_rules": "Для жалоб: подключите файл правил компании",
        "workflow_ready": "Готово — нажмите «Получить черновик»",
        "workflow_done": "Черновик готов — проверьте результат справа",
        "workflow_busy": "Помощник обрабатывает файл…",
        "setup_key_ok": "Ключ подключён",
        "setup_key_pending": "Ключ не указан",
        "setup_file_ok": "{rows} строк · {cols} колонок",
        "setup_file_pending": "Файл не открыт",
        "card_result": "Результат",
        "info_panel_title": "Справка",
        "info_about_title": "О приложении",
        "info_about": (
            "Умный помощник v1.1.0 — настольное приложение для работы с таблицами "
            "Excel (.xlsx) и CSV.\n\n"
            "Что делает приложение:\n"
            "• Открывает ваш файл локально на компьютере\n"
            "• По задаче (шаблону или своему тексту) отправляет данные в Claude (Anthropic)\n"
            "• Получает обработанную таблицу и краткий отчёт об изменениях\n"
            "• Сохраняет черновик — вы проверяете и решаете, использовать ли результат\n\n"
            "Типовые задачи: дубли, даты, сортировка, пробелы, классификация жалоб "
            "по правилам компании и другое.\n\n"
            "Для кого: бухгалтерия, аналитика, поддержка, HR — везде, где таблицы "
            "требуют ручной рутины.\n\n"
            "Важно:\n"
            "• Результат — черновик, не финальный документ; всегда проверяйте вручную\n"
            "• Данные хранятся на вашем устройстве (папка Application Support); "
            "разработчику не передаются\n"
            "• Нужен ваш API-ключ Anthropic (оплата по факту использования)"
        ),
        "info_steps_title": "Как работать",
        "info_steps": (
            "1. Введите API-ключ Anthropic\n"
            "2. Откройте файл (.xlsx / .csv)\n"
            "3. Выберите шаблон или опишите задачу\n"
            "4. Нажмите «Получить черновик»\n"
            "5. Откройте таблицу и проверьте результат\n"
            "6. Для жалоб: сохраните Excel и отфильтруйте по колонке «Отдел»"
        ),
        "info_templates_title": "Шаблоны",
        "info_templates": (
            "Шаблоны — готовые формулировки задач для типовой обработки таблиц.\n\n"
            "Как это работает:\n"
            "1. Нажмите «— Выберите шаблон —» внизу окна\n"
            "2. Выберите пункт из списка — текст подставится в поле задачи\n"
            "3. Кнопка подсвечивается синим, пока текст совпадает с шаблоном\n"
            "4. Измените текст вручную — подсветка снимется\n"
            "5. Или опишите задачу с нуля без шаблона\n\n"
            "В списке: дубли, даты, пробелы, сортировка, пустые строки, "
            "разбивка ФИО, опечатки и «Классификация жалоб».\n\n"
            "Для «Классификация жалоб» дополнительно подключите файл правил компании. "
            "После черновика сохраните файл — в Excel отсортируйте или отфильтруйте по «Отдел». "
            "Дополнительно в «Ещё ▼»: экспорт в папку, письмо. «В отделы» — только при настройке папок."
        ),
        "info_data_title": "Данные и ключ",
        "info_data": (
            "<p style='margin-top:0'><b>Где хранятся данные</b></p>"
            "<p>API-ключ, настройки и история черновиков лежат "
            "<b>только на вашем устройстве</b> (в папке приложения на диске). "
            "Своего сервера у приложения нет — таблицы никуда не выгружаются "
            "и не хранятся у разработчика.</p>"
            "<p>В Anthropic уходит только содержимое файла и текст задачи "
            "<b>в момент запуска</b> помощника. Без вашего нажатия ничего не отправляется.</p>"
            "<p><b>Где взять API-ключ</b></p>"
            "<p>1. Зарегистрируйтесь на сайте Anthropic<br>"
            "2. Пополните баланс (оплата по факту использования)<br>"
            "3. Создайте ключ в консоли и вставьте его в поле «API-ключ» вверху</p>"
            "<p>Создать ключ:<br>"
            "<a href='https://console.anthropic.com/settings/keys'>"
            "console.anthropic.com/settings/keys</a></p>"
        ),
        "info_pick": "Выберите раздел — текст появится ниже.",
        "btn_show": "Показать",
        "btn_hide": "Скрыть",
        "key_loaded": "Ключ загружен",
        "key_saved": "Ключ сохранён",
        "btn_open": "Выбрать файл",
        "btn_change_file": "Сменить файл",
        "btn_detach": "Открепить файл",
        "no_file": "Поддерживаются .xlsx и .csv",
        "stat_rows": "Строк",
        "stat_cols": "Колонок",
        "stat_size": "Размер",
        "stat_format": "Формат",
        "tpl_label": "Шаблоны:",
        "tpl_choose": "— Выберите шаблон —",
        "file_info": "Информация о файле",
        "rules_label": "Правила компании:",
        "btn_attach_rules": "Подключить файл правил",
        "rules_none": "Файл правил не выбран — для жалоб обязателен",
        "dlg_open_rules": "Выберите файл правил",
        "filter_rules": "Правила (*.txt *.md *.csv *.xlsx)",
        "filter_rules_txt": "Текст (*.txt *.md)",
        "filter_rules_table": "Таблицы (*.csv *.xlsx)",
        "filter_txt": "Текст/TSV (*.txt *.tsv)",
        "filter_all": "Все файлы (*)",
        "err_rules_type": "Этот формат файла правил не поддерживается.",
        "err_rules_type_hint": "Допустимо: .txt, .md, .csv или .xlsx",
        "err_no_rules": "Для шаблона жалоб не подключены правила компании.",
        "err_no_rules_hint": (
            "Нажмите «Подключить файл правил» "
            "(например, company_rules.example.txt)."
        ),
        "own_task": "Задача",
        "task_ph": "Опишите, что нужно сделать помощнику…",
        "btn_run": "Получить черновик",
        "progress_run": "Помощник работает…",
        "progress_done": "Черновик готов — действия в шапке панели «Результат».",
        "progress_saved": "Сохранено: {name}",
        "result_warn": (
            "Черновик готов — откройте отчёт и проверьте изменения."
        ),
        "report_title": "Отчёт о проделанной работе",
        "btn_view_report": "Просмотр",
        "btn_close": "Закрыть",
        "changes_title": "Что изменилось",
        "changes_summary": (
            "Строк: {before} → {after}. Изменено ячеек: {cells}. "
            "Удалено строк: {removed}. Добавлено строк: {added}."
        ),
        "changes_structure": (
            "Структура таблицы изменилась (колонки добавлены, удалены или переименованы). "
            "Детали ниже — по общим колонкам."
        ),
        "changes_rows_fewer": (
            "Всего на {n} строк(и) меньше. Список удалённых строк — ниже."
        ),
        "changes_none": "Изменений не обнаружено.",
        "changes_removed_header": "Удалённые строки ({n}):",
        "changes_added_header": "Добавленные строки ({n}):",
        "changes_row_removed": "  Строка {row}: {preview}",
        "changes_row_added": "  Строка {row}: {preview}",
        "changes_cells_header": "Изменённые ячейки ({n}):",
        "changes_cell": "  Строка {row}, «{col}»: {old} → {new}",
        "changes_more": "  … и ещё {n}.",
        "draft_label": "Полный черновик:",
        "save_hint": "Сохраняя файл, вы подтверждаете, что просмотрели черновик.",
        "err_draft_parse": "Ответ не удалось прочитать как таблицу.",
        "err_draft_parse_hint": (
            "Текст черновика ниже — ИИ мог добавить пояснения до или после таблицы. "
            "Запустите снова или сократите задачу."
        ),
        "err_no_draft": "Пока нет черновика для сохранения.",
        "err_no_draft_hint": "Сначала нажмите «Получить черновик» и проверьте результат.",
        "btn_save": "Сохранить",
        "btn_save_tip": "Сохранить черновик в CSV или Excel (меню ▼)",
        "btn_history": "История",
        "btn_history_tip": "Просмотр прошлых черновиков и повторное сохранение (без API)",
        "history_title": "История черновиков",
        "history_empty": "Пока нет сохранённых черновиков.",
        "history_task_label": "Задача: {task}",
        "history_open": "Открыть в работе",
        "history_download_csv": "Скачать CSV",
        "history_download_xlsx": "Скачать Excel",
        "history_delete": "Удалить",
        "history_delete_confirm": "Удалить эту запись из истории?",
        "history_deleted": "Запись удалена.",
        "history_opened": "Черновик загружен из истории.",
        "history_err_load": "Не удалось прочитать запись истории.",
        "history_preview_summary": "Сводка",
        "btn_export_routing": "По отделам",
        "menu_xlsx": "Сохранить как Excel (.xlsx)",
        "menu_csv": "Сохранить как CSV (.csv)",
        "menu_routing": "Экспорт по отделам",
        "dlg_save_routing": "Сохранить экспорт по отделам",
        "sheet_summary": "Сводка",
        "sheet_review": "На проверку",
        "sum_dept": "Отдел",
        "sum_total": "Всего",
        "sum_high": "Высокий приоритет",
        "sum_review": "Требует проверки",
        "sum_unknown": "Не указан",
        "err_no_dept": "В черновике нет колонки «Отдел».",
        "err_no_dept_hint": (
            "Сначала выполните шаблон «Классификация жалоб» с подключёнными правилами, "
            "затем снова экспортируйте через «Ещё ▼» → «Экспорт в папку»."
        ),
        "progress_routing_saved": "Экспорт по отделам: {name}",
        "btn_deliver_inboxes": "В отделы",
        "btn_deliver_inboxes_tip": (
            "Только для настроенных папок (department_folders.demo.txt). "
            "Обычно используйте «Экспорт в папку»."
        ),
        "menu_deliver_inboxes": "В отделы (настроенные папки)",
        "dlg_routing_review_title": "Проверка перед отправкой",
        "dlg_routing_review_hint": (
            "Проверьте, сколько обращений уйдёт в каждый отдел. "
            "После подтверждения файлы появятся в папках отделов, "
            "а статус строк станет «Отправлено»."
        ),
        "dlg_routing_review_intro": "Сводка маршрутизации:",
        "btn_confirm_deliver": "Разложить по отделам",
        "progress_status_sent": "Статус обновлён: Отправлено ({n} строк)",
        "btn_export_folder": "В папку",
        "btn_copy_email": "Письмо",
        "btn_copy_email_tip": "Скопировать готовый текст письма для отдела в буфер обмена",
        "dlg_pick_folder": "Выберите папку",
        "menu_folder": "Экспорт в папку",
        "progress_folder_saved": "Папка создана: {name}",
        "progress_delivered": "Отправлено в {n} отдел(ов)",
        "deliver_done_title": "Отправлено в отделы",
        "deliver_done_body": (
            "Файлы лежат в папках отделов:\n{root}\n\n{details}\n\n"
            "В каждой папке: Excel, «жалобы_прочитать_….txt» и «письмо_….txt»."
        ),
        "err_no_inboxes": "Папки отделов не настроены.",
        "err_no_inboxes_hint": "Создайте или проверьте файл:\n{path}",
        "err_nothing_delivered": "Ни один файл не был разложен по папкам отделов.",
        "err_nothing_delivered_hint": (
            "Проверьте колонку «Отдел» в черновике и записи в department_folders.demo.txt."
        ),
        "sheet_unknown_dept": "Проверить_отдел",
        "warn_unknown_dept": (
            "{n} строк с неизвестным отделом (нет в правилах компании). "
            "См. фильтр «Проверить отдел» или экспорт."
        ),
        "filter_label": "Фильтр:",
        "filter_all": "Все",
        "filter_high": "Высокий приоритет",
        "filter_review": "Требует проверки",
        "filter_unknown_dept": "Проверить отдел",
        "filter_count": "Показано: {n} из {total}",
        "filter_reset": "Показать все",
        "email_subject": "Жалобы — {department} — {date}",
        "email_body": (
            "Тема: Жалобы — {department} — {date}\n\n"
            "Здравствуйте!\n\n"
            "Во вложении {count} обращение(й) для отдела «{department}».\n"
            "Высокий приоритет: {high}\n"
            "Требует проверки: {review}\n\n"
            "{attach_hint}\n\n"
            "С уважением"
        ),
        "email_attach": "Приложите соответствующий Excel-файл из папки экспорта.",
        "email_copied": "Текст письма скопирован в буфер обмена.",
        "dlg_email_title": "Текст письма",
        "dlg_email_hint": (
            "Выберите отдел — ниже появится готовый текст письма. "
            "Нажмите «Скопировать» и вставьте его в почту (⌘V или Ctrl+V)."
        ),
        "btn_copy_clipboard": "Скопировать",
        "email_copied_detail": (
            "Текст для отдела «{department}» скопирован в буфер обмена.\n\n"
            "Откройте почту и вставьте (⌘V на Mac, Ctrl+V на Windows)."
        ),
        "log_title": "Экспорт маршрутизации — журнал",
        "log_date": "Дата",
        "log_total": "Всего",
        "log_review": "Требует проверки",
        "log_unknown": "Неизвестный отдел",
        "log_by_dept": "По отделам:",
        "log_unknown_list": "Неизвестные названия отделов:",
        "operation_log_title": "Журнал последней операции",
        "operation_log_deliver": "Отправка в отделы",
        "operation_log_folder": "Экспорт в папку",
        "operation_log_path": "Папка: {path}",
        "operation_log_files": "Файлы:",
        "btn_open_folder": "Открыть папку",
        "dlg_open": "Открыть файл",
        "dlg_save_csv": "Сохранить CSV",
        "dlg_save_xlsx": "Сохранить Excel",
        "filter_tables": "Таблицы (*.xlsx *.csv *.txt *.tsv)",
        "filter_xlsx": "Excel (*.xlsx)",
        "filter_csv": "CSV (*.csv)",
        "warn_title": "Внимание",
        "err_title": "Ошибка",
        "err_api_title": "Ошибка API",
        "err_api": "API-ключ не введён.",
        "err_api_hint": (
            "Введите ключ Claude (sk-ant-…) вверху слева. "
            "Создать ключ: console.anthropic.com → API Keys."
        ),
        "err_file": "Файл не открыт.",
        "err_file_hint": "Вверху, в блоке «Файл», нажмите «Выбрать файл» и выберите Excel или CSV.",
        "err_file_empty": "В открытом файле нет строк с данными.",
        "err_file_empty_hint": "Проверьте таблицу или выберите другой файл.",
        "err_task": "Поле задачи пустое.",
        "err_task_hint": "Выберите шаблон или кратко опишите, что должен сделать помощник.",
        "err_type": "Этот формат файла не поддерживается.",
        "err_type_hint": "Допустимо: .xlsx, .csv, .txt, .tsv",
        "err_read": "Не удалось прочитать файл.",
        "err_read_hint": "Файл повреждён, защищён паролем или это не таблица?",
        "err_excel": "Не удалось сохранить Excel.",
        "err_excel_hint": "Технические детали:\n{err}",
        "err_api_auth": "API-ключ отклонён (неверный или просроченный).",
        "err_api_auth_hint": "Проверьте ключ вверху слева и при необходимости создайте новый.",
        "err_api_rate": "Слишком много запросов к API (лимит).",
        "err_api_rate_hint": "Подождите 1–2 минуты и повторите. Для больших файлов протестируйте на меньшем числе строк.",
        "err_api_timeout": "Запрос занял слишком много времени (таймаут).",
        "err_api_timeout_hint": "Проверьте интернет и повторите. Очень большие таблицы можно разбить.",
        "err_api_network": "Нет соединения с API.",
        "err_api_network_hint": "Проверьте интернет, файрвол или VPN и повторите.",
        "err_api_billing": "Проблема с балансом или оплатой API.",
        "err_api_billing_hint": "В аккаунте Anthropic проверьте баланс или способ оплаты.",
        "err_api_overload": "Сервис API временно перегружен.",
        "err_api_overload_hint": "Повторите через несколько минут.",
        "err_api_generic": "При обращении к API произошла ошибка.",
        "err_api_generic_hint": "Подробности:\n{detail}",
    },
}

LANG_OPTIONS = [("de", "Deutsch"), ("en", "English"), ("ru", "Русский")]


TASK_TEMPLATES: dict[str, list[tuple[str, str]]] = {
    "de": [
        ("Duplikate entfernen",
         "Entferne vollständig doppelte Zeilen. Zähle Zeilen als Duplikat, "
         "wenn alle Spalten übereinstimmen. Behalte nur das erste Vorkommen."),
        ("Datumsformat DD.MM.JJJJ",
         "Finde alle Datumsspalten und formatiere sie einheitlich als DD.MM.JJJJ. "
         "Unbekannte Werte unverändert lassen."),
        ("Leerzeichen bereinigen",
         "Entferne führende und nachfolgende Leerzeichen in Textzellen. "
         "Ersetze mehrere Leerzeichen innerhalb des Textes durch eines."),
        ("Nach 1. Spalte sortieren",
         "Sortiere alle Zeilen nach der ersten Spalte aufsteigend. "
         "Die Kopfzeile bleibt an erster Position."),
        ("Leere Zeilen löschen",
         "Lösche alle vollständig leeren Zeilen (ohne Werte)."),
        ("Name in Spalten aufteilen",
         "Wenn eine Spalte den vollständigen Namen enthält, "
         "teile sie in Nachname, Vorname und ggf. Zweitname auf."),
        ("Tippfehler korrigieren",
         "Korrigiere offensichtliche Tippfehler in Textspalten. "
         "Zahlen und Datums nicht ändern."),
        ("Beschwerden einordnen", complaints_task_text("de")),
    ],
    "en": [
        ("Remove duplicates",
         "Remove fully duplicate rows. Rows are duplicates if all columns match. "
         "Keep only the first occurrence."),
        ("Date format DD.MM.YYYY",
         "Find all date columns and format them uniformly as DD.MM.YYYY. "
         "Leave unrecognized values unchanged."),
        ("Trim whitespace",
         "Remove leading and trailing spaces in text cells. "
         "Replace multiple consecutive spaces with a single space."),
        ("Sort by 1st column",
         "Sort all rows by the first column ascending. "
         "Keep the header row at the top."),
        ("Delete empty rows",
         "Delete all completely empty rows (no values in any cell)."),
        ("Split full name",
         "If a column contains a full name, split it into "
         "separate Last name, First name, and Middle name columns."),
        ("Fix typos",
         "Fix obvious typos in text columns. Do not change numbers or dates."),
        ("Classify complaints", complaints_task_text("en")),
    ],
    "ru": [
        ("Удалить дубли",
         "Удали полностью дублирующиеся строки. Считай дублями строки, "
         "у которых совпадают все колонки. Оставь только первое вхождение."),
        ("Формат дат ДД.ММ.ГГГГ",
         "Найди все колонки с датами и приведи их к единому формату ДД.ММ.ГГГГ. "
         "Нераспознанные значения оставь как есть."),
        ("Убрать пробелы",
         "Убери лишние пробелы в начале и конце текстовых ячеек. "
         "Замени несколько подряд идущих пробелов на один."),
        ("Сортировка по 1-й колонке",
         "Отсортируй все строки по первой колонке по возрастанию. "
         "Строку заголовков оставь первой."),
        ("Удалить пустые строки",
         "Удали все полностью пустые строки (где нет ни одного значения)."),
        ("Разбить ФИО",
         "Если есть колонка с полным ФИО, разбей её на "
         "отдельные колонки: Фамилия, Имя, Отчество."),
        ("Исправить опечатки",
         "Исправь очевидные опечатки в текстовых колонках. "
         "Не меняй числа и даты."),
        ("Классификация жалоб", complaints_task_text("ru")),
    ],
}


def load_language() -> str:
    if os.path.exists(user_data.LANG_FILE):
        try:
            lang = open(user_data.LANG_FILE, encoding="utf-8").read().strip().lower()
            if lang in TEXTS:
                return lang
        except OSError:
            pass
    return "de"


def save_language(lang: str) -> None:
    try:
        with open(user_data.LANG_FILE, "w", encoding="utf-8") as f:
            f.write(lang)
    except OSError:
        pass


def load_theme() -> str:
    if os.path.exists(user_data.THEME_FILE):
        try:
            theme = open(user_data.THEME_FILE, encoding="utf-8").read().strip().lower()
            if theme in THEME_PALETTES:
                return theme
        except OSError:
            pass
    return "light"


def save_theme(theme: str) -> None:
    try:
        with open(user_data.THEME_FILE, "w", encoding="utf-8") as f:
            f.write(theme)
    except OSError:
        pass


def load_api_key() -> str:
    if os.path.exists(user_data.CONFIG_FILE):
        try:
            with open(user_data.CONFIG_FILE, encoding="utf-8") as f:
                return f.read().strip()
        except OSError:
            pass
    return ""


def load_rules_path() -> str:
    if os.path.exists(user_data.RULES_PATH_FILE):
        try:
            with open(user_data.RULES_PATH_FILE, encoding="utf-8") as f:
                return f.read().strip()
        except OSError:
            pass
    return ""


def save_rules_path(path: str) -> None:
    try:
        with open(user_data.RULES_PATH_FILE, "w", encoding="utf-8") as f:
            f.write(path)
    except OSError:
        pass


def _rules_table_to_text(df: pd.DataFrame) -> str:
    lines = [
        "ПРАВИЛА КОМПАНИИ (таблица; каждая строка — категория / маршрутизация):",
        "",
    ]
    for idx, row in df.iterrows():
        parts: list[str] = []
        for col in df.columns:
            val = row[col]
            if pd.isna(val) or not str(val).strip():
                continue
            parts.append(f"{col}: {str(val).strip()}")
        if parts:
            lines.append(f"{int(idx) + 1}. " + " | ".join(parts))
    lines.append(RULES_TABLE_FOOTER)
    return "\n".join(lines).strip()


def load_company_rules() -> str:
    path = load_rules_path()
    if not path or not os.path.exists(path):
        return ""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".txt", ".md"):
            with open(path, encoding="utf-8") as f:
                return f.read().strip()
        if ext == ".csv":
            return _rules_table_to_text(pd.read_csv(path))
        if ext in (".xlsx", ".xlsm"):
            return _rules_table_to_text(pd.read_excel(path, engine="openpyxl"))
    except (OSError, ValueError, pd.errors.ParserError):
        pass
    return ""


def save_api_key(key: str) -> None:
    try:
        with open(user_data.CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(key.strip())
    except OSError:
        pass


def read_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return pd.read_excel(path, engine="openpyxl")
    if ext == ".csv":
        return pd.read_csv(path)
    if ext in (".txt", ".tsv"):
        with open(path, encoding="utf-8") as f:
            text = f.read()
        sep = "\t" if ext == ".tsv" else _detect_csv_separator(text[:8192])
        return pd.read_csv(StringIO(text), sep=sep, engine="python")
    raise ValueError("bad_format")


def extract_llm_csv(text: str) -> str:
    """Вырезает CSV из ответа ИИ (без markdown и пояснений до/после)."""
    raw = text.strip().lstrip("\ufeff")
    if "```" in raw:
        blocks: list[str] = []
        for block in raw.split("```"):
            block = block.strip()
            if block.lower().startswith("csv"):
                block = block[3:].lstrip()
            if block:
                blocks.append(block)
        if blocks:
            raw = max(blocks, key=len)

    lines = raw.splitlines()
    csv_lines: list[str] = []
    started = False
    for line in lines:
        stripped = line.strip()
        if not stripped:
            if started:
                csv_lines.append(line)
            continue
        has_sep = "," in line or ";" in line or "\t" in line
        if not started:
            if has_sep:
                started = True
                csv_lines.append(line)
            continue
        if has_sep:
            csv_lines.append(line)
        elif started:
            break
    return "\n".join(csv_lines) if csv_lines else raw


def _detect_csv_separator(text: str) -> str:
    try:
        return csv.Sniffer().sniff(text[:8192], delimiters=",;\t").delimiter
    except csv.Error:
        header = text.splitlines()[0] if text else ""
        if header.count(";") > header.count(","):
            return ";"
        if "\t" in header:
            return "\t"
        return ","


def parse_csv_text(csv_text: str) -> pd.DataFrame:
    text = extract_llm_csv(csv_text)
    if not text.strip():
        raise ValueError("empty_csv")

    separators = []
    sniffed = _detect_csv_separator(text)
    separators.append(sniffed)
    for sep in (",", ";", "\t"):
        if sep not in separators:
            separators.append(sep)

    last_err: Exception | None = None
    for sep in separators:
        for kwargs in (
            {"sep": sep, "engine": "python"},
            {"sep": sep, "engine": "python", "on_bad_lines": "skip"},
        ):
            try:
                df = pd.read_csv(StringIO(text), **kwargs)
                if len(df.columns) >= 2 or (len(df.columns) == 1 and len(df) > 0):
                    return df
            except Exception as exc:
                last_err = exc
    raise ValueError(str(last_err) if last_err else "parse_failed")


def _measure_cell_width(value, cap: int = 42) -> int:
    """Сколько символов учитывать при расчёте ширины (без длинных хвостов)."""
    if pd.isna(value):
        return 0
    text = str(value).strip()
    if not text:
        return 0
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    longest_line = max((len(part) for part in text.split("\n")), default=0)
    return min(longest_line, cap)


def _fit_column_width(series: pd.Series, header: str) -> float:
    """Ширина колонки под типичное содержимое + капля запаса."""
    header_len = len(str(header).strip())
    measured = [_measure_cell_width(v) for v in series]
    measured = [m for m in measured if m > 0]

    if measured:
        measured.sort()
        p_idx = min(len(measured) - 1, int(len(measured) * 0.88))
        body_len = measured[p_idx]
    else:
        body_len = 0

    width = max(header_len, body_len) + 1.2
    return max(6.0, min(width, 50.0))


def _write_formatted_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """Записывает лист Excel с переносом текста и шириной колонок под содержимое."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wrap = Alignment(wrap_text=True, vertical="top")
    header_align = Alignment(wrap_text=True, vertical="center")

    df.to_excel(writer, index=False, sheet_name=sheet_name)
    sheet = writer.sheets[sheet_name]
    for col_idx, column in enumerate(df.columns, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = _fit_column_width(
            df[column], str(column),
        )
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = header_align if cell.row == 1 else wrap


def save_dataframe_excel(df: pd.DataFrame, path: str) -> None:
    """Сохраняет Excel: ширина каждой колонки под свой текст, длинное переносится."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_formatted_sheet(writer, "Sheet1", df)


ROUTING_DEPT_NAMES = ("Отдел", "Department", "Abteilung")
ROUTING_PRIO_NAMES = ("Приоритет", "Priority", "Priorität")
ROUTING_CAT_NAMES = ("Категория", "Category", "Kategorie")
REVIEW_MARKERS = (
    "требует проверки", "needs review", "prüfung erforderlich", "prüfung nötig",
)
HIGH_PRIORITY_MARKERS = ("высокий", "high", "hoch")
ALLOWED_DEPT_MARKERS = (
    "РАЗРЕШЁННЫЕ ОТДЕЛЫ",
    "ALLOWED DEPARTMENTS",
    "ZULÄSSIGE ABTEILUNGEN",
)
FILTER_ALL = "__all__"
FILTER_HIGH = "__high__"
FILTER_REVIEW = "__review__"
FILTER_UNKNOWN = "__unknown__"

INBOX_ROOT = os.path.expanduser("~/Desktop/SmartSorter_Почта_отделов")
DEPARTMENT_INBOXES_FILE = "department_folders.demo.txt"
INBOX_SPECIAL_ALIASES = (
    "на проверку", "needs review", "prüfung", "проверить_отдел",
    "check_department", "abteilung_prüfen", "сводка", "summary",
)


def detect_complaint_text_column(df: pd.DataFrame) -> str | None:
    found = find_routing_column(df, COMPLAINT_TEXT_NAMES)
    if found:
        return found
    best_col: str | None = None
    best_avg = 0.0
    for col in df.columns:
        lengths = [
            len(str(v).strip())
            for v in df[col]
            if not pd.isna(v) and str(v).strip()
        ]
        if not lengths:
            continue
        avg = sum(lengths) / len(lengths)
        if avg > best_avg:
            best_avg = avg
            best_col = str(col)
    return best_col if best_avg >= 12 else None


def build_request_context(df: pd.DataFrame, file_path: str) -> str:
    """Контекст таблицы — ИИ видит структуру и колонку с текстом жалобы."""
    lines = [
        "TABLE METADATA:",
        f"- File: {os.path.basename(file_path)}",
        f"- Data rows: {len(df)}",
        f"- Columns ({len(df.columns)}): {', '.join(str(c) for c in df.columns)}",
    ]
    text_col = detect_complaint_text_column(df)
    if text_col:
        lines.append(f"- Complaint text column: «{text_col}» — read this cell IN FULL for each row.")
    lines.extend((
        "- Output columns and routing logic come ONLY from <company_rules>.",
        "- Classify by meaning of the whole complaint, not isolated words.",
        "- Apply <company_rules> together with this context.",
    ))
    return "\n".join(lines)


def build_api_user_message(
    csv_data: str,
    task: str,
    *,
    file_path: str,
    df: pd.DataFrame,
    company_rules: str = "",
) -> str:
    """Собирает запрос: контекст → правила → данные → задача."""
    parts = [
        f"<context>\n{build_request_context(df, file_path)}\n</context>\n",
    ]
    if company_rules:
        parts.append(f"<company_rules>\n{company_rules}\n</company_rules>\n")
    parts.append(f"<source_csv>\n{csv_data}\n</source_csv>\n")
    parts.append(f"<user_task>\n{task}\n</user_task>")
    return "\n".join(parts)


def build_system_prompt(company_rules: str = "") -> str:
    if company_rules:
        return SYSTEM_PROMPT + RULES_SYSTEM_APPEND
    return SYSTEM_PROMPT


def find_routing_column(df: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    for col in df.columns:
        cl = str(col).strip().lower()
        for cand in candidates:
            c = cand.lower()
            if cl == c or c in cl:
                return col
    return None


def _is_review_value(value) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    return any(marker in text for marker in REVIEW_MARKERS)


def _is_high_priority_value(value) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    return any(text == m or text.startswith(m + " ") for m in HIGH_PRIORITY_MARKERS)


def _safe_excel_sheet_name(name: str, used: set[str]) -> str:
    raw = str(name).strip() if str(name).strip() else "?"
    for ch in (":", "\\", "/", "?", "*", "[", "]"):
        raw = raw.replace(ch, "_")
    base = raw[:31]
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(candidate)
    return candidate


def build_routing_summary(
    df: pd.DataFrame,
    dept_col: str,
    prio_col: str | None,
    cat_col: str | None,
    labels: dict[str, str],
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    grouped = df.groupby(dept_col, dropna=False)
    for dept, group in grouped:
        dept_name = str(dept).strip() if not pd.isna(dept) else labels["unknown"]
        high = (
            int(group[prio_col].map(_is_high_priority_value).sum())
            if prio_col
            else 0
        )
        review = (
            int(group[cat_col].map(_is_review_value).sum())
            if cat_col
            else 0
        )
        rows.append({
            labels["dept"]: dept_name,
            labels["total"]: len(group),
            labels["high"]: high,
            labels["review"]: review,
        })
    summary = pd.DataFrame(rows)
    if not summary.empty:
        summary = summary.sort_values(labels["total"], ascending=False)
    return summary


@dataclass
class RoutingExportStats:
    total: int
    by_department: dict[str, int]
    review: int
    unknown_dept: int
    unknown_names: list[str]


@dataclass
class DepartmentInbox:
    name: str
    folder: str
    email: str = ""


def _norm_dept_name(name: str) -> str:
    return str(name).strip().casefold()


def _safe_file_stem(name: str) -> str:
    stem = str(name).strip() or "unnamed"
    for ch in (":", "\\", "/", "?", "*", "[", "]", '"', "<", ">", "|"):
        stem = stem.replace(ch, "_")
    return stem[:80]


def _looks_like_rules_section_header(stripped: str) -> bool:
    if stripped.startswith("- ") or stripped.startswith("#"):
        return False
    head = stripped.split("(")[0].split(":")[0].strip()
    if len(head) < 3:
        return False
    alpha = [c for c in head if c.isalpha()]
    if not alpha:
        return False
    upper = sum(
        1 for c in alpha
        if c.isupper() or c in "АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ"
    )
    return upper / len(alpha) >= 0.75


def parse_allowed_departments(rules_text: str) -> list[str]:
    """Список отделов из блока РАЗРЕШЁННЫЕ ОТДЕЛЫ в файле правил."""
    if not rules_text.strip():
        return []
    lines = rules_text.splitlines()
    in_section = False
    depts: list[str] = []
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        upper = stripped.upper()
        if any(marker in upper for marker in ALLOWED_DEPT_MARKERS):
            in_section = True
            continue
        if in_section:
            if _looks_like_rules_section_header(stripped):
                break
            numbered = re.match(r"^[\d]+[.)]\s*(.+)$", stripped)
            if numbered:
                dept = numbered.group(1).strip()
                if len(dept) <= 80 and "→" not in dept:
                    depts.append(dept)
                continue
            if stripped.startswith("- "):
                dept = stripped[2:].strip()
                if len(dept) <= 80 and "→" not in dept:
                    depts.append(dept)
    if depts:
        return depts
    for line in lines:
        match = re.search(
            r"(?:Отдел|Department|Abteilung):\s*(.+?)(?:\s*\||$)",
            line,
            re.IGNORECASE,
        )
        if match:
            dept = match.group(1).strip()
            if dept and dept not in depts:
                depts.append(dept)
    return depts


def department_inboxes_config_path() -> str:
    return os.path.join(APP_DIR, DEPARTMENT_INBOXES_FILE)


def _default_department_inboxes() -> dict[str, DepartmentInbox]:
    """Папки отделов на рабочем столе — используются, если нет конфиг-файла."""
    pairs = (
        ("Служба поддержки", "support@demo.local"),
        ("Бухгалтерия", "buchhaltung@demo.local"),
        ("Документооборот", "docs@demo.local"),
        ("Контроль качества", "qa@demo.local"),
        ("Руководитель смены", "shift-lead@demo.local"),
        ("IT-поддержка", "it-helpdesk@demo.local"),
        ("На проверку", ""),
        ("Сводка", ""),
    )
    result: dict[str, DepartmentInbox] = {}
    for name, email in pairs:
        folder = INBOX_ROOT if name in ("На проверку", "Сводка") else os.path.join(INBOX_ROOT, name)
        if name == "На проверку":
            folder = os.path.join(INBOX_ROOT, "_На_проверку")
        elif name == "Сводка":
            folder = os.path.join(INBOX_ROOT, "_Сводка")
        result[name.casefold()] = DepartmentInbox(name=name, folder=folder, email=email)
    return result


def _parse_department_inboxes_text(text: str) -> dict[str, DepartmentInbox]:
    inboxes: dict[str, DepartmentInbox] = {}
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or not stripped.startswith("- "):
            continue
        parts = [p.strip() for p in stripped[2:].strip().split("|")]
        if len(parts) < 2:
            continue
        name, folder = parts[0], os.path.expanduser(parts[1])
        email = parts[2] if len(parts) > 2 else ""
        inboxes[name.casefold()] = DepartmentInbox(name=name, folder=folder, email=email)
    return inboxes


def department_inboxes_configured() -> bool:
    """True, если есть файл с путями отделов (не встроенные папки по умолчанию)."""
    path = department_inboxes_config_path()
    if not os.path.exists(path):
        return False
    try:
        with open(path, encoding="utf-8") as cfg:
            return bool(_parse_department_inboxes_text(cfg.read()))
    except OSError:
        return False


def load_department_inboxes() -> dict[str, DepartmentInbox]:
    """Читает department_folders.demo.txt; без файла — пусто (не гадаем пути)."""
    path = department_inboxes_config_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as cfg:
            return _parse_department_inboxes_text(cfg.read())
    except OSError:
        return {}


def _find_department_inbox(
    name: str,
    inboxes: dict[str, DepartmentInbox],
) -> DepartmentInbox | None:
    key = re.sub(r"^00_", "", name.strip()).casefold()
    if key in inboxes:
        return inboxes[key]
    for alias in INBOX_SPECIAL_ALIASES:
        if alias in key:
            for special in ("на проверку", "сводка"):
                if special in alias and special in inboxes:
                    return inboxes[special]
    for inbox_key, inbox in inboxes.items():
        if inbox_key in key or key in inbox_key:
            return inbox
    return None


def load_allowed_departments() -> list[str]:
    rules_text = load_company_rules()
    depts = parse_allowed_departments(rules_text)
    if depts:
        return depts
    path = load_rules_path()
    if path and os.path.splitext(path)[1].lower() == ".csv":
        try:
            rules_df = pd.read_csv(path)
            dept_col = find_routing_column(rules_df, ROUTING_DEPT_NAMES)
            if dept_col:
                return sorted({
                    str(v).strip()
                    for v in rules_df[dept_col]
                    if not pd.isna(v) and str(v).strip()
                })
        except (OSError, ValueError, pd.errors.ParserError):
            pass
    return []


def _unknown_department_mask(
    series: pd.Series,
    allowed_departments: list[str],
) -> pd.Series:
    if not allowed_departments:
        return pd.Series([False] * len(series), index=series.index)
    allowed = {_norm_dept_name(d) for d in allowed_departments}

    def is_unknown(value) -> bool:
        if pd.isna(value) or not str(value).strip():
            return True
        return _norm_dept_name(str(value)) not in allowed

    return series.map(is_unknown)


def _routing_slices(
    df: pd.DataFrame,
    allowed_departments: list[str] | None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    dept_col = find_routing_column(df, ROUTING_DEPT_NAMES)
    if not dept_col:
        raise ValueError("no_dept")
    allowed = allowed_departments or []
    cat_col = find_routing_column(df, ROUTING_CAT_NAMES)
    unknown_mask = _unknown_department_mask(df[dept_col], allowed)
    unknown_df = df[unknown_mask].copy()
    known_df = df[~unknown_mask].copy()
    review_df = (
        df[df[cat_col].map(_is_review_value)].copy()
        if cat_col
        else df.iloc[0:0].copy()
    )
    return known_df, unknown_df, review_df, dept_col


def _routing_export_stats(
    df: pd.DataFrame,
    dept_col: str,
    unknown_df: pd.DataFrame,
    review_df: pd.DataFrame,
) -> RoutingExportStats:
    by_department: dict[str, int] = {}
    for dept, group in df.groupby(dept_col, dropna=False):
        label = str(dept).strip() if not pd.isna(dept) else "?"
        by_department[label] = len(group)
    unknown_names = sorted({
        str(v).strip()
        for v in unknown_df[dept_col]
        if not pd.isna(v) and str(v).strip()
    })
    return RoutingExportStats(
        total=len(df),
        by_department=by_department,
        review=len(review_df),
        unknown_dept=len(unknown_df),
        unknown_names=unknown_names,
    )


def write_routing_workbook(
    writer,
    df: pd.DataFrame,
    labels: dict[str, str],
    allowed_departments: list[str] | None = None,
) -> RoutingExportStats:
    """Пишет сводку, проверку, неизвестные отделы и листы по отделам."""
    allowed = allowed_departments or []
    known_df, unknown_df, review_df, dept_col = _routing_slices(df, allowed)
    prio_col = find_routing_column(df, ROUTING_PRIO_NAMES)
    cat_col = find_routing_column(df, ROUTING_CAT_NAMES)
    used_sheets: set[str] = set()

    summary_df = build_routing_summary(df, dept_col, prio_col, cat_col, labels)
    _write_formatted_sheet(
        writer,
        _safe_excel_sheet_name(labels["sheet_summary"], used_sheets),
        summary_df,
    )
    if not review_df.empty:
        _write_formatted_sheet(
            writer,
            _safe_excel_sheet_name(labels["sheet_review"], used_sheets),
            review_df.reset_index(drop=True),
        )
    if not unknown_df.empty:
        _write_formatted_sheet(
            writer,
            _safe_excel_sheet_name(labels["sheet_unknown_dept"], used_sheets),
            unknown_df.reset_index(drop=True),
        )
    for dept, group in grouped_sorted(known_df, dept_col):
        dept_label = str(dept).strip() if not pd.isna(dept) else labels["unknown"]
        _write_formatted_sheet(
            writer,
            _safe_excel_sheet_name(dept_label, used_sheets),
            group.reset_index(drop=True),
        )
    return _routing_export_stats(df, dept_col, unknown_df, review_df)


def save_routing_export(
    df: pd.DataFrame,
    path: str,
    labels: dict[str, str],
    allowed_departments: list[str] | None = None,
) -> RoutingExportStats:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        return write_routing_workbook(writer, df, labels, allowed_departments)


def build_routing_log(stats: RoutingExportStats, labels: dict[str, str]) -> str:
    lines = [
        labels["log_title"],
        f"{labels['log_date']}: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"{labels['log_total']}: {stats.total}",
        f"{labels['log_review']}: {stats.review}",
        f"{labels['log_unknown']}: {stats.unknown_dept}",
        "",
        labels["log_by_dept"],
    ]
    for dept, count in sorted(stats.by_department.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"  - {dept}: {count}")
    if stats.unknown_names:
        lines.extend(("", labels["log_unknown_list"],))
        for name in stats.unknown_names:
            lines.append(f"  - {name}")
    return "\n".join(lines)


def build_routing_review_text(
    df: pd.DataFrame,
    labels: dict[str, str],
    allowed_departments: list[str] | None,
    intro: str,
) -> str:
    allowed = allowed_departments or []
    known_df, unknown_df, review_df, dept_col = _routing_slices(df, allowed)
    stats = _routing_export_stats(df, dept_col, unknown_df, review_df)
    lines = [intro, ""]
    lines.append(f"{labels['log_total']}: {stats.total}")
    lines.append(f"{labels['log_review']}: {stats.review}")
    lines.append(f"{labels['log_unknown']}: {stats.unknown_dept}")
    lines.extend(("", labels["log_by_dept"]))
    for dept, count in sorted(stats.by_department.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"  • {dept}: {count}")
    if stats.unknown_names:
        lines.extend(("", labels["log_unknown_list"]))
        for name in stats.unknown_names:
            lines.append(f"  • {name}")
    if stats.review > 0:
        lines.extend(("", f"→ {labels['sheet_review']}: {stats.review}"))
    return "\n".join(lines)


def export_routing_folder(
    df: pd.DataFrame,
    parent_dir: str,
    labels: dict[str, str],
    allowed_departments: list[str] | None = None,
) -> tuple[str, RoutingExportStats]:
    """Создаёт папку с файлами по отделам, сводкой и log.txt."""
    allowed = allowed_departments or []
    known_df, unknown_df, review_df, dept_col = _routing_slices(df, allowed)
    prio_col = find_routing_column(df, ROUTING_PRIO_NAMES)
    cat_col = find_routing_column(df, ROUTING_CAT_NAMES)

    folder = os.path.join(
        parent_dir,
        f"routing_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}",
    )
    os.makedirs(folder, exist_ok=True)

    summary_df = build_routing_summary(df, dept_col, prio_col, cat_col, labels)
    save_dataframe_excel(
        summary_df,
        os.path.join(folder, f"00_{_safe_file_stem(labels['sheet_summary'])}.xlsx"),
    )
    if not review_df.empty:
        review_stem = _safe_file_stem(labels["sheet_review"])
        review_data = review_df.reset_index(drop=True)
        save_dataframe_excel(
            review_data,
            os.path.join(folder, f"{review_stem}.xlsx"),
        )
        _save_complaints_readable(
            os.path.join(folder, f"{review_stem}_прочитать.txt"),
            review_data,
            f"Жалобы — {labels['sheet_review']}",
        )
    if not unknown_df.empty:
        unknown_stem = _safe_file_stem(labels["sheet_unknown_dept"])
        unknown_data = unknown_df.reset_index(drop=True)
        save_dataframe_excel(
            unknown_data,
            os.path.join(folder, f"{unknown_stem}.xlsx"),
        )
        _save_complaints_readable(
            os.path.join(folder, f"{unknown_stem}_прочитать.txt"),
            unknown_data,
            f"Жалобы — {labels['sheet_unknown_dept']}",
        )
    for dept, group in grouped_sorted(known_df, dept_col):
        dept_label = str(dept).strip() if not pd.isna(dept) else labels["unknown"]
        dept_stem = _safe_file_stem(dept_label)
        dept_data = group.reset_index(drop=True)
        save_dataframe_excel(
            dept_data,
            os.path.join(folder, f"{dept_stem}.xlsx"),
        )
        _save_complaints_readable(
            os.path.join(folder, f"{dept_stem}_прочитать.txt"),
            dept_data,
            f"Жалобы — {dept_label}",
        )

    stats = _routing_export_stats(df, dept_col, unknown_df, review_df)
    log_path = os.path.join(folder, "log.txt")
    with open(log_path, "w", encoding="utf-8") as log_file:
        log_file.write(build_routing_log(stats, labels))
    return folder, stats


def _write_inbox_letter(
    path: str,
    inbox: DepartmentInbox,
    body: str,
    xlsx_name: str,
    readable_name: str | None,
) -> None:
    header = f"Кому: {inbox.email or '(email не указан)'}\n"
    header += f"Папка: {inbox.folder}\n"
    header += f"Таблица: {xlsx_name}\n"
    if readable_name:
        header += f"Прочитать тексты: {readable_name}\n"
    header += "-" * 50 + "\n\n"
    with open(path, "w", encoding="utf-8") as out:
        out.write(header + body)


def _deliver_dataframe_to_inbox(
    data: pd.DataFrame,
    inbox_key: str,
    title: str,
    email_dept: str,
    inboxes: dict[str, DepartmentInbox],
    labels: dict[str, str],
    stamp: str,
) -> str | None:
    if data.empty:
        return None
    inbox = _find_department_inbox(inbox_key, inboxes)
    if inbox is None:
        return f"⚠ нет папки для: {title}"
    os.makedirs(inbox.folder, exist_ok=True)
    stem = _safe_file_stem(title)
    subset = data.reset_index(drop=True)
    xlsx_name = f"{stamp}_{stem}.xlsx"
    readable_name = f"жалобы_прочитать_{stamp}.txt"
    save_dataframe_excel(subset, os.path.join(inbox.folder, xlsx_name))
    _save_complaints_readable(
        os.path.join(inbox.folder, readable_name),
        subset,
        f"Жалобы — {title}",
    )
    dept_col = find_routing_column(subset, ROUTING_DEPT_NAMES)
    if dept_col and email_dept in subset[dept_col].astype(str).str.strip().values:
        body = build_department_email(subset, email_dept, labels)
    else:
        date = datetime.now().strftime("%d.%m.%Y")
        body = (
            f"Тема: Жалобы — {title} — {date}\n\n"
            f"Здравствуйте!\n\n"
            f"Файлы для отдела «{title}» лежат в этой папке.\n"
            f"Откройте «{readable_name}», чтобы прочитать тексты жалоб.\n\n"
            f"С уважением"
        )
    _write_inbox_letter(
        os.path.join(inbox.folder, f"письмо_{stamp}.txt"),
        inbox,
        body,
        xlsx_name,
        readable_name,
    )
    return f"✓ {title} → {inbox.folder}"


def deliver_to_department_inboxes(
    df: pd.DataFrame,
    labels: dict[str, str],
    allowed_departments: list[str] | None = None,
) -> tuple[list[str], str]:
    """Кладёт Excel, тексты жалоб и письмо в папки отделов (без выбора папки)."""
    inboxes = load_department_inboxes()
    if not inboxes:
        raise ValueError("no_inboxes")
    allowed = allowed_departments or []
    known_df, unknown_df, review_df, dept_col = _routing_slices(df, allowed)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    log: list[str] = []

    if not review_df.empty:
        line = _deliver_dataframe_to_inbox(
            review_df,
            labels["sheet_review"],
            labels["sheet_review"],
            labels["sheet_review"],
            inboxes,
            labels,
            stamp,
        )
        if line:
            log.append(line)

    if not unknown_df.empty:
        line = _deliver_dataframe_to_inbox(
            unknown_df,
            labels["sheet_unknown_dept"],
            labels["sheet_unknown_dept"],
            labels["sheet_unknown_dept"],
            inboxes,
            labels,
            stamp,
        )
        if line:
            log.append(line)

    for dept, group in grouped_sorted(known_df, dept_col):
        dept_label = str(dept).strip() if not pd.isna(dept) else labels["unknown"]
        line = _deliver_dataframe_to_inbox(
            group,
            dept_label,
            dept_label,
            dept_label,
            inboxes,
            labels,
            stamp,
        )
        if line:
            log.append(line)

    if not log:
        raise ValueError("nothing_delivered")
    return log, INBOX_ROOT


def filter_draft_dataframe(
    df: pd.DataFrame,
    mode: str,
    allowed_departments: list[str] | None = None,
) -> pd.DataFrame:
    dept_col = find_routing_column(df, ROUTING_DEPT_NAMES)
    prio_col = find_routing_column(df, ROUTING_PRIO_NAMES)
    cat_col = find_routing_column(df, ROUTING_CAT_NAMES)
    allowed = allowed_departments or []

    if mode == FILTER_ALL:
        return df
    if mode == FILTER_HIGH and prio_col:
        return df[df[prio_col].map(_is_high_priority_value)].copy()
    if mode == FILTER_REVIEW and cat_col:
        return df[df[cat_col].map(_is_review_value)].copy()
    if mode == FILTER_UNKNOWN and dept_col:
        return df[_unknown_department_mask(df[dept_col], allowed)].copy()
    if dept_col and mode != FILTER_ALL:
        mask = df[dept_col].astype(str).str.strip() == str(mode).strip()
        return df[mask].copy()
    return df


def build_complaints_readable_text(
    df: pd.DataFrame,
    title: str = "",
) -> str:
    """Человекочитаемый текст жалоб — для .txt рядом с экспортом по отделам."""
    if df.empty:
        return f"{title}\n\n(нет обращений)\n" if title else "(нет обращений)\n"

    text_col = detect_complaint_text_column(df)
    cat_col = find_routing_column(df, ROUTING_CAT_NAMES)
    prio_col = find_routing_column(df, ROUTING_PRIO_NAMES)
    reason_col = find_routing_column(
        df, ("Причина_выбора_категории", "Reason", "Grund"),
    )
    summary_col = find_routing_column(
        df, ("Краткое_резюме", "Summary", "Kurzfassung"),
    )
    meta_cols: list[str] = []
    for names in (
        ("Дата", "Date", "Datum"),
        ("ФИО", "Name", "Имя"),
        ("Email", "E-mail", "Почта"),
    ):
        found = find_routing_column(df, names)
        if found and found not in meta_cols:
            meta_cols.append(found)

    lines: list[str] = []
    if title:
        lines.extend((title, "=" * min(len(title), 60), ""))
    lines.append(f"Всего обращений: {len(df)}")
    lines.append(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    lines.append("")

    for i, (_, row) in enumerate(df.iterrows(), 1):
        lines.append("─" * 50)
        lines.append(f"Обращение {i}")
        lines.append("─" * 50)
        for col in meta_cols:
            val = row.get(col)
            if not pd.isna(val) and str(val).strip():
                lines.append(f"{col}: {val}")
        if cat_col:
            val = row.get(cat_col)
            if not pd.isna(val) and str(val).strip():
                lines.append(f"Категория: {val}")
        if prio_col:
            val = row.get(prio_col)
            if not pd.isna(val) and str(val).strip():
                lines.append(f"Приоритет: {val}")
        if summary_col:
            val = row.get(summary_col)
            if not pd.isna(val) and str(val).strip():
                lines.append(f"Кратко: {val}")
        lines.append("")
        if text_col:
            val = row.get(text_col)
            if not pd.isna(val) and str(val).strip():
                lines.append("Текст жалобы:")
                lines.append(str(val).strip())
            else:
                lines.append("Текст жалобы: (не указан)")
        else:
            lines.append("Текст жалобы: (колонка не найдена)")
        if reason_col:
            val = row.get(reason_col)
            if not pd.isna(val) and str(val).strip():
                lines.append("")
                lines.append(f"Почему такая категория: {val}")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def _save_complaints_readable(path: str, df: pd.DataFrame, title: str) -> None:
    with open(path, "w", encoding="utf-8") as out:
        out.write(build_complaints_readable_text(df, title))


def build_department_email(
    df: pd.DataFrame,
    department: str,
    labels: dict[str, str],
) -> str:
    dept_col = find_routing_column(df, ROUTING_DEPT_NAMES)
    prio_col = find_routing_column(df, ROUTING_PRIO_NAMES)
    cat_col = find_routing_column(df, ROUTING_CAT_NAMES)
    if not dept_col:
        return ""
    subset = df[df[dept_col].astype(str).str.strip() == department.strip()].copy()
    high = (
        int(subset[prio_col].map(_is_high_priority_value).sum())
        if prio_col
        else 0
    )
    review = (
        int(subset[cat_col].map(_is_review_value).sum())
        if cat_col
        else 0
    )
    return labels["email_body"].format(
        department=department,
        count=len(subset),
        high=high,
        review=review,
        date=datetime.now().strftime("%d.%m.%Y"),
        attach_hint=labels["email_attach"],
    )


def grouped_sorted(df: pd.DataFrame, dept_col: str):
    groups = list(df.groupby(dept_col, dropna=False))
    groups.sort(key=lambda item: str(item[0]).strip().lower() if not pd.isna(item[0]) else "")
    return groups


def _cell_key(value) -> str:
    """Нормализованное значение ячейки для сравнения строк."""
    if pd.isna(value):
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return format(value, ".15g")
    text = str(value).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    try:
        num = float(text.replace(",", "."))
        if num == int(num):
            return str(int(num))
        return format(num, ".15g")
    except ValueError:
        return text


def _cell_str(value) -> str:
    text = _cell_key(value)
    return text if text else "—"


@dataclass
class TableChanges:
    structure_changed: bool
    rows_before: int
    rows_after: int
    changed_cells: list[tuple[int, str, str, str]]
    removed_rows: list[tuple[int, str]]
    added_rows: list[tuple[int, str]]


def _common_columns(left: pd.DataFrame, right: pd.DataFrame) -> list:
    right_cols = set(right.columns)
    return [col for col in left.columns if col in right_cols]


def _row_tuple(df: pd.DataFrame, idx: int, columns: list) -> tuple[str, ...]:
    return tuple(_cell_key(df.iloc[idx][col]) for col in columns)


def _row_preview(
    df: pd.DataFrame,
    idx: int,
    columns: list | None = None,
    max_cols: int = 4,
    max_len: int = 90,
) -> str:
    cols = list(columns or df.columns)[:max_cols]
    parts = [_cell_str(df.iloc[idx][col]) for col in cols]
    text = " · ".join(parts)
    total_cols = len(columns or df.columns)
    extra = total_cols - len(cols)
    if extra > 0:
        text += f" · (+{extra})"
    if len(text) > max_len:
        return text[: max_len - 1] + "…"
    return text


def _rows_only_in_left(
    left: pd.DataFrame,
    right: pd.DataFrame,
    columns: list,
) -> list[tuple[int, str]]:
    if not columns:
        return []
    pool = Counter(_row_tuple(right, i, columns) for i in range(len(right)))
    unmatched: list[tuple[int, str]] = []
    for i in range(len(left)):
        sig = _row_tuple(left, i, columns)
        if pool[sig] > 0:
            pool[sig] -= 1
        else:
            unmatched.append((i + 2, _row_preview(left, i, columns)))
    return unmatched


def _effective_row_counts(changes: TableChanges) -> tuple[int, int]:
    removed = len(changes.removed_rows)
    added = len(changes.added_rows)
    net_removed = max(0, changes.rows_before - changes.rows_after)
    net_added = max(0, changes.rows_after - changes.rows_before)
    if removed == 0 and net_removed > 0:
        removed = net_removed
    if added == 0 and net_added > 0:
        added = net_added
    return removed, added


def compare_tables(original: pd.DataFrame, draft: pd.DataFrame) -> TableChanges:
    rows_before = len(original)
    rows_after = len(draft)
    structure_changed = list(original.columns) != list(draft.columns)
    changed_cells: list[tuple[int, str, str, str]] = []
    common_cols = _common_columns(original, draft)

    removed_rows = _rows_only_in_left(original, draft, common_cols)
    added_rows = _rows_only_in_left(draft, original, common_cols)

    same_shape = (
        list(original.columns) == list(draft.columns)
        and rows_before == rows_after
        and not removed_rows
        and not added_rows
    )
    if same_shape:
        for row_idx in range(rows_before):
            for col in original.columns:
                old_val = _cell_str(original.iloc[row_idx][col])
                new_val = _cell_str(draft.iloc[row_idx][col])
                if old_val != new_val:
                    changed_cells.append(
                        (row_idx + 2, str(col), old_val, new_val),
                    )

    return TableChanges(
        structure_changed=structure_changed,
        rows_before=rows_before,
        rows_after=rows_after,
        changed_cells=changed_cells,
        removed_rows=removed_rows,
        added_rows=added_rows,
    )


class WorkReportDialog(QDialog):
    """Окно с подробным отчётом о том, что изменил ассистент."""

    def __init__(
        self,
        parent: QWidget,
        title: str,
        summary: str,
        details: str,
        close_text: str,
        configure_edit,
    ):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumSize(620, 480)
        self.resize(720, 560)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 16)
        lay.setSpacing(12)

        self.lbl_title = QLabel(title)
        self.lbl_title.setObjectName("stats_title")
        lay.addWidget(self.lbl_title)

        self.lbl_summary = QLabel(summary)
        self.lbl_summary.setObjectName("changes_summary")
        self.lbl_summary.setWordWrap(True)
        lay.addWidget(self.lbl_summary)

        self.text_details = QTextEdit()
        configure_edit(self.text_details, "output", min_height=280, read_only=True)
        self.text_details.setPlainText(details)
        lay.addWidget(self.text_details, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_close = QPushButton(close_text)
        btn_close.setObjectName("primary")
        btn_close.setMinimumWidth(120)
        btn_close.clicked.connect(self.accept)
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)


class HistoryDialog(QDialog):
    """Просмотр локального архива черновиков."""

    def __init__(self, parent: "MainWindow", configure_edit) -> None:
        super().__init__(parent)
        self._main = parent
        self._configure_edit = configure_edit
        self._records: list[HistoryRecord] = []
        self._current: HistoryRecord | None = None

        self.setWindowTitle(parent.tr("history_title"))
        self.setMinimumSize(760, 480)
        self.resize(900, 560)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 16)
        lay.setSpacing(12)

        body = QHBoxLayout()
        body.setSpacing(14)
        self.list_entries = QListWidget()
        self.list_entries.setMinimumWidth(240)
        self.list_entries.setMaximumWidth(300)
        self.list_entries.currentItemChanged.connect(self._on_pick)
        body.addWidget(self.list_entries)

        right = QVBoxLayout()
        right.setSpacing(8)
        self.lbl_summary = QLabel()
        self.lbl_summary.setObjectName("changes_summary")
        self.lbl_summary.setWordWrap(True)
        right.addWidget(self.lbl_summary)

        self.lbl_task = QLabel()
        self.lbl_task.setObjectName("file_empty")
        self.lbl_task.setWordWrap(True)
        right.addWidget(self.lbl_task)

        self.text_preview = QTextEdit()
        self._configure_edit(self.text_preview, "output", min_height=280, read_only=True)
        right.addWidget(self.text_preview, stretch=1)
        body.addLayout(right, stretch=1)
        lay.addLayout(body, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        self.btn_open = QPushButton()
        self.btn_open.setObjectName("primary")
        self.btn_open.clicked.connect(self._open_in_work)
        self.btn_csv = QPushButton()
        self.btn_csv.setObjectName("secondary")
        self.btn_csv.clicked.connect(self._download_csv)
        self.btn_xlsx = QPushButton()
        self.btn_xlsx.setObjectName("secondary")
        self.btn_xlsx.clicked.connect(self._download_xlsx)
        self.btn_delete = QPushButton()
        self.btn_delete.setObjectName("secondary")
        self.btn_delete.clicked.connect(self._delete_entry)
        self.btn_close = QPushButton()
        self.btn_close.setObjectName("secondary")
        self.btn_close.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_open)
        btn_row.addWidget(self.btn_csv)
        btn_row.addWidget(self.btn_xlsx)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_delete)
        btn_row.addWidget(self.btn_close)
        lay.addLayout(btn_row)

        self._apply_texts()
        self._reload_list()

    def _tr(self, key: str, **kwargs) -> str:
        return self._main.tr(key, **kwargs)

    def _apply_texts(self) -> None:
        self.setWindowTitle(self._tr("history_title"))
        self.btn_open.setText(self._tr("history_open"))
        self.btn_csv.setText(self._tr("history_download_csv"))
        self.btn_xlsx.setText(self._tr("history_download_xlsx"))
        self.btn_delete.setText(self._tr("history_delete"))
        self.btn_close.setText(self._tr("btn_close"))

    @staticmethod
    def _format_when(iso: str) -> str:
        try:
            return datetime.fromisoformat(iso).strftime("%d.%m.%Y %H:%M")
        except ValueError:
            return iso

    def _reload_list(self) -> None:
        self.list_entries.clear()
        self._records = list_history_entries()
        self._current = None
        self._set_preview_empty()

        for record in self._records:
            meta = record.meta
            when = self._format_when(meta.created_at)
            task_line = meta.task_preview
            if len(task_line) > 72:
                task_line = task_line[:69] + "…"
            text = f"{when}\n{meta.source_name}\n{task_line}"
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, meta.id)
            self.list_entries.addItem(item)

        if self._records:
            self.list_entries.setCurrentRow(0)
        else:
            self.lbl_summary.setText(self._tr("history_empty"))
            self._set_actions_enabled(False)

    def _set_preview_empty(self) -> None:
        self.lbl_task.setText("")
        self.text_preview.clear()
        self._set_actions_enabled(False)

    def _set_actions_enabled(self, enabled: bool) -> None:
        self.btn_open.setEnabled(enabled)
        self.btn_csv.setEnabled(enabled)
        self.btn_xlsx.setEnabled(enabled)
        self.btn_delete.setEnabled(enabled)

    def _on_pick(self, current: QListWidgetItem | None, _previous) -> None:
        if current is None:
            self._current = None
            self._set_preview_empty()
            if not self._records:
                self.lbl_summary.setText(self._tr("history_empty"))
            return

        entry_id = current.data(Qt.ItemDataRole.UserRole)
        record = load_history_entry(entry_id) if entry_id else None
        if record is None:
            self._current = None
            self.lbl_summary.setText(self._tr("history_err_load"))
            self._set_actions_enabled(False)
            return

        self._current = record
        meta = record.meta
        self.lbl_summary.setText(
            meta.report_summary
            or self._tr(
                "changes_summary",
                before=meta.rows_before,
                after=meta.rows_after,
                cells=meta.cells_changed,
                removed=0,
                added=0,
            ),
        )
        self.lbl_task.setText(self._tr("history_task_label", task=meta.task_preview))
        try:
            _original, draft_df = load_history_dataframes(record)
            preview = draft_df.head(80).to_csv(index=False)
            if len(draft_df) > 80:
                preview += f"\n… (+{len(draft_df) - 80} rows)"
            self.text_preview.setPlainText(preview)
            self._set_actions_enabled(True)
        except Exception:
            self.text_preview.setPlainText("")
            self.lbl_summary.setText(self._tr("history_err_load"))
            self._set_actions_enabled(False)

    def _draft_df(self) -> pd.DataFrame | None:
        if self._current is None:
            return None
        try:
            _original, draft_df = load_history_dataframes(self._current)
            return draft_df
        except Exception:
            return None

    def _download_csv(self) -> None:
        df = self._draft_df()
        if df is None or self._current is None:
            self._main._show_warning("history_err_load")
            return
        default = os.path.splitext(self._current.meta.source_name)[0] + "_draft.csv"
        path, _ = QFileDialog.getSaveFileName(
            self, self._main.tr("dlg_save_csv"), default, self._main.tr("filter_csv"),
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        df.to_csv(path, index=False, encoding="utf-8")
        self._main.lbl_progress.setText(
            self._main.tr("progress_saved", name=os.path.basename(path)),
        )

    def _download_xlsx(self) -> None:
        df = self._draft_df()
        if df is None or self._current is None:
            self._main._show_warning("history_err_load")
            return
        default = os.path.splitext(self._current.meta.source_name)[0] + "_draft.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, self._main.tr("dlg_save_xlsx"), default, self._main.tr("filter_xlsx"),
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            save_dataframe_excel(df, path)
            self._main.lbl_progress.setText(
                self._main.tr("progress_saved", name=os.path.basename(path)),
            )
        except Exception as e:
            self._main._show_error("err_excel", err=e)

    def _open_in_work(self) -> None:
        if self._current is None:
            return
        if self._main._apply_history_record(self._current):
            self.accept()

    def _delete_entry(self) -> None:
        if self._current is None:
            return
        answer = QMessageBox.question(
            self,
            self._main.tr("warn_title"),
            self._tr("history_delete_confirm"),
        )
        if answer != QMessageBox.StandardButton.Yes:
            return
        if delete_history_entry(self._current.meta.id):
            self._main.lbl_progress.setText(self._tr("history_deleted"))
            self._reload_list()


def _populate_table_widget(table: QTableWidget, df: pd.DataFrame) -> None:
    table.clear()
    if df.empty:
        table.setRowCount(0)
        table.setColumnCount(0)
        return
    rows, cols = len(df), len(df.columns)
    table.setRowCount(rows)
    table.setColumnCount(cols)
    table.setHorizontalHeaderLabels([str(c) for c in df.columns])
    for row_idx in range(rows):
        for col_idx in range(cols):
            value = df.iat[row_idx, col_idx]
            text = "" if pd.isna(value) else str(value)
            table.setItem(row_idx, col_idx, QTableWidgetItem(text))
    header = table.horizontalHeader()
    if header is not None:
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        header.setStretchLastSection(True)


class DraftTableDialog(QDialog):
    """Просмотр черновика внутри приложения."""

    def __init__(self, parent: "MainWindow") -> None:
        super().__init__(parent)
        self._main = parent
        self.setWindowTitle(parent.tr("dlg_draft_title"))
        self.setMinimumSize(720, 420)
        self.resize(920, 560)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 16)
        lay.setSpacing(10)

        self.lbl_warning = QLabel()
        self.lbl_warning.setObjectName("result_warn")
        self.lbl_warning.setWordWrap(True)
        self.lbl_warning.hide()
        lay.addWidget(self.lbl_warning)

        self.frame_filter = QFrame()
        self.frame_filter.setObjectName("filter_bar")
        filter_lay = QHBoxLayout(self.frame_filter)
        filter_lay.setContentsMargins(12, 8, 12, 8)
        filter_lay.setSpacing(10)
        self.lbl_filter = QLabel(parent.tr("filter_label"))
        self.lbl_filter.setObjectName("filter_bar_label")
        self.combo_filter = QComboBox()
        self.combo_filter.setMinimumWidth(200)
        self.combo_filter.currentIndexChanged.connect(self._on_filter_changed)
        self.lbl_filter_count = QLabel()
        self.lbl_filter_count.setObjectName("filter_count")
        self.btn_filter_reset = QPushButton(parent.tr("filter_reset"))
        self.btn_filter_reset.setObjectName("secondary")
        self.btn_filter_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_filter_reset.clicked.connect(self._reset_filter)
        filter_lay.addWidget(self.lbl_filter)
        filter_lay.addWidget(self.combo_filter)
        filter_lay.addWidget(self.lbl_filter_count)
        filter_lay.addStretch()
        filter_lay.addWidget(self.btn_filter_reset)
        self.frame_filter.hide()
        lay.addWidget(self.frame_filter)

        self.lbl_rows = QLabel()
        self.lbl_rows.setObjectName("changes_summary")
        lay.addWidget(self.lbl_rows)

        self.table = QTableWidget()
        self.table.setObjectName("draft_table")
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        lay.addWidget(self.table, stretch=1)

        btn_row = QHBoxLayout()
        self.btn_open_excel = QPushButton()
        self.btn_open_excel.setObjectName("secondary")
        self.btn_open_excel.clicked.connect(self._open_in_excel)
        self.btn_close = QPushButton()
        self.btn_close.setObjectName("primary")
        self.btn_close.clicked.connect(self.accept)
        btn_row.addWidget(self.btn_open_excel)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_close)
        lay.addLayout(btn_row)

        self._apply_texts()
        self._setup_filter()
        self._reload_table()

    def _tr(self, key: str, **kwargs) -> str:
        return self._main.tr(key, **kwargs)

    def _apply_texts(self) -> None:
        self.setWindowTitle(self._tr("dlg_draft_title"))
        self.lbl_filter.setText(self._tr("filter_label"))
        self.btn_filter_reset.setText(self._tr("filter_reset"))
        self.btn_open_excel.setText(self._tr("btn_open_draft"))
        self.btn_open_excel.setToolTip(self._tr("btn_open_draft_tip"))
        self.btn_close.setText(self._tr("btn_close"))
        warning = self._main._routing_warning_text()
        if warning:
            self.lbl_warning.setText(warning)
            self.lbl_warning.show()

    def _setup_filter(self) -> None:
        show_filter = (
            self._main._routing_workflow_active()
            and self._main._df_draft is not None
            and self._main._draft_has_routing(self._main._df_draft)
        )
        self.frame_filter.setVisible(show_filter)
        if not show_filter:
            return
        self.combo_filter.blockSignals(True)
        current = self._main._filter_mode
        self.combo_filter.clear()
        for label, mode in self._main._filter_combo_items():
            self.combo_filter.addItem(label, mode)
        idx = self.combo_filter.findData(current)
        self.combo_filter.setCurrentIndex(idx if idx >= 0 else 0)
        self.combo_filter.blockSignals(False)
        self.btn_filter_reset.setVisible(self._main._filter_mode != FILTER_ALL)

    def _filtered_dataframe(self) -> pd.DataFrame | None:
        if self._main._df_draft is None:
            return None
        return filter_draft_dataframe(
            self._main._df_draft,
            self._main._filter_mode,
            self._main._allowed_departments,
        )

    def _reload_table(self) -> None:
        full_df = self._main._df_draft
        filtered = self._filtered_dataframe()
        if full_df is None or filtered is None:
            self.table.clear()
            self.lbl_rows.setText("")
            return
        _populate_table_widget(self.table, filtered)
        self.lbl_rows.setText(
            self._tr("dlg_draft_rows", total=len(full_df), shown=len(filtered)),
        )
        self.lbl_filter_count.setText(
            self._tr("filter_count", n=len(filtered), total=len(full_df)),
        )
        self.btn_filter_reset.setVisible(self._main._filter_mode != FILTER_ALL)

    def _on_filter_changed(self, _index: int) -> None:
        mode = self.combo_filter.currentData()
        if mode is None:
            return
        self._main._filter_mode = mode
        self._reload_table()

    def _reset_filter(self) -> None:
        self._main._filter_mode = FILTER_ALL
        idx = self.combo_filter.findData(FILTER_ALL)
        if idx >= 0:
            self.combo_filter.blockSignals(True)
            self.combo_filter.setCurrentIndex(idx)
            self.combo_filter.blockSignals(False)
        self._reload_table()

    def _open_in_excel(self) -> None:
        self._main._open_draft_externally()


class RoutingReviewDialog(QDialog):
    """Проверка раскладки по отделам перед отправкой."""

    def __init__(
        self,
        parent: QWidget,
        title: str,
        hint: str,
        body: str,
        confirm_text: str,
        cancel_text: str,
        configure_edit,
    ):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumSize(480, 400)
        self.resize(540, 440)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 16)
        lay.setSpacing(12)

        lbl_hint = QLabel(hint)
        lbl_hint.setWordWrap(True)
        lbl_hint.setObjectName("changes_summary")
        lay.addWidget(lbl_hint)

        self.text_body = QTextEdit()
        configure_edit(self.text_body, "output", min_height=240, read_only=True)
        self.text_body.setPlainText(body)
        lay.addWidget(self.text_body, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton(cancel_text)
        btn_cancel.setObjectName("secondary")
        btn_cancel.clicked.connect(self.reject)
        btn_ok = QPushButton(confirm_text)
        btn_ok.setObjectName("primary")
        btn_ok.setMinimumWidth(180)
        btn_ok.clicked.connect(self.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)


class DepartmentEmailDialog(QDialog):
    """Выбор отдела и превью текста письма перед копированием."""

    def __init__(
        self,
        parent: QWidget,
        title: str,
        hint: str,
        departments: list[str],
        df: pd.DataFrame,
        labels: dict[str, str],
        copy_text: str,
        cancel_text: str,
        configure_edit,
    ):
        super().__init__(parent)
        self._df = df
        self._labels = labels
        self.setWindowTitle(title)
        self.setMinimumSize(520, 420)
        self.resize(580, 460)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 16)
        lay.setSpacing(12)

        lbl_hint = QLabel(hint)
        lbl_hint.setWordWrap(True)
        lbl_hint.setObjectName("changes_summary")
        lay.addWidget(lbl_hint)

        self.combo_dept = QComboBox()
        self.combo_dept.addItems(departments)
        lay.addWidget(self.combo_dept)

        self.text_preview = QTextEdit()
        configure_edit(self.text_preview, "output", min_height=220, read_only=True)
        lay.addWidget(self.text_preview, stretch=1)
        self.combo_dept.currentTextChanged.connect(self._refresh_preview)
        self._refresh_preview(self.combo_dept.currentText())

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton(cancel_text)
        btn_cancel.setObjectName("secondary")
        btn_cancel.clicked.connect(self.reject)
        btn_copy = QPushButton(copy_text)
        btn_copy.setObjectName("primary")
        btn_copy.setMinimumWidth(130)
        btn_copy.clicked.connect(self.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_copy)
        lay.addLayout(btn_row)

    def _refresh_preview(self, department: str):
        if not department:
            self.text_preview.clear()
            return
        self.text_preview.setPlainText(
            build_department_email(self._df, department, self._labels),
        )

    def selected_department(self) -> str:
        return self.combo_dept.currentText()

    def email_text(self) -> str:
        return self.text_preview.toPlainText()


ERROR_CODE_PREFIX = "code:"


def classify_api_error(raw: str) -> str:
    """Сопоставляет текст исключения API с ключом TEXTS."""
    low = raw.lower()
    if any(
        x in low
        for x in (
            "invalid api key",
            "authentication",
            "401",
            "unauthorized",
            "invalid x-api-key",
            "api key not found",
        )
    ):
        return "err_api_auth"
    if any(x in low for x in ("rate limit", "429", "too many requests")):
        return "err_api_rate"
    if any(x in low for x in ("timeout", "timed out")):
        return "err_api_timeout"
    if any(
        x in low
        for x in (
            "connection",
            "network",
            "resolve",
            "dns",
            "errno",
            "connect error",
        )
    ):
        return "err_api_network"
    if any(
        x in low
        for x in ("credit", "billing", "balance", "payment", "402", "insufficient")
    ):
        return "err_api_billing"
    if any(
        x in low
        for x in ("overloaded", "529", "503", "temporarily unavailable")
    ):
        return "err_api_overload"
    return "err_api_generic"


def format_error_message(tr_fn, key: str, **kwargs) -> str:
    """Сообщение об ошибке: что случилось + что делать."""
    problem = tr_fn(key, **kwargs)
    hint_key = f"{key}_hint"
    hint = tr_fn(hint_key, **kwargs)
    if hint == hint_key:
        return problem
    return f"{problem}\n\n{hint}"


def _prepare_rounded_menu(menu: QMenu) -> None:
    """Убирает системную прямоугольную рамку popup на macOS."""
    menu.setWindowFlags(
        Qt.WindowType.Popup
        | Qt.WindowType.FramelessWindowHint
        | Qt.WindowType.NoDropShadowWindowHint,
    )
    menu.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)


class HoverComboBox(QComboBox):
    """Тема/язык: hover + выпадающее меню (как у шаблонов), без системного списка."""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)
        self._get_palette = None
        self._menu_open = False
        self._hover_active = False

    def set_palette_provider(self, provider) -> None:
        self._get_palette = provider

    def _palette(self) -> dict[str, str]:
        if self._get_palette:
            return self._get_palette()
        return THEME_PALETTES["light"]

    def _interaction_blocked(self) -> bool:
        if self._menu_open:
            return True
        app = QApplication.instance()
        return bool(app and app.activeModalWidget() is not None)

    def _combo_stylesheet(self, hovered: bool) -> str:
        p = self._palette()
        if hovered:
            bg, border = p["BG_ACCENT"], p["BLUE"]
        else:
            bg, border = p["BG_CARD"], p["BORDER"]
        return f"""
            QComboBox {{
                background-color: {bg};
                color: {p["TEXT"]};
                border: 1px solid {border};
                border-radius: 8px;
                padding: 7px 28px 7px 12px;
                min-width: 110px;
            }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: center right;
                width: 24px;
                border: none;
            }}
            QComboBox::down-arrow {{
                width: 0; height: 0;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid {p["TEXT_MUTED"]};
                margin-right: 8px;
            }}
        """

    def _menu_stylesheet(self) -> str:
        p = self._palette()
        return f"""
            QMenu {{
                background-color: {p["BG_CARD"]};
                color: {p["TEXT"]};
                border: 1px solid {p["BORDER"]};
                border-radius: 8px;
                padding: 4px;
                margin: 0px;
            }}
            QMenu::item {{
                padding: 8px 14px;
                border-radius: 6px;
                font-size: 13px;
                color: {p["TEXT"]};
                background: transparent;
            }}
            QMenu::item:selected {{
                background-color: {p["BG_ACCENT"]};
                color: {p["BLUE"]};
            }}
        """

    def refresh_theme(self) -> None:
        self._set_hovered(self._hover_active and not self._interaction_blocked())

    def _set_hovered(self, active: bool) -> None:
        self._hover_active = active
        if self._interaction_blocked():
            active = False
        self.setProperty("hover", "true" if active else "false")
        self.setStyleSheet(self._combo_stylesheet(active))

    def enterEvent(self, event) -> None:
        if not self._interaction_blocked():
            self._set_hovered(True)
        super().enterEvent(event)

    def leaveEvent(self, event) -> None:
        if not self._interaction_blocked():
            self._set_hovered(False)
        super().leaveEvent(event)

    def showPopup(self) -> None:
        if self.count() == 0:
            return
        self._set_hovered(False)
        self._menu_open = True
        menu = QMenu(self)
        _prepare_rounded_menu(menu)
        menu.setStyleSheet(self._menu_stylesheet())
        menu.setMinimumWidth(self.width())
        menu.aboutToHide.connect(self._on_popup_closed)
        for i in range(self.count()):
            action = menu.addAction(self.itemText(i))
            if i == self.currentIndex():
                fnt = action.font()
                fnt.setWeight(QFont.Weight.DemiBold)
                action.setFont(fnt)
            action.triggered.connect(
                lambda _checked=False, idx=i: self.setCurrentIndex(idx),
            )
        menu.exec(self.mapToGlobal(QPoint(0, self.height() + 4)))

    def hidePopup(self) -> None:
        self._menu_open = False

    def _on_popup_closed(self) -> None:
        self._menu_open = False
        QTimer.singleShot(0, self._after_hide_popup)

    def _after_hide_popup(self) -> None:
        if self._interaction_blocked():
            self._set_hovered(False)
            return
        self._set_hovered(self.underMouse())


class DetachButton(QLabel):
    """Компактный крестик, выровненный по центру."""

    _SIZE = 24

    clicked = pyqtSignal()

    def __init__(self, parent: QWidget | None = None):
        super().__init__("", parent)
        self.setObjectName("detach")
        self.setFixedSize(self._SIZE, self._SIZE)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        color = self.palette().color(QPalette.ColorRole.WindowText)
        painter.setPen(
            QPen(color, 1.5, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap),
        )
        pad = 7
        w, h = self.width(), self.height()
        painter.drawLine(pad, pad, w - pad, h - pad)
        painter.drawLine(pad, h - pad, w - pad, pad)
        painter.end()

    def mouseReleaseEvent(self, event: QMouseEvent | None) -> None:
        if event and event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mouseReleaseEvent(event)


class TaskEnterFilter(QObject):
    """Enter — отправить задачу; Shift+Enter — новая строка."""

    def __init__(self, run_cb, parent: QObject | None = None):
        super().__init__(parent)
        self._run_cb = run_cb

    def eventFilter(self, watched: QObject, event: QEvent | None) -> bool:
        if (
            event
            and event.type() == QEvent.Type.KeyPress
            and isinstance(event, QKeyEvent)
            and event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter)
        ):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                return False
            QTimer.singleShot(0, self._run_cb)
            return True
        return False


class TaskComposerFocusFilter(QObject):
    """Подсветка рамки композера при фокусе в поле ввода."""

    def __init__(self, frame: QFrame, parent: QObject | None = None):
        super().__init__(parent)
        self._frame = frame

    def eventFilter(self, watched: QObject, event: QEvent | None) -> bool:
        if event and event.type() == QEvent.Type.FocusIn:
            self._frame.setProperty("focused", "true")
            self._frame.update()
        elif event and event.type() == QEvent.Type.FocusOut:
            self._frame.setProperty("focused", "false")
            self._frame.update()
        return False


class Worker(QObject):
    """Фоновый запрос к API — через threading, не QThread (стабильнее SSL на macOS)."""

    finished_ok = pyqtSignal(str)
    finished_err = pyqtSignal(str)

    def __init__(
        self,
        api_key: str,
        file_path: str,
        task: str,
        company_rules: str = "",
        parent: QObject | None = None,
    ):
        super().__init__(parent)
        self.api_key = api_key
        self.file_path = file_path
        self.task = task
        self.company_rules = company_rules
        self._thread: threading.Thread | None = None

    def isRunning(self) -> bool:
        return self._thread is not None and self._thread.is_alive()

    def start(self) -> None:
        self._thread = threading.Thread(
            target=self._run,
            name="smartsorter-api",
            daemon=True,
        )
        self._thread.start()

    def _run(self) -> None:
        try:
            df = read_table(self.file_path)
            if df.empty:
                self.finished_err.emit(f"{ERROR_CODE_PREFIX}err_file_empty")
                return
            csv_data = df.to_csv(index=False)
            user_message = build_api_user_message(
                csv_data,
                self.task,
                file_path=self.file_path,
                df=df,
                company_rules=self.company_rules,
            )
            client = Anthropic(api_key=self.api_key)
            response = client.messages.create(
                model="claude-opus-4-5-20251101",
                max_tokens=16384,
                temperature=0,
                system=build_system_prompt(self.company_rules),
                messages=[{"role": "user", "content": user_message}],
            )
            result = "".join(
                b.text for b in response.content if hasattr(b, "text")
            ).strip()
            self.finished_ok.emit(result)
        except ValueError as e:
            if str(e) == "bad_format":
                self.finished_err.emit(f"{ERROR_CODE_PREFIX}err_type")
            else:
                self.finished_err.emit(f"{ERROR_CODE_PREFIX}err_read")
        except Exception as e:
            self.finished_err.emit(str(e))


def _app_icon_pixmap(size: int = 44) -> QPixmap:
    pixmap = QPixmap(APP_ICON_PATH)
    if pixmap.isNull():
        return QPixmap()
    return pixmap.scaled(
        size,
        size,
        Qt.AspectRatioMode.KeepAspectRatio,
        Qt.TransformationMode.SmoothTransformation,
    )


def _apply_app_window_icon(window: QMainWindow) -> None:
    icon = QIcon(APP_ICON_PATH)
    if icon.isNull():
        return
    window.setWindowIcon(icon)
    app = QApplication.instance()
    if app:
        app.setWindowIcon(icon)


def _paint_info_glyph(painter: QPainter, glyph: str, rect: QRectF, color: QColor) -> None:
    pen = QPen(color, 1.75, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap)
    painter.setPen(pen)
    painter.setBrush(Qt.BrushStyle.NoBrush)
    cx = rect.center().x()
    cy = rect.center().y()
    w = rect.width()
    h = rect.height()
    if glyph == "about":
        painter.drawEllipse(rect)
        painter.drawLine(QPointF(cx, cy + h * 0.05), QPointF(cx, cy + h * 0.28))
        painter.setBrush(QBrush(color))
        painter.drawEllipse(QPointF(cx, cy - h * 0.18), 1.4, 1.4)
        painter.setBrush(Qt.BrushStyle.NoBrush)
    elif glyph == "steps":
        for y_frac in (-0.26, 0.0, 0.26):
            y = cy + h * y_frac
            painter.drawEllipse(QRectF(cx - w * 0.34, y - 2.0, 4.0, 4.0))
            painter.drawLine(QPointF(cx - w * 0.2, y), QPointF(cx + w * 0.34, y))
    elif glyph == "templates":
        painter.drawRoundedRect(
            QRectF(cx - w * 0.3, cy - h * 0.24, w * 0.5, h * 0.36), 2.0, 2.0,
        )
        painter.drawRoundedRect(
            QRectF(cx - w * 0.18, cy - h * 0.1, w * 0.5, h * 0.36), 2.0, 2.0,
        )
        for y_frac in (-0.02, 0.1, 0.22):
            y = cy + h * y_frac
            painter.drawLine(QPointF(cx - w * 0.1, y), QPointF(cx + w * 0.26, y))
    elif glyph == "data":
        painter.drawEllipse(QRectF(cx - w * 0.1, cy - h * 0.34, w * 0.34, h * 0.34))
        painter.drawLine(
            QPointF(cx + w * 0.08, cy - h * 0.08),
            QPointF(cx + w * 0.3, cy + h * 0.3),
        )
        painter.drawLine(
            QPointF(cx + w * 0.18, cy + h * 0.18),
            QPointF(cx + w * 0.32, cy + h * 0.18),
        )
        painter.drawLine(
            QPointF(cx + w * 0.22, cy + h * 0.26),
            QPointF(cx + w * 0.36, cy + h * 0.26),
        )


class WorkflowNavItem(QFrame):
    """Кликабельный шаг мастера в боковой панели."""

    clicked = pyqtSignal(str)

    def __init__(self, key: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._key = key
        self.setObjectName("wf_item")
        self.setProperty("state", "idle")
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit(self._key)
        super().mouseReleaseEvent(event)


class InfoGlyphButton(QPushButton):
    """Иконка справки — тонкие линии в стиле SF Symbols."""

    hovered = pyqtSignal()
    unhovered = pyqtSignal()

    def __init__(self, glyph: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._glyph = glyph
        self.setObjectName("info_icon")
        self.setFixedSize(44, 44)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setText("")
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)

    def enterEvent(self, event) -> None:
        self.hovered.emit()
        super().enterEvent(event)

    def leaveEvent(self, event) -> None:
        self.unhovered.emit()
        super().leaveEvent(event)

    def set_active(self, active: bool) -> None:
        self.setProperty("active", "true" if active else "false")
        self.update()

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        if self.width() < 22 or self.height() < 22:
            return
        painter = QPainter(self)
        if not painter.isActive():
            return
        try:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            if self.property("active") == "true":
                color = QColor(self.palette().color(QPalette.ColorRole.Highlight))
            elif self.underMouse():
                color = QColor(self.palette().color(QPalette.ColorRole.WindowText))
            else:
                color = QColor(self.palette().color(QPalette.ColorRole.PlaceholderText))
                color.setAlpha(210)
            inner = QRectF(self.rect()).adjusted(11, 11, -11, -11)
            if inner.width() > 4 and inner.height() > 4:
                _paint_info_glyph(painter, self._glyph, inner, color)
        finally:
            painter.end()


class InfoHoverCard(QFrame):
    """Закруглённая подсказка при наведении — поверх контента, справа от иконки."""

    def __init__(self, host: QWidget) -> None:
        super().__init__(host)
        self.setObjectName("info_hover_card")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(12, 8, 12, 8)
        self._label = QLabel()
        self._label.setObjectName("info_hover_text")
        lay.addWidget(self._label)
        self.hide()

    def apply_palette(self, palette: dict[str, str]) -> None:
        p = palette
        self.setStyleSheet(f"""
            QFrame#info_hover_card {{
                background: {p["BG_CARD"]};
                border: 1px solid {p["BORDER"]};
                border-radius: 10px;
            }}
            QLabel#info_hover_text {{
                color: {p["TEXT"]};
                font-size: 12px;
                font-weight: 600;
                background: transparent;
            }}
        """)

    def show_for(self, anchor: QWidget, text: str) -> None:
        host = self.parentWidget()
        if host is None:
            return
        self._label.setText(text)
        self.adjustSize()
        pos = anchor.mapTo(host, QPoint(anchor.width() + 10, 0))
        pos.setY(pos.y() + max(0, (anchor.height() - self.height()) // 2))
        self.move(pos)
        self.show()
        self.raise_()


class InfoSideDrawer(QFrame):
    """Выдвижная панель справки справа от иконок."""

    DRAWER_WIDTH = 292

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("info_drawer")
        self._open_key = ""
        self.setMinimumWidth(0)
        self.setMaximumWidth(0)
        self.setSizePolicy(
            QSizePolicy.Policy.Fixed,
            QSizePolicy.Policy.Expanding,
        )

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 18, 16, 16)
        lay.setSpacing(12)

        hdr = QHBoxLayout()
        hdr.setSpacing(8)
        self.lbl_title = QLabel()
        self.lbl_title.setObjectName("info_drawer_title")
        self.lbl_title.setWordWrap(True)
        hdr.addWidget(self.lbl_title, stretch=1)
        self.btn_close = QPushButton("\u2039")
        self.btn_close.setObjectName("info_drawer_close")
        self.btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_close.setToolTip("")
        hdr.addWidget(self.btn_close)
        lay.addLayout(hdr)

        self.text_body = QTextBrowser()
        self.text_body.setObjectName("info_drawer_body")
        self.text_body.setOpenExternalLinks(True)
        self.text_body.setFrameShape(QFrame.Shape.NoFrame)
        lay.addWidget(self.text_body, stretch=1)

        self._anim = QPropertyAnimation(self, b"maximumWidth", self)
        self._anim.setDuration(240)
        self._anim.setEasingCurve(QEasingCurve.Type.OutCubic)

    def is_open_for(self, key: str) -> bool:
        return self._open_key == key and self.maximumWidth() > 0

    def is_visible_drawer(self) -> bool:
        return self.maximumWidth() > 0

    def open_with(self, key: str, title: str, body: str) -> None:
        self._open_key = key
        self.lbl_title.setText(title)
        self._set_body(key, body)
        self._animate_to(self.DRAWER_WIDTH)

    def _set_body(self, key: str, body: str) -> None:
        if key == "data":
            self.text_body.setHtml(body)
        else:
            self.text_body.setPlainText(body)

    def close_drawer(self) -> None:
        self._open_key = ""
        self._animate_to(0)

    def _animate_to(self, width: int) -> None:
        self._anim.stop()
        self._anim.setStartValue(self.maximumWidth())
        self._anim.setEndValue(width)
        self._anim.start()


class MainWindow(QMainWindow):
    LAYOUT_BUILD = "2.2.0"
    TASK_INPUT_MIN_H = 88
    TASK_INPUT_MAX_H = 200
    TASK_SEND_ROW_H = 44
    LEFT_PANEL_MIN = 260
    LEFT_PANEL_MAX = 340

    def __init__(self):
        super().__init__()
        self.lang = load_language()
        self.theme = load_theme()
        self.input_file_path = ""
        self.result_csv = ""
        self._df_original: pd.DataFrame | None = None
        self._df_draft: pd.DataFrame | None = None
        self._worker: Worker | None = None
        self._worker_busy = False
        self._pending_worker_result = ""
        self._pending_worker_error = ""
        self._worker_busy = False
        self._key_hidden = True
        self._last_changes: TableChanges | None = None
        self._report_summary = ""
        self._report_details = ""
        self._allowed_departments: list[str] = []
        self._filter_mode = FILTER_ALL
        self._last_operation_path = ""
        self._info_section_key = ""
        self._info_nav_buttons: dict[str, QPushButton] = {}
        self._body_host: QWidget | None = None
        self._info_shell: QWidget | None = None
        self._info_hover_card: InfoHoverCard | None = None
        self._info_drawer: InfoSideDrawer | None = None
        self._info_hover_btn: InfoGlyphButton | None = None
        self._selected_tpl_index: int | None = None
        self._wf_nav: dict[str, tuple[WorkflowNavItem, QLabel, QLabel]] = {}
        self._file_rows = 0
        self._file_cols = 0
        self.setMinimumSize(1120, 760)
        self.resize(1280, 880)
        self._build_ui()
        _apply_app_window_icon(self)
        self._setup_shortcuts()
        self._apply_theme()
        self._apply_language()

    def tr(self, key: str, **kwargs) -> str:
        text = TEXTS.get(self.lang, TEXTS["en"]).get(key, key)
        return text.format(**kwargs) if kwargs else text

    def _error_text(self, key: str, **kwargs) -> str:
        return format_error_message(self.tr, key, **kwargs)

    def _show_warning(self, key: str, **kwargs) -> None:
        QMessageBox.warning(self, self.tr("warn_title"), self._error_text(key, **kwargs))

    def _show_error(self, key: str, **kwargs) -> None:
        QMessageBox.critical(self, self.tr("err_title"), self._error_text(key, **kwargs))

    def _show_api_error(self, raw: str) -> None:
        key = classify_api_error(raw)
        if key == "err_api_generic":
            body = self._error_text(key, detail=raw)
        else:
            body = self._error_text(key)
        QMessageBox.critical(self, self.tr("err_api_title"), body)

    TOOLBAR_BTN_WIDTH = 100

    @staticmethod
    def _configure_toolbar_button(btn: QPushButton, *, width: int | None = None) -> None:
        btn.setMinimumWidth(width or MainWindow.TOOLBAR_BTN_WIDTH)
        height = 30 if btn.objectName() == "toolbar_btn" else 34
        btn.setFixedHeight(height)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)

    @staticmethod
    def _configure_text_edit(
        editor: QTextEdit,
        object_name: str,
        *,
        min_height: int = 120,
        read_only: bool = False,
        max_height: int | None = None,
    ) -> None:
        editor.setObjectName(object_name)
        editor.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        editor.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        editor.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        editor.setFrameShape(QFrame.Shape.StyledPanel)
        editor.document().setDocumentMargin(8)
        editor.setMinimumHeight(min_height)
        if max_height is not None:
            editor.setMaximumHeight(max_height)
        editor.setReadOnly(read_only)

    def _elide_label_text(
        self,
        label: QLabel,
        text: str,
        *,
        reserve: int = 36,
    ) -> None:
        fm = QFontMetrics(label.font())
        width = label.width() if label.width() > reserve else 220
        label.setText(
            fm.elidedText(text, Qt.TextElideMode.ElideMiddle, width - reserve),
        )

    def _show_file_empty_state(self, hint: str | None = None) -> None:
        self._file_rows = 0
        self._file_cols = 0
        self.btn_open.show()
        self.btn_change_file.hide()
        self.frame_file_chip.hide()
        self.lbl_file.setText(hint or self.tr("no_file"))
        self.lbl_file.setObjectName("file_empty")
        self.lbl_file.show()
        self._update_result_file_display()
        self._refresh_workflow_ui()

    def _show_file_selected(self, path: str) -> None:
        self.btn_open.hide()
        self.lbl_file.hide()
        self.frame_file_chip.show()
        self.btn_change_file.show()
        self._elide_label_text(
            self.lbl_file_name,
            os.path.basename(path),
            reserve=40,
        )
        self._update_result_file_display()
        self._refresh_workflow_ui()

    def _refresh_elided_labels(self) -> None:
        if self.input_file_path:
            self._elide_label_text(
                self.lbl_file_name,
                os.path.basename(self.input_file_path),
                reserve=40,
            )
            self._update_result_file_display()
        rules_path = load_rules_path()
        if rules_path and os.path.exists(rules_path):
            self._elide_label_text(
                self.lbl_rules_file,
                os.path.basename(rules_path),
                reserve=40,
            )

    def resizeEvent(self, event: QResizeEvent | None) -> None:
        super().resizeEvent(event)
        self._refresh_elided_labels()

    def _set_wf_state(self, key: str, state: str) -> None:
        item = self._wf_nav.get(key)
        if item is None:
            return
        frame, dot, _label = item
        if frame.property("state") != state:
            frame.setProperty("state", state)
            frame.style().unpolish(frame)
            frame.style().polish(frame)
        if state == "done":
            dot.setText("\u2713")
        else:
            dot.setText({"setup": "1", "work": "2", "result": "3"}[key])

    def _update_workflow_nav(self) -> None:
        if not self._wf_nav:
            return
        has_key = bool(
            (hasattr(self, "entry_api") and self.entry_api.text().strip())
            or load_api_key(),
        )
        has_file = bool(self.input_file_path)
        has_draft = self._draft_dataframe() is not None

        if has_key and has_file:
            setup_state = "done"
        elif has_key or has_file:
            setup_state = "active"
        else:
            setup_state = "idle"

        if has_draft:
            work_state = "done"
            result_state = "done"
        elif has_file:
            work_state = "active"
            result_state = "idle"
        else:
            work_state = "idle"
            result_state = "idle"

        self._set_wf_state("setup", setup_state)
        self._set_wf_state("work", work_state)
        self._set_wf_state("result", result_state)

    def _on_wf_nav_clicked(self, key: str) -> None:
        if key == "setup":
            self.entry_api.setFocus()
        elif key == "work":
            self.entry_task.setFocus()
        elif key == "result":
            if self._draft_dataframe() is not None:
                self.btn_view_draft.setFocus()
            else:
                self.lbl_result_card_title.setFocus()

    def _build_nav_sidebar(self) -> QFrame:
        sidebar = QFrame()
        sidebar.setObjectName("nav_sidebar")
        sidebar.setFixedWidth(248)
        lay = QVBoxLayout(sidebar)
        lay.setContentsMargins(16, 22, 16, 18)
        lay.setSpacing(0)

        brand = QHBoxLayout()
        brand.setSpacing(10)
        self.lbl_app_icon = QLabel()
        self.lbl_app_icon.setObjectName("app_icon")
        self.lbl_app_icon.setFixedSize(40, 40)
        self.lbl_app_icon.setPixmap(_app_icon_pixmap(40))
        self.lbl_app_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        brand.addWidget(self.lbl_app_icon, alignment=Qt.AlignmentFlag.AlignTop)
        title_box = QVBoxLayout()
        title_box.setSpacing(3)
        self.lbl_app_title = QLabel()
        self.lbl_app_title.setObjectName("sidebar_title")
        self.lbl_app_title.setWordWrap(True)
        self.lbl_app_sub = QLabel()
        self.lbl_app_sub.setObjectName("sidebar_sub")
        self.lbl_app_sub.setWordWrap(True)
        title_box.addWidget(self.lbl_app_title)
        title_box.addWidget(self.lbl_app_sub)
        brand.addLayout(title_box, stretch=1)
        lay.addLayout(brand)
        lay.addSpacing(22)

        wf_title = QLabel()
        wf_title.setObjectName("sidebar_section")
        self.lbl_wf_section = wf_title
        lay.addWidget(wf_title)
        lay.addSpacing(10)

        self._wf_nav.clear()
        for idx, key in enumerate(("setup", "work", "result"), start=1):
            item = WorkflowNavItem(key, sidebar)
            item.clicked.connect(self._on_wf_nav_clicked)
            row = QHBoxLayout(item)
            row.setContentsMargins(10, 8, 10, 8)
            row.setSpacing(10)
            dot = QLabel(str(idx))
            dot.setObjectName("wf_dot")
            dot.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl = QLabel()
            lbl.setObjectName("wf_label")
            row.addWidget(dot, alignment=Qt.AlignmentFlag.AlignVCenter)
            row.addWidget(lbl, stretch=1, alignment=Qt.AlignmentFlag.AlignVCenter)
            self._wf_nav[key] = (item, dot, lbl)
            lay.addWidget(item)
        lay.addStretch(1)

        help_title = QLabel()
        help_title.setObjectName("sidebar_section")
        self.lbl_help_section = help_title
        lay.addWidget(help_title)
        lay.addSpacing(8)

        self._info_nav_buttons.clear()
        for key, title_key, _body_key in self.INFO_SECTIONS:
            btn = QPushButton()
            btn.setObjectName("info_nav_btn")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.clicked.connect(
                lambda _checked=False, section=key: self._open_info_section(section),
            )
            self._info_nav_buttons[key] = btn
            lay.addWidget(btn)
        lay.addSpacing(16)

        self.lbl_theme = QLabel()
        self.lbl_theme.setObjectName("lang_label")
        self.combo_theme = HoverComboBox()
        self.combo_theme.setObjectName("theme_combo")
        self.combo_theme.set_palette_provider(self._palette)
        self.combo_theme.setCursor(Qt.CursorShape.PointingHandCursor)
        self.combo_theme.currentIndexChanged.connect(self._on_theme_changed)
        theme_row = QHBoxLayout()
        theme_row.setSpacing(8)
        theme_row.addWidget(self.lbl_theme)
        theme_row.addWidget(self.combo_theme, stretch=1)
        lay.addLayout(theme_row)

        self.lbl_lang = QLabel()
        self.lbl_lang.setObjectName("lang_label")
        self.combo_lang = HoverComboBox()
        self.combo_lang.setObjectName("lang_combo")
        self.combo_lang.set_palette_provider(self._palette)
        self.combo_lang.setCursor(Qt.CursorShape.PointingHandCursor)
        for code, name in LANG_OPTIONS:
            self.combo_lang.addItem(name, code)
        self.combo_lang.currentIndexChanged.connect(self._on_language_changed)
        lang_row = QHBoxLayout()
        lang_row.setSpacing(8)
        lang_row.addWidget(self.lbl_lang)
        lang_row.addWidget(self.combo_lang, stretch=1)
        lay.addLayout(lang_row)

        return sidebar

    def _build_input_column(self) -> QWidget:
        col = QWidget()
        col.setObjectName("input_column")
        lay = QVBoxLayout(col)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(14)
        lay.addWidget(self._build_setup_panel(), stretch=0)
        task_dock = self._build_task_dock()
        task_dock.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding,
        )
        lay.addWidget(task_dock, stretch=1)
        return col

    def _build_ui(self):
        # layout 2.2 — единая шапка результата, текстовая справка, без дублей подсказок
        root = QWidget()
        self.setCentralWidget(root)
        outer = QVBoxLayout(root)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        body = QWidget()
        self._body_host = body
        body_lay = QHBoxLayout(body)
        body_lay.setContentsMargins(0, 0, 0, 0)
        body_lay.setSpacing(0)
        body_lay.addWidget(self._build_nav_sidebar())

        content = QWidget()
        content_lay = QHBoxLayout(content)
        content_lay.setContentsMargins(0, 0, 0, 0)
        content_lay.setSpacing(0)

        self._info_drawer = InfoSideDrawer(content)
        self._info_drawer.btn_close.clicked.connect(self._close_info_drawer)
        content_lay.addWidget(self._info_drawer)

        main_stage = QWidget()
        main_stage.setObjectName("main_stage")
        stage_lay = QVBoxLayout(main_stage)
        stage_lay.setContentsMargins(20, 16, 20, 12)
        stage_lay.setSpacing(10)

        cmd_bar = QFrame()
        cmd_bar.setObjectName("command_bar")
        cmd_lay = QHBoxLayout(cmd_bar)
        cmd_lay.setContentsMargins(14, 10, 14, 10)
        cmd_lay.setSpacing(12)
        cmd_title_box = QVBoxLayout()
        cmd_title_box.setSpacing(2)
        self.lbl_workspace_title = QLabel()
        self.lbl_workspace_title.setObjectName("command_title")
        self.lbl_workspace_sub = QLabel()
        self.lbl_workspace_sub.setObjectName("command_sub")
        cmd_title_box.addWidget(self.lbl_workspace_title)
        cmd_title_box.addWidget(self.lbl_workspace_sub)
        cmd_lay.addLayout(cmd_title_box, stretch=1)
        self.lbl_version_badge = QLabel()
        self.lbl_version_badge.setObjectName("version_badge")
        self.lbl_version_badge.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
        )
        cmd_lay.addWidget(
            self.lbl_version_badge,
            alignment=Qt.AlignmentFlag.AlignVCenter,
        )
        stage_lay.addWidget(cmd_bar, stretch=0)

        split_host = QWidget()
        split = QHBoxLayout(split_host)
        split.setContentsMargins(0, 0, 0, 0)
        split.setSpacing(12)
        split.addWidget(self._build_input_column(), stretch=11)

        self.frame_result_card = self._result_section()
        self.frame_result_card.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding,
        )
        self.frame_result_card.setMinimumWidth(340)
        split.addWidget(self.frame_result_card, stretch=13)
        stage_lay.addWidget(split_host, stretch=1)

        content_lay.addWidget(main_stage, stretch=1)
        body_lay.addWidget(content, stretch=1)

        self._info_hover_card = InfoHoverCard(body)
        self._info_hover_card.hide()
        self._info_shell = None
        outer.addWidget(body, stretch=1)
        self._update_info_panel()

    def _card(self) -> tuple[QFrame, QVBoxLayout, QLabel]:
        frame = QFrame()
        frame.setObjectName("card")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(20, 18, 20, 18)
        lay.setSpacing(12)
        title = QLabel()
        title.setObjectName("card_title")
        lay.addWidget(title)
        return frame, lay, title

    INFO_SECTIONS: tuple[tuple[str, str, str], ...] = (
        ("about", "info_about_title", "info_about"),
        ("steps", "info_steps_title", "info_steps"),
        ("templates", "info_templates_title", "info_templates"),
        ("data", "info_data_title", "info_data"),
    )
    def _show_info_hover(self, btn: InfoGlyphButton, key: str) -> None:
        if QApplication.activeModalWidget() is not None:
            return
        if self._info_drawer and self._info_drawer.is_visible_drawer():
            return
        if self._info_hover_card is None:
            return
        title_key = ""
        for section, t_key, _ in self.INFO_SECTIONS:
            if section == key:
                title_key = t_key
                break
        if not title_key:
            return
        self._info_hover_btn = btn
        self._info_hover_card.apply_palette(self._palette())
        self._info_hover_card.show_for(btn, self.tr(title_key))

    def _hide_info_hover(self) -> None:
        if self._info_hover_card is not None:
            self._info_hover_card.hide()
        self._info_hover_btn = None

    def _set_info_icon_active(self, key: str | None) -> None:
        for section, btn in self._info_nav_buttons.items():
            active = section == key
            btn.setProperty("active", "true" if active else "false")
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def _close_info_drawer(self) -> None:
        if self._info_drawer is None:
            return
        self._info_drawer.close_drawer()
        self._info_section_key = ""
        self._set_info_icon_active(None)

    def _open_info_section(self, key: str) -> None:
        self._hide_info_hover()
        if self._info_drawer and self._info_drawer.is_open_for(key):
            self._close_info_drawer()
            return
        title_key = ""
        body_key = ""
        for section, t_key, b_key in self.INFO_SECTIONS:
            if section == key:
                title_key = t_key
                body_key = b_key
                break
        if not title_key or self._info_drawer is None:
            return
        self._info_section_key = key
        self._info_drawer.open_with(
            key,
            self.tr(title_key),
            self.tr(body_key),
        )
        self._set_info_icon_active(key)

    def _update_info_panel(self) -> None:
        if self._info_drawer is not None:
            self._info_drawer.btn_close.setToolTip(self.tr("btn_close"))
        if self._info_section_key and self._info_drawer is not None:
            for section, t_key, b_key in self.INFO_SECTIONS:
                if section == self._info_section_key:
                    self._info_drawer.lbl_title.setText(self.tr(t_key))
                    self._info_drawer._set_body(section, self.tr(b_key))
                    break

    def _build_setup_panel(self) -> QFrame:
        panel = QFrame()
        panel.setObjectName("stage_panel")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(12)

        self.lbl_card_api = QLabel()
        self.lbl_card_api.setObjectName("panel_label")
        lay.addWidget(self.lbl_card_api)
        api_box = QHBoxLayout()
        api_box.setSpacing(8)
        self.entry_api = QLineEdit()
        self.entry_api.setEchoMode(QLineEdit.EchoMode.Password)
        self.entry_api.setPlaceholderText("sk-ant-api03-...")
        self.entry_api.setFont(QFont("Menlo", 10))
        self.entry_api.editingFinished.connect(self._save_key)
        self.entry_api.textChanged.connect(self._on_api_key_changed)
        self.btn_eye = QPushButton()
        self.btn_eye.setObjectName("secondary")
        self.btn_eye.setFixedWidth(92)
        self.btn_eye.clicked.connect(self._toggle_key)
        api_box.addWidget(self.entry_api, stretch=1)
        api_box.addWidget(self.btn_eye)
        lay.addLayout(api_box)
        self.lbl_api_status = QLabel()
        self.lbl_api_status.setObjectName("status_pending")
        lay.addWidget(self.lbl_api_status)

        sep = QFrame()
        sep.setObjectName("panel_sep")
        sep.setFixedHeight(1)
        lay.addWidget(sep)

        self.lbl_card_file = QLabel()
        self.lbl_card_file.setObjectName("panel_label")
        lay.addWidget(self.lbl_card_file)
        file_box = QHBoxLayout()
        file_box.setSpacing(8)
        self.btn_open = QPushButton()
        self.btn_open.setObjectName("secondary")
        self.btn_open.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_open.setMinimumWidth(120)
        self.btn_open.clicked.connect(self._pick_file)
        self.lbl_file = QLabel()
        self.lbl_file.setObjectName("file_empty")
        self.lbl_file.setWordWrap(False)
        self.frame_file_chip = QFrame()
        self.frame_file_chip.setObjectName("file_chip")
        self.frame_file_chip.hide()
        chip_lay = QHBoxLayout(self.frame_file_chip)
        chip_lay.setContentsMargins(10, 4, 8, 4)
        chip_lay.setSpacing(6)
        self.lbl_file_name = QLabel()
        self.lbl_file_name.setObjectName("file_chip_name")
        self.lbl_file_name.setWordWrap(False)
        self.lbl_file_name.setMinimumHeight(24)
        self.lbl_file_name.setAlignment(
            Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
        )
        self.btn_clear_file = DetachButton()
        self.btn_clear_file.clicked.connect(self._clear_file)
        chip_lay.addWidget(
            self.lbl_file_name, stretch=1,
            alignment=Qt.AlignmentFlag.AlignVCenter,
        )
        chip_lay.addWidget(
            self.btn_clear_file,
            alignment=Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight,
        )
        self.btn_change_file = QPushButton()
        self.btn_change_file.setObjectName("secondary")
        self.btn_change_file.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_change_file.clicked.connect(self._pick_file)
        self.btn_change_file.hide()
        file_box.addWidget(self.btn_open)
        file_box.addWidget(self.frame_file_chip, stretch=1)
        file_box.addWidget(self.lbl_file, stretch=1)
        file_box.addWidget(self.btn_change_file)
        lay.addLayout(file_box)
        self.lbl_file_meta = QLabel()
        self.lbl_file_meta.setObjectName("file_meta")
        lay.addWidget(self.lbl_file_meta)
        return panel

    def _build_task_dock(self) -> QFrame:
        dock = QFrame()
        dock.setObjectName("stage_panel")
        dock_lay = QVBoxLayout(dock)
        dock_lay.setContentsMargins(16, 14, 16, 14)
        dock_lay.setSpacing(10)

        self.lbl_card_task = QLabel()
        self.lbl_card_task.setObjectName("panel_label")
        dock_lay.addWidget(self.lbl_card_task)

        tpl_row = QHBoxLayout()
        tpl_row.setSpacing(8)
        tpl_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.lbl_tpl = QLabel()
        self.lbl_tpl.setObjectName("inline_label")
        self.btn_tpl = QPushButton()
        self.btn_tpl.setObjectName("tpl_pick")
        self.btn_tpl.setMinimumHeight(36)
        self.btn_tpl.setMaximumHeight(36)
        self.btn_tpl.setSizePolicy(
            QSizePolicy.Policy.Fixed,
            QSizePolicy.Policy.Fixed,
        )
        self.btn_tpl.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_tpl.clicked.connect(self._show_tpl_menu)
        tpl_row.addWidget(self.lbl_tpl)
        tpl_row.addWidget(self.btn_tpl)
        tpl_row.addStretch(1)
        dock_lay.addLayout(tpl_row)

        self.frame_rules_block = QFrame()
        rules_block_lay = QVBoxLayout(self.frame_rules_block)
        rules_block_lay.setContentsMargins(0, 0, 0, 0)
        rules_block_lay.setSpacing(6)
        rules_row = QHBoxLayout()
        rules_row.setSpacing(8)
        self.lbl_rules = QLabel()
        self.lbl_rules.setObjectName("inline_label")
        self.btn_attach_rules = QPushButton()
        self.btn_attach_rules.setObjectName("attach_rules")
        self.btn_attach_rules.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_attach_rules.clicked.connect(self._pick_rules_file)
        rules_row.addWidget(self.lbl_rules)
        rules_row.addWidget(self.btn_attach_rules)
        rules_row.addStretch(1)
        rules_block_lay.addLayout(rules_row)

        rules_status = QHBoxLayout()
        rules_status.setSpacing(8)
        self.lbl_rules_file = QLabel()
        self.lbl_rules_file.setObjectName("rules_hint")
        self.lbl_rules_file.setWordWrap(True)
        self.lbl_rules_file.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
        )
        self.btn_clear_rules = DetachButton()
        self.btn_clear_rules.hide()
        self.btn_clear_rules.clicked.connect(self._clear_rules_file)
        rules_status.addStretch(1)
        rules_status.addWidget(self.lbl_rules_file)
        rules_status.addWidget(self.btn_clear_rules)
        rules_block_lay.addLayout(rules_status)
        self.frame_rules_block.hide()
        dock_lay.addWidget(self.frame_rules_block)

        self.frame_task_composer = QFrame()
        self.frame_task_composer.setObjectName("task_composer")
        composer_lay = QVBoxLayout(self.frame_task_composer)
        composer_lay.setContentsMargins(0, 0, 0, 0)
        composer_lay.setSpacing(0)

        self.entry_task = QTextEdit()
        self._configure_text_edit(
            self.entry_task,
            "task_input",
            min_height=self.TASK_INPUT_MIN_H,
        )
        self.entry_task.document().setDocumentMargin(4)
        self._task_enter_filter = TaskEnterFilter(self._run, self)
        self.entry_task.installEventFilter(self._task_enter_filter)
        self._task_focus_filter = TaskComposerFocusFilter(
            self.frame_task_composer, self,
        )
        self.entry_task.installEventFilter(self._task_focus_filter)
        self.entry_task.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding,
        )
        self.entry_task.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded,
        )
        self.entry_task.textChanged.connect(self._on_task_changed)
        composer_lay.addWidget(self.entry_task)

        send_row = QWidget()
        send_row.setObjectName("task_send_row")
        send_row.setFixedHeight(self.TASK_SEND_ROW_H)
        send_outer = QVBoxLayout(send_row)
        send_outer.setContentsMargins(10, 0, 8, 8)
        send_outer.setSpacing(6)
        self.progress_run = QProgressBar()
        self.progress_run.setObjectName("run_progress")
        self.progress_run.setTextVisible(False)
        self.progress_run.setRange(0, 100)
        self.progress_run.setValue(0)
        self.progress_run.hide()
        send_outer.addWidget(self.progress_run)
        send_lay = QHBoxLayout()
        send_lay.setContentsMargins(0, 0, 0, 0)
        send_lay.setSpacing(10)
        self.lbl_progress = QLabel("")
        self.lbl_progress.setObjectName("task_progress")
        self.lbl_progress.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
        )
        self.lbl_progress.setWordWrap(True)
        self.btn_run = QPushButton()
        self.btn_run.setObjectName("task_send")
        self.btn_run.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_run.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.btn_run.clicked.connect(self._run)
        send_lay.addWidget(self.lbl_progress, stretch=1)
        send_lay.addWidget(self.btn_run)
        send_outer.addLayout(send_lay)
        composer_lay.addWidget(send_row)

        dock_lay.addWidget(self.frame_task_composer, stretch=1)
        return dock

    def _result_section(self) -> QFrame:
        card = QFrame()
        self.frame_result_card = card
        card.setObjectName("card_result")
        outer = QVBoxLayout(card)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        hdr = QFrame()
        hdr.setObjectName("result_hdr")
        hdr_lay = QVBoxLayout(hdr)
        hdr_lay.setContentsMargins(16, 12, 16, 10)
        hdr_lay.setSpacing(0)

        title_row = QHBoxLayout()
        title_row.setSpacing(12)
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self.lbl_result_card_title = QLabel()
        self.lbl_result_card_title.setObjectName("card_title")
        self.lbl_result_subtitle = QLabel()
        self.lbl_result_subtitle.setObjectName("result_hdr_sub")
        self.lbl_result_subtitle.setWordWrap(True)
        title_col.addWidget(self.lbl_result_card_title)
        title_col.addWidget(self.lbl_result_subtitle)
        title_row.addLayout(title_col, stretch=1)

        actions = QHBoxLayout()
        actions.setSpacing(6)
        self.btn_history = QPushButton()
        self.btn_history.setObjectName("toolbar_btn")
        self._configure_toolbar_button(self.btn_history, width=96)
        self.btn_history.clicked.connect(self._show_history_dialog)
        self.btn_open_draft = QPushButton()
        self.btn_open_draft.setObjectName("secondary")
        self.btn_open_draft.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_open_draft.setMinimumHeight(32)
        self.btn_open_draft.setMaximumHeight(32)
        self.btn_open_draft.hide()
        self.btn_open_draft.clicked.connect(self._open_draft_externally)
        self.btn_view_draft = QPushButton()
        self.btn_view_draft.setObjectName("secondary")
        self.btn_view_draft.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_view_draft.setMinimumHeight(32)
        self.btn_view_draft.setMaximumHeight(32)
        self.btn_view_draft.hide()
        self.btn_view_draft.clicked.connect(self._show_draft_table_dialog)
        self.btn_save = QPushButton()
        self.btn_save.setObjectName("primary")
        self.btn_save.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_save.setMinimumHeight(32)
        self.btn_save.setMaximumHeight(32)
        self.btn_save.setEnabled(False)
        self.btn_save.hide()
        self.btn_save.clicked.connect(self._save_csv)
        self.btn_menu = QPushButton()
        self.btn_menu.setObjectName("toolbar_btn")
        self._configure_toolbar_button(self.btn_menu, width=72)
        self.btn_menu.setEnabled(False)
        self.btn_menu.hide()
        self.btn_menu.clicked.connect(self._show_save_menu)
        actions.addWidget(self.btn_history)
        actions.addWidget(self.btn_open_draft)
        actions.addWidget(self.btn_view_draft)
        actions.addWidget(self.btn_save)
        actions.addWidget(self.btn_menu)
        title_row.addLayout(actions)
        hdr_lay.addLayout(title_row)

        self.lbl_report_title = QLabel()
        self.lbl_report_title.setObjectName("stats_title")
        self.lbl_report_title.hide()
        self.btn_view_report = QPushButton()
        self.btn_view_report.setObjectName("toolbar_btn")
        self._configure_toolbar_button(self.btn_view_report, width=100)
        self.btn_view_report.setEnabled(False)
        self.btn_view_report.hide()
        self.btn_view_report.clicked.connect(self._show_report_dialog)
        self.btn_export_folder = QPushButton()
        self.btn_export_folder.hide()
        self.btn_export_folder.clicked.connect(self._export_routing_folder)
        self.btn_copy_email = QPushButton()
        self.btn_copy_email.hide()
        self.btn_copy_email.clicked.connect(self._copy_department_email)
        self.btn_export_routing = QPushButton()
        self.btn_export_routing.hide()
        self.btn_export_routing.clicked.connect(self._save_routing)

        outer.addWidget(hdr)

        body = QFrame()
        body.setObjectName("result_body")
        body_lay = QVBoxLayout(body)
        body_lay.setContentsMargins(16, 16, 16, 12)
        body_lay.setSpacing(12)

        self.frame_result_empty = QFrame()
        self.frame_result_empty.setObjectName("result_empty_panel")
        empty_lay = QVBoxLayout(self.frame_result_empty)
        empty_lay.setContentsMargins(8, 12, 8, 12)
        empty_lay.addStretch(1)
        self.lbl_result_empty_title = QLabel()
        self.lbl_result_empty_title.setObjectName("result_empty_title")
        self.lbl_result_empty_title.setWordWrap(True)
        self.lbl_result_empty_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_result_empty = QLabel()
        self.lbl_result_empty.setObjectName("result_empty_hint")
        self.lbl_result_empty.setWordWrap(True)
        self.lbl_result_empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_lay.addWidget(self.lbl_result_empty_title)
        empty_lay.addWidget(self.lbl_result_empty)
        empty_lay.addStretch(2)
        body_lay.addWidget(self.frame_result_empty, stretch=1)

        self.frame_result_ready = QFrame()
        self.frame_result_ready.setObjectName("result_ready_panel")
        self.frame_result_ready.hide()
        ready_lay = QVBoxLayout(self.frame_result_ready)
        ready_lay.setContentsMargins(14, 12, 14, 12)
        ready_lay.setSpacing(6)
        self.lbl_result_ready_title = QLabel()
        self.lbl_result_ready_title.setObjectName("result_ready_title")
        self.lbl_result_ready_title.setWordWrap(True)
        self.lbl_result_ready_hint = QLabel()
        self.lbl_result_ready_hint.setObjectName("result_ready_hint")
        self.lbl_result_ready_hint.setWordWrap(True)
        ready_lay.addWidget(self.lbl_result_ready_title)
        ready_lay.addWidget(self.lbl_result_ready_hint)
        body_lay.addWidget(self.frame_result_ready)

        self.lbl_ai_banner = QLabel()
        self.lbl_ai_banner.setObjectName("result_alert")
        self.lbl_ai_banner.setWordWrap(True)
        self.lbl_ai_banner.hide()
        body_lay.addWidget(self.lbl_ai_banner)

        self.lbl_result_warning = QLabel()
        self.lbl_result_warning.setObjectName("result_warn")
        self.lbl_result_warning.setWordWrap(True)
        self.lbl_result_warning.hide()
        body_lay.addWidget(self.lbl_result_warning)

        self.frame_operation_log = QFrame()
        self.frame_operation_log.setObjectName("operation_log_frame")
        log_lay = QVBoxLayout(self.frame_operation_log)
        log_lay.setContentsMargins(0, 0, 0, 0)
        log_lay.setSpacing(6)
        log_hdr = QHBoxLayout()
        self.lbl_operation_log_title = QLabel()
        self.lbl_operation_log_title.setObjectName("operation_log_title")
        self.btn_open_operation_folder = QPushButton()
        self.btn_open_operation_folder.setObjectName("secondary")
        self.btn_open_operation_folder.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_open_operation_folder.clicked.connect(self._open_operation_folder)
        self.btn_open_operation_folder.hide()
        log_hdr.addWidget(self.lbl_operation_log_title)
        log_hdr.addStretch()
        log_hdr.addWidget(self.btn_open_operation_folder)
        log_lay.addLayout(log_hdr)
        self.text_operation_log = QTextEdit()
        self.text_operation_log.setObjectName("operation_log")
        self.text_operation_log.setReadOnly(True)
        self.text_operation_log.setMaximumHeight(96)
        self.text_operation_log.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        log_lay.addWidget(self.text_operation_log)
        self.frame_operation_log.hide()
        body_lay.addWidget(self.frame_operation_log)

        body_lay.addStretch()
        outer.addWidget(body, stretch=1)

        self._routing_warning_text_value = ""
        self._update_result_panel()
        return card

    def _routing_warning_text(self) -> str:
        return getattr(self, "_routing_warning_text_value", "")

    def _update_result_file_display(self) -> None:
        if not hasattr(self, "lbl_result_subtitle"):
            return
        if not self.input_file_path:
            self.lbl_result_subtitle.setText(self.tr("result_no_file"))
            return
        name = os.path.basename(self.input_file_path)
        has_draft = self._draft_dataframe() is not None
        key = "result_file_ready" if has_draft else "result_file_waiting"
        self.lbl_result_subtitle.setText(self.tr(key, name=name))

    def _update_result_panel(self) -> None:
        if not hasattr(self, "btn_view_draft"):
            return
        self._update_result_file_display()
        has_draft = self._draft_dataframe() is not None
        self.frame_result_empty.setVisible(not has_draft)
        self.frame_result_ready.setVisible(has_draft)
        ready = "true" if has_draft else "false"
        if self.frame_result_card.property("ready") != ready:
            self.frame_result_card.setProperty("ready", ready)
            self.frame_result_card.style().unpolish(self.frame_result_card)
            self.frame_result_card.style().polish(self.frame_result_card)
        for btn in (
            self.btn_save,
            self.btn_menu,
            self.btn_view_draft,
            self.btn_open_draft,
        ):
            btn.setVisible(has_draft)
        if has_draft:
            self.btn_save.setEnabled(True)
            self.btn_menu.setEnabled(True)
            self.lbl_ai_banner.show()
        else:
            self.btn_save.setEnabled(False)
            self.btn_menu.setEnabled(False)
            self.lbl_ai_banner.hide()
            self.lbl_result_warning.hide()
        self._refresh_workflow_ui()

    def _show_draft_table_dialog(self) -> None:
        if self._draft_dataframe() is None:
            self._show_warning("err_no_draft")
            return
        dlg = DraftTableDialog(self)
        dlg.exec()

    def _open_draft_externally(self) -> None:
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        path = os.path.join(
            tempfile.gettempdir(),
            f"smartsorter_draft_{os.getpid()}.xlsx",
        )
        try:
            save_dataframe_excel(df, path)
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))
            self.lbl_progress.setText(
                self.tr("progress_opened", name=os.path.basename(path)),
            )
        except Exception as e:
            self._show_error("err_excel", err=e)

    def _palette(self) -> dict[str, str]:
        return THEME_PALETTES.get(self.theme, THEME_PALETTES["light"])

    def _dropdown_menu_style(self) -> str:
        p = self._palette()
        return f"""
            QMenu {{
                background: {p["BG_CARD"]}; color: {p["TEXT"]};
                border: 1px solid {p["BORDER"]};
                border-radius: 8px; padding: 4px; margin: 0px;
            }}
            QMenu::item {{
                padding: 8px 14px; border-radius: 6px; font-size: 13px;
                background: transparent;
            }}
            QMenu::item:selected {{
                background: {p["BG_ACCENT"]}; color: {p["BLUE"]};
            }}
        """

    def _tpl_button_width(self) -> int:
        fm = QFontMetrics(self.btn_tpl.font())
        width = fm.horizontalAdvance(f"{self.tr('tpl_choose')}   ▼")
        for templates in TASK_TEMPLATES.values():
            for label, _ in templates:
                text = f"  {label}   ▼"
                width = max(width, fm.horizontalAdvance(text))
        return width + 28

    def _apply_tpl_button_width(self) -> None:
        self.btn_tpl.setFixedWidth(self._tpl_button_width())

    def _show_tpl_menu(self):
        menu = QMenu(self)
        _prepare_rounded_menu(menu)
        menu.setStyleSheet(self._dropdown_menu_style())
        menu.setFixedWidth(self.btn_tpl.width())
        templates = TASK_TEMPLATES.get(self.lang, TASK_TEMPLATES["en"])
        for i, (label, task_text) in enumerate(templates):
            menu.addAction(
                label,
                lambda _checked=False, idx=i, t=task_text: self._apply_template(idx, t),
            )
        menu.exec(self.btn_tpl.mapToGlobal(self.btn_tpl.rect().bottomLeft()))

    def _update_tpl_button(self) -> None:
        self._apply_tpl_button_width()
        templates = TASK_TEMPLATES.get(self.lang, TASK_TEMPLATES["en"])
        idx = self._selected_tpl_index
        if idx is not None and 0 <= idx < len(templates):
            label, _ = templates[idx]
            self.btn_tpl.setText(f"  {label}   ▼")
            self.btn_tpl.setToolTip(label)
            self.btn_tpl.setProperty("selected", "true")
            self.btn_tpl.setStyleSheet("")
        else:
            self._selected_tpl_index = None
            self.btn_tpl.setText(f"{self.tr('tpl_choose')}   ▼")
            self.btn_tpl.setToolTip(self.tr("tpl_choose"))
            self.btn_tpl.setProperty("selected", "false")
            self.btn_tpl.setStyleSheet("")

    def _rebuild_templates(self):
        templates = TASK_TEMPLATES.get(self.lang, TASK_TEMPLATES["en"])
        if self._selected_tpl_index is not None and 0 <= self._selected_tpl_index < len(templates):
            _, task_text = templates[self._selected_tpl_index]
            self.entry_task.blockSignals(True)
            self.entry_task.setPlainText(task_text)
            self.entry_task.blockSignals(False)
        self._update_tpl_button()

    def _on_theme_changed(self, _index: int):
        theme = self.combo_theme.currentData()
        if theme and theme != self.theme:
            self.combo_theme.hidePopup()
            self.combo_lang.hidePopup()
            self.theme = theme
            save_theme(theme)
            self._apply_theme()

    def _apply_theme(self):
        app = QApplication.instance()
        if not app:
            return
        apply_platform_theme(app, self.theme)
        app.setStyleSheet(build_stylesheet(self._palette()))
        if self._info_hover_card is not None:
            self._info_hover_card.apply_palette(self._palette())
        self._apply_task_placeholder_style()
        QTimer.singleShot(0, self._refresh_after_theme_change)

    def _refresh_after_theme_change(self) -> None:
        for combo in self.findChildren(HoverComboBox):
            combo.setProperty("hover", "false")
            combo.refresh_theme()
            if combo.underMouse():
                combo._set_hovered(True)

    def _rebuild_theme_combo(self):
        self.combo_theme.blockSignals(True)
        self.combo_theme.clear()
        self.combo_theme.addItem(self.tr("theme_light"), "light")
        self.combo_theme.addItem(self.tr("theme_dark"), "dark")
        idx = self.combo_theme.findData(self.theme)
        if idx >= 0:
            self.combo_theme.setCurrentIndex(idx)
        self.combo_theme.blockSignals(False)

    def _on_language_changed(self, _index: int):
        lang = self.combo_lang.currentData()
        if lang and lang != self.lang:
            self.lang = lang
            save_language(lang)
            self._apply_language()

    def _apply_language(self):
        idx = self.combo_lang.findData(self.lang)
        if idx >= 0:
            self.combo_lang.blockSignals(True)
            self.combo_lang.setCurrentIndex(idx)
            self.combo_lang.blockSignals(False)

        self.setWindowTitle(
            f"{self.tr('window_title')}  v{APP_VERSION}",
        )
        self.lbl_app_title.setText(self.tr("app_title"))
        self.lbl_app_sub.setText(self.tr("app_sub"))
        self.lbl_workspace_title.setText(self.tr("workspace_title"))
        self.lbl_workspace_sub.setText(self.tr("workspace_sub"))
        self.lbl_version_badge.setText(f"v{APP_VERSION}")
        self.lbl_ai_banner.setText(self.tr("ai_banner"))
        self.lbl_theme.setText(self.tr("theme_label"))
        self.lbl_lang.setText(self.tr("lang_label"))
        self._rebuild_theme_combo()
        self.lbl_card_api.setText(self.tr("card_api"))
        self.lbl_card_file.setText(self.tr("card_file"))
        self.lbl_card_task.setText(self.tr("card_task"))
        if self._wf_nav:
            self._wf_nav["setup"][2].setText(self.tr("wf_setup"))
            self._wf_nav["work"][2].setText(self.tr("wf_work"))
            self._wf_nav["result"][2].setText(self.tr("wf_result"))
        if hasattr(self, "lbl_wf_section"):
            self.lbl_wf_section.setText(self.tr("sidebar_workflow"))
        if hasattr(self, "lbl_help_section"):
            self.lbl_help_section.setText(self.tr("sidebar_help"))
        for key, title_key, _ in self.INFO_SECTIONS:
            btn = self._info_nav_buttons.get(key)
            if btn is not None:
                btn.setText(self.tr(title_key))
        self._update_workflow_nav()
        self._update_info_panel()
        self.lbl_tpl.setText(self.tr("tpl_label"))
        self.lbl_rules.setText(self.tr("rules_label"))
        self.btn_attach_rules.setText(self.tr("btn_attach_rules"))
        self.btn_clear_rules.setToolTip(self.tr("btn_detach"))
        self._update_rules_display()
        self.btn_clear_file.setToolTip(self.tr("btn_detach"))
        self.entry_task.setPlaceholderText(self.tr("task_ph"))
        self._apply_task_placeholder_style()
        self.btn_run.setText(self.tr("btn_run"))
        self.btn_run.setToolTip(self.tr("btn_run"))
        self.lbl_result_empty.setText(self.tr("result_empty_hint"))
        self.lbl_result_empty_title.setText(self.tr("result_empty_title"))
        self.btn_open.setText(self.tr("btn_open"))
        self.btn_change_file.setText(self.tr("btn_change_file"))
        if self.input_file_path:
            self._show_file_selected(self.input_file_path)
        elif self.btn_open.isVisible():
            self.lbl_file.setText(self.tr("no_file"))
        self.btn_history.setText(self.tr("btn_history"))
        self.btn_history.setToolTip(self.tr("btn_history_tip"))
        self.btn_save.setText(self.tr("btn_save"))
        self.btn_save.setToolTip(self.tr("btn_save_tip"))
        self.btn_export_folder.setText(self.tr("btn_export_folder"))
        self.btn_copy_email.setText(self.tr("btn_copy_email"))
        self.btn_copy_email.setToolTip(self.tr("btn_copy_email_tip"))
        self.btn_export_routing.setText(self.tr("btn_export_routing"))
        self.btn_menu.setText(f"{self.tr('btn_more')}  ▼")
        self.btn_menu.setToolTip(self.tr("menu_tip"))
        self.lbl_result_card_title.setText(self.tr("card_result"))
        self.lbl_result_ready_title.setText(self.tr("result_ready_title"))
        self.lbl_result_ready_hint.setText(self.tr("result_actions_hint"))
        self.btn_view_draft.setText(self.tr("btn_view_draft"))
        self.btn_view_draft.setToolTip(self.tr("btn_view_draft_tip"))
        self.btn_open_draft.setText(self.tr("btn_open_draft"))
        self.btn_open_draft.setToolTip(self.tr("btn_open_draft_tip"))
        self._update_result_file_display()
        self.lbl_report_title.setText(self.tr("report_title"))
        self.btn_view_report.setText(self.tr("btn_view_report"))
        self.lbl_operation_log_title.setText(self.tr("operation_log_title"))
        self.btn_open_operation_folder.setText(self.tr("btn_open_folder"))

        self.btn_eye.setText(self.tr("btn_hide") if not self._key_hidden else self.tr("btn_show"))

        if not self.input_file_path:
            self._reset_file_display()
        else:
            try:
                df = read_table(self.input_file_path)
                self._show_file_info(self.input_file_path, df)
            except Exception:
                self._reset_file_display()

        self._rebuild_templates()

        if self._last_changes is not None:
            self._show_changes(self._last_changes)

        self._update_result_panel()
        self._refresh_workflow_ui()

        key = load_api_key()
        if key and not self.entry_api.text():
            self.entry_api.setText(key)

    def _apply_task_placeholder_style(self) -> None:
        if not hasattr(self, "entry_task"):
            return
        p = self._palette()
        pal = self.entry_task.palette()
        pal.setColor(QPalette.ColorRole.PlaceholderText, QColor(p["TEXT_SEC"]))
        self.entry_task.setPalette(pal)

    def _on_task_changed(self) -> None:
        if self._selected_tpl_index is not None:
            templates = TASK_TEMPLATES.get(self.lang, TASK_TEMPLATES["en"])
            if self._selected_tpl_index < len(templates):
                _, expected = templates[self._selected_tpl_index]
                if self.entry_task.toPlainText() != expected:
                    self._selected_tpl_index = None
                    self._update_tpl_button()
        self._refresh_workflow_ui()

    def _match_template_index(self, lang: str, task_text: str) -> int | None:
        normalized = task_text.strip()
        if not normalized:
            return None
        templates = TASK_TEMPLATES.get(lang, TASK_TEMPLATES["en"])
        for i, (_, tpl_text) in enumerate(templates):
            if normalized == tpl_text.strip():
                return i
        return None

    def _restore_task_from_history(
        self, lang: str, task_text: str, template_index: int | None,
    ) -> None:
        if lang and lang != self.lang:
            lang_idx = self.combo_lang.findData(lang)
            if lang_idx >= 0:
                self.combo_lang.blockSignals(True)
                self.combo_lang.setCurrentIndex(lang_idx)
                self.combo_lang.blockSignals(False)
                self.lang = lang
                save_language(lang)
                self._apply_language()

        tpl_idx = template_index
        if tpl_idx is None:
            tpl_idx = self._match_template_index(self.lang, task_text)

        if tpl_idx is not None:
            templates = TASK_TEMPLATES.get(self.lang, TASK_TEMPLATES["en"])
            if 0 <= tpl_idx < len(templates):
                self._apply_template(tpl_idx, task_text)
                return

        self._selected_tpl_index = None
        self.entry_task.blockSignals(True)
        self.entry_task.setPlainText(task_text)
        self.entry_task.blockSignals(False)
        self._update_tpl_button()

    def _apply_template(self, index: int, task_text: str):
        self._selected_tpl_index = index
        self.entry_task.blockSignals(True)
        self.entry_task.setPlainText(task_text)
        self.entry_task.blockSignals(False)
        self._update_tpl_button()
        self._refresh_workflow_ui()
        self.entry_task.setFocus()

    def _is_complaints_workflow(self) -> bool:
        _, needs_rules = self._prepare_task(self.entry_task.toPlainText().strip())
        return needs_rules

    def _has_rules_attached(self) -> bool:
        path = load_rules_path()
        return bool(path and os.path.exists(path))

    def _routing_workflow_active(self) -> bool:
        return self._is_complaints_workflow() and self._has_rules_attached()

    def _setup_shortcuts(self) -> None:
        QShortcut(QKeySequence.StandardKey.Open, self, self._pick_file)
        QShortcut(QKeySequence("Ctrl+Return"), self, self._run)
        QShortcut(QKeySequence("Meta+Return"), self, self._run)

    def _has_api_key(self) -> bool:
        return bool(self.entry_api.text().strip() or load_api_key())

    def _run_readiness(self) -> tuple[bool, str]:
        if self._worker_busy:
            return False, "workflow_busy"
        if not self._has_api_key():
            return False, "workflow_need_key"
        if not self.input_file_path:
            return False, "workflow_need_file"
        task, needs_rules = self._prepare_task(self.entry_task.toPlainText().strip())
        if not task:
            return False, "workflow_need_task"
        if needs_rules and not self._has_rules_attached():
            return False, "workflow_need_rules"
        return True, "workflow_ready"

    def _on_api_key_changed(self, _text: str = "") -> None:
        self._refresh_workflow_ui()

    def _update_setup_status(self) -> None:
        if not hasattr(self, "lbl_api_status"):
            return
        if self._has_api_key():
            self.lbl_api_status.setText(self.tr("setup_key_ok"))
            self.lbl_api_status.setObjectName("status_ok")
        else:
            self.lbl_api_status.setText(self.tr("setup_key_pending"))
            self.lbl_api_status.setObjectName("status_pending")
        self.lbl_api_status.style().unpolish(self.lbl_api_status)
        self.lbl_api_status.style().polish(self.lbl_api_status)

        if self.input_file_path:
            self.lbl_file_meta.setText(
                self.tr(
                    "setup_file_ok",
                    rows=self._file_rows,
                    cols=self._file_cols,
                ),
            )
            self.lbl_file_meta.show()
        else:
            self.lbl_file_meta.setText(self.tr("setup_file_pending"))
            self.lbl_file_meta.show()

    def _update_input_lock(self) -> None:
        if not hasattr(self, "frame_task_composer"):
            return
        unlocked = bool(self.input_file_path) and not self._worker_busy
        self.entry_task.setReadOnly(not unlocked)
        self.btn_tpl.setEnabled(unlocked)
        locked = "false" if unlocked else "true"
        if self.frame_task_composer.property("locked") != locked:
            self.frame_task_composer.setProperty("locked", locked)
            self.frame_task_composer.style().unpolish(self.frame_task_composer)
            self.frame_task_composer.style().polish(self.frame_task_composer)

    def _set_run_progress_busy(self, busy: bool) -> None:
        if not hasattr(self, "progress_run"):
            return
        if busy:
            self.progress_run.setRange(0, 0)
            self.progress_run.show()
            self.lbl_progress.setText(self.tr("progress_run"))
        else:
            self.progress_run.hide()
            self.progress_run.setRange(0, 100)
            self.progress_run.setValue(0)

    def _update_run_state(self) -> None:
        if not hasattr(self, "btn_run"):
            return
        ready, hint_key = self._run_readiness()
        self.btn_run.setEnabled(ready)
        self._set_run_progress_busy(self._worker_busy)
        self.btn_run.setToolTip(
            self.tr("btn_run") if ready else self.tr(hint_key),
        )

    def _update_workspace_hint(self) -> None:
        if not hasattr(self, "lbl_workspace_sub"):
            return
        if self._draft_dataframe() is not None:
            self.lbl_workspace_sub.setText(self.tr("workflow_done"))
            return
        if self._worker_busy:
            self.lbl_workspace_sub.setText(self.tr("workflow_busy"))
            return
        _, hint_key = self._run_readiness()
        self.lbl_workspace_sub.setText(self.tr(hint_key))

    def _update_rules_visibility(self) -> None:
        if not hasattr(self, "frame_rules_block"):
            return
        show = self._is_complaints_workflow()
        self.frame_rules_block.setVisible(show)

    def _refresh_workflow_ui(self) -> None:
        self._update_workflow_nav()
        self._update_setup_status()
        self._update_input_lock()
        self._update_rules_visibility()
        self._update_run_state()
        self._update_workspace_hint()

    def _update_rules_display(self):
        path = load_rules_path()
        if path and os.path.exists(path):
            self.lbl_rules_file.setWordWrap(False)
            self.lbl_rules_file.setObjectName("file_chip_name")
            self.lbl_rules_file.setStyleSheet("")
            self._elide_label_text(
                self.lbl_rules_file,
                os.path.basename(path),
                reserve=40,
            )
            self.btn_clear_rules.show()
        else:
            self.lbl_rules_file.setWordWrap(True)
            self.lbl_rules_file.setText(self.tr("rules_none"))
            self.lbl_rules_file.setObjectName("rules_hint")
            self.lbl_rules_file.setStyleSheet("")
            self.btn_clear_rules.hide()
        self._refresh_workflow_ui()

    def _pick_rules_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("dlg_open_rules"), APP_DIR,
            ";;".join((
                self.tr("filter_rules"),
                self.tr("filter_rules_txt"),
                self.tr("filter_rules_table"),
                self.tr("filter_all"),
            )),
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        if ext not in RULES_EXTENSIONS:
            self._show_warning("err_rules_type")
            return
        save_rules_path(path)
        self._update_rules_display()

    def _clear_rules_file(self):
        save_rules_path("")
        self._update_rules_display()

    @staticmethod
    def _prepare_task(task: str) -> tuple[str, bool]:
        needs_rules = task.startswith(RULES_TASK_MARKER)
        clean = task.replace(RULES_TASK_MARKER, "", 1).strip()
        return clean, needs_rules

    def _save_key(self):
        key = self.entry_api.text().strip()
        if key:
            save_api_key(key)
        self._refresh_workflow_ui()

    def _toggle_key(self):
        self._key_hidden = not self._key_hidden
        if self._key_hidden:
            self.entry_api.setEchoMode(QLineEdit.EchoMode.Password)
            self.btn_eye.setText(self.tr("btn_show"))
        else:
            self.entry_api.setEchoMode(QLineEdit.EchoMode.Normal)
            self.btn_eye.setText(self.tr("btn_hide"))
        self.entry_api.setFocus()

    def _reset_file_display(self):
        self._show_file_empty_state()
        self._df_original = None
        self._clear_result()

    def _clear_result(self):
        self.result_csv = ""
        self._df_draft = None
        self._last_changes = None
        self._report_summary = ""
        self._report_details = ""
        self.lbl_report_title.hide()
        self.btn_view_report.hide()
        self.btn_view_report.setEnabled(False)
        self.btn_save.setEnabled(False)
        self.btn_menu.setEnabled(False)
        self.btn_export_routing.setEnabled(False)
        self.btn_export_folder.setEnabled(False)
        self.btn_copy_email.setEnabled(False)
        self._allowed_departments = []
        self._filter_mode = FILTER_ALL
        self._routing_warning_text_value = ""
        self._clear_operation_log()
        self._update_result_panel()
        self._refresh_workflow_ui()

    def _clear_operation_log(self) -> None:
        self._last_operation_path = ""
        self.text_operation_log.clear()
        self.btn_open_operation_folder.hide()
        self.frame_operation_log.hide()

    def _show_operation_log(self, text: str, folder_path: str = "") -> None:
        self._last_operation_path = folder_path
        self.text_operation_log.setPlainText(text)
        self.lbl_operation_log_title.setText(self.tr("operation_log_title"))
        self.btn_open_operation_folder.setText(self.tr("btn_open_folder"))
        if folder_path and os.path.isdir(folder_path):
            self.btn_open_operation_folder.show()
        else:
            self.btn_open_operation_folder.hide()
        self.frame_operation_log.show()

    def _open_operation_folder(self) -> None:
        path = self._last_operation_path
        if path and os.path.isdir(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def _routing_delivery_log_text(
        self,
        df: pd.DataFrame,
        deliver_lines: list[str],
        root: str,
    ) -> str:
        labels = self._routing_labels()
        allowed = self._allowed_departments or []
        known_df, unknown_df, review_df, dept_col = _routing_slices(df, allowed)
        stats = _routing_export_stats(df, dept_col, unknown_df, review_df)
        parts = [
            self.tr("operation_log_deliver"),
            self.tr("operation_log_path", path=root),
            "",
            build_routing_log(stats, labels),
            "",
            self.tr("operation_log_files"),
            *deliver_lines,
        ]
        return "\n".join(parts)

    def _routing_labels(self) -> dict[str, str]:
        return {
            "sheet_summary": self.tr("sheet_summary"),
            "sheet_review": self.tr("sheet_review"),
            "sheet_unknown_dept": self.tr("sheet_unknown_dept"),
            "dept": self.tr("sum_dept"),
            "total": self.tr("sum_total"),
            "high": self.tr("sum_high"),
            "review": self.tr("sum_review"),
            "unknown": self.tr("sum_unknown"),
            "email_body": self.tr("email_body"),
            "email_attach": self.tr("email_attach"),
            "log_title": self.tr("log_title"),
            "log_date": self.tr("log_date"),
            "log_total": self.tr("log_total"),
            "log_review": self.tr("log_review"),
            "log_unknown": self.tr("log_unknown"),
            "log_by_dept": self.tr("log_by_dept"),
            "log_unknown_list": self.tr("log_unknown_list"),
        }

    def _count_unknown_departments(self, df: pd.DataFrame) -> int:
        if not self._allowed_departments:
            return 0
        dept_col = find_routing_column(df, ROUTING_DEPT_NAMES)
        if not dept_col:
            return 0
        return int(_unknown_department_mask(df[dept_col], self._allowed_departments).sum())

    def _filter_combo_items(self) -> list[tuple[str, object]]:
        items: list[tuple[str, object]] = [(self.tr("filter_all"), FILTER_ALL)]
        if self._df_draft is None:
            return items
        prio_col = find_routing_column(self._df_draft, ROUTING_PRIO_NAMES)
        cat_col = find_routing_column(self._df_draft, ROUTING_CAT_NAMES)
        dept_col = find_routing_column(self._df_draft, ROUTING_DEPT_NAMES)
        if prio_col:
            items.append((self.tr("filter_high"), FILTER_HIGH))
        if cat_col:
            items.append((self.tr("filter_review"), FILTER_REVIEW))
        if dept_col and self._allowed_departments:
            items.append((self.tr("filter_unknown_dept"), FILTER_UNKNOWN))
        if dept_col:
            depts = sorted({
                str(v).strip()
                for v in self._df_draft[dept_col]
                if not pd.isna(v) and str(v).strip()
            })
            for dept in depts:
                items.append((dept, dept))
        return items

    def _set_routing_warning(self):
        self._routing_warning_text_value = ""
        if self._df_draft is None or not self._routing_workflow_active():
            if hasattr(self, "lbl_result_warning"):
                self.lbl_result_warning.hide()
            return
        unknown_n = self._count_unknown_departments(self._df_draft)
        if unknown_n > 0:
            self._routing_warning_text_value = self.tr("warn_unknown_dept", n=unknown_n)
            self.lbl_result_warning.setText(self._routing_warning_text_value)
            self.lbl_result_warning.show()
        else:
            self.lbl_result_warning.hide()

    def _draft_has_routing(self, df: pd.DataFrame | None) -> bool:
        return df is not None and find_routing_column(df, ROUTING_DEPT_NAMES) is not None

    def _format_changes(self, changes: TableChanges) -> tuple[str, str]:
        removed_count, added_count = _effective_row_counts(changes)
        summary = self.tr(
            "changes_summary",
            before=changes.rows_before,
            after=changes.rows_after,
            cells=len(changes.changed_cells),
            removed=removed_count,
            added=added_count,
        )

        lines: list[str] = []
        row_limit = 25
        cell_limit = 40

        if changes.structure_changed:
            lines.append(self.tr("changes_structure"))

        net_fewer = max(0, changes.rows_before - changes.rows_after)
        if net_fewer > 0 and not changes.removed_rows and removed_count > 0:
            lines.append(self.tr("changes_rows_fewer", n=net_fewer))

        if changes.removed_rows:
            lines.append(self.tr("changes_removed_header", n=len(changes.removed_rows)))
            for row, preview in changes.removed_rows[:row_limit]:
                lines.append(self.tr("changes_row_removed", row=row, preview=preview))
            extra = len(changes.removed_rows) - row_limit
            if extra > 0:
                lines.append(self.tr("changes_more", n=extra))

        if changes.added_rows:
            if lines:
                lines.append("")
            lines.append(self.tr("changes_added_header", n=len(changes.added_rows)))
            for row, preview in changes.added_rows[:row_limit]:
                lines.append(self.tr("changes_row_added", row=row, preview=preview))
            extra = len(changes.added_rows) - row_limit
            if extra > 0:
                lines.append(self.tr("changes_more", n=extra))

        if changes.changed_cells:
            if lines:
                lines.append("")
            lines.append(self.tr("changes_cells_header", n=len(changes.changed_cells)))
            for row, col, old, new in changes.changed_cells[:cell_limit]:
                lines.append(self.tr("changes_cell", row=row, col=col, old=old, new=new))
            extra = len(changes.changed_cells) - cell_limit
            if extra > 0:
                lines.append(self.tr("changes_more", n=extra))

        if not lines:
            return summary, self.tr("changes_none")
        return summary, "\n".join(lines)

    def _show_changes(self, changes: TableChanges):
        self._last_changes = changes
        summary, details = self._format_changes(changes)
        self._report_summary = summary
        self._report_details = details
        self.lbl_report_title.setText(self.tr("report_title"))
        self.lbl_report_title.show()
        self.btn_view_report.setText(self.tr("btn_view_report"))
        self.btn_view_report.show()
        self.btn_view_report.setEnabled(True)
        self._update_result_panel()

    def _show_report_dialog(self):
        if not self._report_details:
            return
        dlg = WorkReportDialog(
            self,
            title=self.tr("report_title"),
            summary=self._report_summary,
            details=self._report_details,
            close_text=self.tr("btn_close"),
            configure_edit=self._configure_text_edit,
        )
        dlg.exec()

    def _show_history_dialog(self) -> None:
        dlg = HistoryDialog(self, self._configure_text_edit)
        dlg.exec()

    def _show_file_from_history(self, source_name: str, source_path: str) -> None:
        if source_path and os.path.exists(source_path):
            self.input_file_path = source_path
            self._show_file_selected(source_path)
            return
        self.input_file_path = ""
        self.btn_open.hide()
        self.lbl_file.hide()
        self.frame_file_chip.show()
        self.btn_change_file.show()
        self._elide_label_text(self.lbl_file_name, source_name, reserve=40)

    def _apply_history_record(self, record: HistoryRecord) -> bool:
        try:
            original_df, draft_df = load_history_dataframes(record)
        except Exception:
            self._show_warning("history_err_load")
            return False

        meta = record.meta
        self._show_file_from_history(meta.source_name, meta.source_path)
        task_text = meta.task_text or meta.task_preview
        self._restore_task_from_history(
            meta.lang, task_text, meta.template_index,
        )

        self.result_csv = draft_df.to_csv(index=False)
        self._df_draft = draft_df
        self._df_original = original_df
        self._file_rows = len(original_df)
        self._file_cols = len(original_df.columns)
        self._allowed_departments = load_allowed_departments()
        self._filter_mode = FILTER_ALL

        if original_df is not None:
            self._show_changes(compare_tables(original_df, draft_df))
        else:
            self._last_changes = None
            self._report_summary = meta.report_summary
            self._report_details = meta.report_details
            if meta.report_summary:
                self.lbl_report_title.setText(self.tr("report_title"))
                self.lbl_report_title.show()
                self.btn_view_report.show()
                self.btn_view_report.setEnabled(True)

        self.btn_save.setEnabled(True)
        self.btn_menu.setEnabled(True)
        self._refresh_workflow_ui()
        self._set_routing_warning()
        self._update_result_panel()
        self.lbl_progress.setText(self.tr("history_opened"))
        return True

    def _persist_history_entry(self) -> None:
        if self._df_draft is None:
            return
        original = self._df_original
        if original is None and self.input_file_path:
            try:
                original = read_table(self.input_file_path)
                self._df_original = original
            except Exception:
                return
        if original is None:
            return
        cells = len(self._last_changes.changed_cells) if self._last_changes else 0
        try:
            save_history_entry(
                source_path=self.input_file_path,
                task=self.entry_task.toPlainText().strip(),
                lang=self.lang,
                original_df=original,
                draft_df=self._df_draft,
                report_summary=self._report_summary,
                report_details=self._report_details,
                cells_changed=cells,
                template_index=self._selected_tpl_index,
            )
        except Exception:
            pass

    def _show_file_info(self, path: str, df: pd.DataFrame):
        self._file_rows = len(df)
        self._file_cols = len(df.columns)
        self._show_file_selected(path)
        self.entry_task.setFocus()
        if not self._worker_busy:
            self.lbl_progress.setText("")

    def _clear_file(self):
        self.input_file_path = ""
        self._reset_file_display()

    def _pick_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("dlg_open"), "",
            ";;".join((
                self.tr("filter_tables"),
                self.tr("filter_xlsx"),
                self.tr("filter_csv"),
                self.tr("filter_txt"),
                self.tr("filter_all"),
            )),
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        if ext not in TABLE_EXTENSIONS:
            self._show_error("err_type")
            return
        try:
            df = read_table(path)
            if df.empty:
                self._show_warning("err_file_empty")
                return
            self.input_file_path = path
            self._df_original = df.copy()
            self._clear_result()
            self._show_file_info(path, df)
        except ValueError:
            self.input_file_path = ""
            self._show_file_empty_state(self.tr("err_read"))
            self._show_error("err_read")
        except Exception as e:
            self.input_file_path = ""
            self._show_file_empty_state(self.tr("err_read"))
            self._show_error("err_excel", err=e)

    def _run(self):
        api_key = self.entry_api.text().strip()
        raw_task = self.entry_task.toPlainText().strip()
        task, needs_rules = self._prepare_task(raw_task)
        if not api_key:
            self._show_warning("err_api")
            self.entry_api.setFocus()
            return
        if not self.input_file_path:
            self._show_warning("err_file")
            self.btn_open.setFocus()
            return
        if not task:
            self._show_warning("err_task")
            self.entry_task.setFocus()
            return
        company_rules = load_company_rules()
        if needs_rules and not company_rules:
            self._show_warning("err_no_rules")
            self.btn_attach_rules.setFocus()
            return

        if self._worker_busy or (
            self._worker is not None and self._worker.isRunning()
        ):
            return

        self._clear_result()
        self._worker_busy = True
        self._refresh_workflow_ui()
        self.btn_run.setEnabled(False)
        self._worker = Worker(
            api_key,
            self.input_file_path,
            task,
            company_rules=company_rules,
            parent=self,
        )
        self._worker.finished_ok.connect(self._on_worker_ok)
        self._worker.finished_err.connect(self._on_worker_err)
        self._worker.start()

    def _on_worker_ok(self, result: str) -> None:
        """Слот потока — откладываем тяжёлую обработку на цикл UI (стабильнее на macOS)."""
        self._worker_busy = False
        self._pending_worker_result = result
        QTimer.singleShot(0, self._apply_pending_worker_result)

    def _on_worker_err(self, msg: str) -> None:
        self._worker_busy = False
        self._pending_worker_error = msg
        QTimer.singleShot(0, self._apply_pending_worker_error)

    def _apply_pending_worker_result(self) -> None:
        self._apply_worker_result(self._pending_worker_result)

    def _apply_pending_worker_error(self) -> None:
        self._apply_worker_error(self._pending_worker_error)

    def _apply_worker_result(self, result: str) -> None:
        self._worker_busy = False
        self.result_csv = result
        try:
            self._df_draft = parse_csv_text(result)
            task_text = self.entry_task.toPlainText().strip()
            _, needs_rules = self._prepare_task(task_text)
            if needs_rules:
                self._df_draft = enrich_complaints_draft(self._df_draft)
                self.result_csv = self._df_draft.to_csv(index=False)

            self._allowed_departments = load_allowed_departments()
            self._filter_mode = FILTER_ALL

            if self._df_original is None and self.input_file_path:
                try:
                    self._df_original = read_table(self.input_file_path)
                except Exception:
                    self._df_original = None

            if self._df_original is not None:
                self._show_changes(
                    self._compare_draft_changes(self._df_original, self._df_draft, needs_rules),
                )

            self.btn_save.setEnabled(True)
            self.btn_menu.setEnabled(True)
            self._refresh_workflow_ui()
            self._set_routing_warning()
            self._persist_history_entry()
            self._update_result_panel()
            self.lbl_progress.setText(self.tr("progress_done"))
            self.btn_view_draft.setFocus()
        except Exception as exc:
            self._worker_busy = False
            self._df_draft = None
            self.result_csv = ""
            self._update_result_panel()
            self._refresh_workflow_ui()
            self.lbl_progress.setText("")
            self._show_api_error(str(exc))

    def _compare_draft_changes(
        self,
        original: pd.DataFrame,
        draft: pd.DataFrame,
        complaints_workflow: bool,
    ) -> TableChanges:
        if complaints_workflow and list(original.columns) != list(draft.columns):
            return TableChanges(
                structure_changed=True,
                rows_before=len(original),
                rows_after=len(draft),
                changed_cells=[],
                removed_rows=[],
                added_rows=[],
            )
        return compare_tables(original, draft)

    def _apply_worker_error(self, msg: str) -> None:
        self._worker_busy = False
        self.lbl_progress.setText("")
        self._refresh_workflow_ui()
        if msg.startswith(ERROR_CODE_PREFIX):
            self._show_error(msg[len(ERROR_CODE_PREFIX):])
        else:
            self._show_api_error(msg)

    def _show_save_menu(self):
        menu = QMenu(self)
        _prepare_rounded_menu(menu)
        menu.setStyleSheet(self._dropdown_menu_style())
        menu.addAction(self.tr("menu_xlsx"), self._save_excel)
        menu.addAction(self.tr("menu_csv"), self._save_csv)
        if self._routing_workflow_active():
            df = self._draft_dataframe()
            has_routing = df is not None and self._draft_has_routing(df)
            menu.addSeparator()
            menu.addAction(self.tr("menu_folder"), self._export_routing_folder)
            menu.addAction(self.tr("btn_copy_email"), self._copy_department_email)
            if has_routing:
                menu.addAction(self.tr("menu_routing"), self._save_routing)
            if has_routing and department_inboxes_configured():
                deliver = menu.addAction(
                    self.tr("menu_deliver_inboxes"),
                    self._deliver_to_department_inboxes,
                )
                deliver.setToolTip(self.tr("btn_deliver_inboxes_tip"))
        menu.exec(self.btn_menu.mapToGlobal(self.btn_menu.rect().bottomLeft()))

    def _draft_dataframe(self) -> pd.DataFrame | None:
        if self._df_draft is not None:
            return self._df_draft
        if not self.result_csv:
            return None
        try:
            return parse_csv_text(self.result_csv)
        except Exception:
            return None

    def _save_csv(self):
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("dlg_save_csv"), "draft.csv", self.tr("filter_csv"),
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        df.to_csv(path, index=False, encoding="utf-8")
        self.lbl_progress.setText(self.tr("progress_saved", name=os.path.basename(path)))

    def _save_excel(self):
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("dlg_save_xlsx"), "draft.xlsx", self.tr("filter_xlsx"),
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            save_dataframe_excel(df, path)
            self.lbl_progress.setText(self.tr("progress_saved", name=os.path.basename(path)))
        except Exception as e:
            self._show_error("err_excel", err=e)

    def _save_routing(self):
        if not self._routing_workflow_active():
            return
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        if not self._draft_has_routing(df):
            self._show_warning("err_no_dept")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("dlg_save_routing"), "routing.xlsx", self.tr("filter_xlsx"),
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            save_routing_export(
                df, path, self._routing_labels(), self._allowed_departments or None,
            )
            self.lbl_progress.setText(
                self.tr("progress_routing_saved", name=os.path.basename(path)),
            )
        except ValueError:
            self._show_warning("err_no_dept")
        except Exception as e:
            self._show_error("err_excel", err=e)

    def _deliver_to_department_inboxes(self):
        if not self._routing_workflow_active():
            return
        if not department_inboxes_configured():
            self._show_warning(
                "err_no_inboxes", path=department_inboxes_config_path(),
            )
            return
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        if not self._draft_has_routing(df):
            self._show_warning("err_no_dept")
            return

        labels = self._routing_labels()
        review_body = build_routing_review_text(
            df,
            labels,
            self._allowed_departments or None,
            self.tr("dlg_routing_review_intro"),
        )
        dlg = RoutingReviewDialog(
            self,
            title=self.tr("dlg_routing_review_title"),
            hint=self.tr("dlg_routing_review_hint"),
            body=review_body,
            confirm_text=self.tr("btn_confirm_deliver"),
            cancel_text=self.tr("btn_close"),
            configure_edit=self._configure_text_edit,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        try:
            log, root = deliver_to_department_inboxes(
                df,
                labels,
                self._allowed_departments or None,
            )
            delivered = sum(1 for line in log if line.startswith("✓"))
            self._df_draft = mark_complaints_sent(df)
            self.result_csv = self._df_draft.to_csv(index=False)
            self._update_result_panel()
            self.lbl_progress.setText(
                self.tr("progress_delivered", n=delivered)
                + " · "
                + self.tr("progress_status_sent", n=len(self._df_draft)),
            )
            self._show_operation_log(
                self._routing_delivery_log_text(self._df_draft, log, root),
                folder_path=root,
            )
        except ValueError as exc:
            if str(exc) == "no_inboxes":
                self._show_warning(
                    "err_no_inboxes", path=department_inboxes_config_path(),
                )
            elif str(exc) == "nothing_delivered":
                self._show_warning("err_nothing_delivered")
            else:
                self._show_warning("err_no_dept")
        except Exception as e:
            self._show_error("err_excel", err=e)

    def _export_routing_folder(self):
        if not self._routing_workflow_active():
            return
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        if not self._draft_has_routing(df):
            self._show_warning("err_no_dept")
            return
        parent = QFileDialog.getExistingDirectory(self, self.tr("dlg_pick_folder"))
        if not parent:
            return
        try:
            labels = self._routing_labels()
            folder, stats = export_routing_folder(
                df, parent, labels, self._allowed_departments or None,
            )
            self.lbl_progress.setText(
                self.tr("progress_folder_saved", name=os.path.basename(folder)),
            )
            log_body = "\n".join((
                self.tr("operation_log_folder"),
                self.tr("operation_log_path", path=folder),
                "",
                build_routing_log(stats, labels),
            ))
            self._show_operation_log(log_body, folder_path=folder)
        except ValueError:
            self._show_warning("err_no_dept")
        except Exception as e:
            self._show_error("err_excel", err=e)

    def _copy_department_email(self):
        if not self._routing_workflow_active():
            return
        df = self._draft_dataframe()
        if df is None:
            self._show_warning("err_no_draft")
            return
        if not self._draft_has_routing(df):
            self._show_warning("err_no_dept")
            return
        dept_col = find_routing_column(df, ROUTING_DEPT_NAMES)
        if not dept_col:
            return
        depts = sorted({
            str(v).strip()
            for v in df[dept_col]
            if not pd.isna(v) and str(v).strip()
        })
        if not depts:
            return
        dlg = DepartmentEmailDialog(
            self,
            title=self.tr("dlg_email_title"),
            hint=self.tr("dlg_email_hint"),
            departments=depts,
            df=df,
            labels=self._routing_labels(),
            copy_text=self.tr("btn_copy_clipboard"),
            cancel_text=self.tr("btn_close"),
            configure_edit=self._configure_text_edit,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        dept = dlg.selected_department()
        text = dlg.email_text()
        QApplication.clipboard().setText(text)
        self.lbl_progress.setText(self.tr("email_copied"))
        QMessageBox.information(
            self,
            self.tr("dlg_email_title"),
            self.tr("email_copied_detail", department=dept),
        )


def main():
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
